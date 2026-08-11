"""
app.py — Flask web dashboard (отдельный сервис).

Google OAuth авторизация, ролевой доступ:
  admin    — всё + /admin панель
  employee — всё (все проекты, WB-отчёты, "Сегодня")
  seller   — только свои проекты, без WB, без "Сегодня"
  pending  — ожидает одобрения
"""
import os
import json
import math
import logging
import functools
from datetime import date, datetime, timedelta
from decimal import Decimal

from flask import (
    Flask, jsonify, request, Response,
    session, render_template, redirect,
)
from werkzeug.middleware.proxy_fix import ProxyFix
from authlib.integrations.flask_client import OAuth

from config import DATABASE_URL
from db import get_conn

# ─── App setup ────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)
logging.basicConfig(level=logging.INFO)

_secret = os.environ.get("SECRET_KEY", "")
if not _secret:
    raise RuntimeError("SECRET_KEY env var is not set — refusing to start")
app.secret_key = _secret

app.config.update(
    PERMANENT_SESSION_LIFETIME = timedelta(hours=8),
    SESSION_COOKIE_HTTPONLY    = True,
    SESSION_COOKIE_SAMESITE    = "Lax",
    SESSION_COOKIE_SECURE      = os.environ.get("FLASK_ENV") != "development",
)

DASHBOARD_URL = os.environ.get(
    "DASHBOARD_URL", "https://web-production-1275f.up.railway.app"
).rstrip("/")
ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "")

# ─── Создаём таблицы дизайна при старте веб-процесса ─────────────────────────
def _ensure_design_tables():
    """Идемпотентная миграция — всё в одном соединении."""
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                # Таблица задач (project включён сразу)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS design_tasks (
                        id          SERIAL PRIMARY KEY,
                        title       TEXT NOT NULL,
                        description TEXT DEFAULT '',
                        priority    TEXT DEFAULT 'med',
                        assignee    TEXT,
                        due         DATE,
                        project     TEXT,
                        done        BOOLEAN DEFAULT FALSE,
                        done_at     TIMESTAMPTZ,
                        created_by  TEXT DEFAULT '',
                        created_at  TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS design_attachments (
                        id          SERIAL PRIMARY KEY,
                        task_id     INTEGER NOT NULL REFERENCES design_tasks(id) ON DELETE CASCADE,
                        attach_role TEXT NOT NULL DEFAULT 'brief',
                        attach_type TEXT NOT NULL DEFAULT 'link',
                        name        TEXT NOT NULL,
                        url         TEXT DEFAULT '',
                        file_data   BYTEA,
                        mime_type   TEXT DEFAULT '',
                        created_by  TEXT DEFAULT '',
                        created_at  TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
                # Добавляем колонки если таблицы уже существовали без них
                cur.execute("ALTER TABLE design_tasks ADD COLUMN IF NOT EXISTS project TEXT")
                cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS design_access BOOLEAN DEFAULT FALSE")
                cur.execute("ALTER TABLE design_tasks ADD COLUMN IF NOT EXISTS assignee_email TEXT DEFAULT ''")
                cur.execute("ALTER TABLE design_tasks ADD COLUMN IF NOT EXISTS created_by_email TEXT DEFAULT ''")
                # Автор задачи (email) для уведомлений
                cur.execute("ALTER TABLE tasks ADD COLUMN IF NOT EXISTS created_by_email TEXT DEFAULT ''")
                cur.execute("ALTER TABLE tasks ADD COLUMN IF NOT EXISTS assignee_email TEXT DEFAULT ''")
                # Таблица уведомлений
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS notifications (
                        id          SERIAL PRIMARY KEY,
                        user_email  TEXT NOT NULL,
                        type        TEXT NOT NULL,
                        title       TEXT NOT NULL,
                        message     TEXT NOT NULL DEFAULT '',
                        task_id     INTEGER,
                        read        BOOLEAN DEFAULT FALSE,
                        created_at  TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
                cur.execute("CREATE INDEX IF NOT EXISTS idx_notif_user_unread ON notifications(user_email, read) WHERE read = FALSE")
                # Комментарии, история и «сегодня» для design_tasks
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS design_comments (
                        id         SERIAL PRIMARY KEY,
                        task_id    INTEGER NOT NULL REFERENCES design_tasks(id) ON DELETE CASCADE,
                        author     TEXT NOT NULL DEFAULT '',
                        text       TEXT NOT NULL DEFAULT '',
                        created_at TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS design_task_history (
                        id         SERIAL PRIMARY KEY,
                        task_id    INTEGER NOT NULL REFERENCES design_tasks(id) ON DELETE CASCADE,
                        changed_by TEXT NOT NULL DEFAULT '',
                        field      TEXT NOT NULL DEFAULT '',
                        old_value  TEXT,
                        new_value  TEXT,
                        changed_at TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS design_today_claims (
                        id           SERIAL PRIMARY KEY,
                        task_id      INTEGER NOT NULL REFERENCES design_tasks(id) ON DELETE CASCADE,
                        user_name    TEXT NOT NULL,
                        claimed_date DATE NOT NULL DEFAULT CURRENT_DATE,
                        UNIQUE (task_id, claimed_date)
                    )
                """)
                cur.execute("ALTER TABLE notifications ADD COLUMN IF NOT EXISTS dismissed BOOLEAN DEFAULT FALSE")
                # Юнит-экономика
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS unit_economics (
                        id               SERIAL PRIMARY KEY,
                        brand            TEXT NOT NULL DEFAULT '',
                        wb_article       BIGINT,
                        seller_article   TEXT DEFAULT '',
                        cost_price       NUMERIC(12,2) DEFAULT 0,
                        logistics_to_wb  NUMERIC(12,2) DEFAULT 0,
                        packaging        NUMERIC(12,2) DEFAULT 0,
                        overhead         NUMERIC(12,2) DEFAULT 0,
                        defect_pct       NUMERIC(5,2) DEFAULT 0,
                        width_cm         NUMERIC(8,2),
                        length_cm        NUMERIC(8,2),
                        height_cm        NUMERIC(8,2),
                        liters           NUMERIC(8,3),
                        redemption_pct   NUMERIC(5,2) DEFAULT 100,
                        warehouse        TEXT DEFAULT '',
                        irp              NUMERIC(8,3) DEFAULT 1.0,
                        logistics_ktr    NUMERIC(12,2),
                        reception_coef   TEXT DEFAULT 'x0',
                        storage_per_day  NUMERIC(10,4) DEFAULT 0,
                        commission_pct   NUMERIC(5,2) DEFAULT 36,
                        wb_price         NUMERIC(12,2) DEFAULT 0,
                        drr_pct          NUMERIC(5,2) DEFAULT 0,
                        drr_external_rub NUMERIC(12,2) DEFAULT 0,
                        stock            INTEGER DEFAULT 0,
                        project          TEXT NOT NULL DEFAULT '',
                        created_at       TIMESTAMPTZ DEFAULT NOW(),
                        updated_at       TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
                cur.execute("ALTER TABLE unit_economics ADD COLUMN IF NOT EXISTS project TEXT NOT NULL DEFAULT ''")
                # quantity → stock (безопасный переименование)
                cur.execute("""
                    DO $$ BEGIN
                        IF EXISTS (
                            SELECT 1 FROM information_schema.columns
                            WHERE table_name='unit_economics' AND column_name='quantity'
                        ) THEN
                            ALTER TABLE unit_economics RENAME COLUMN quantity TO stock;
                        END IF;
                    END $$
                """)
                cur.execute("ALTER TABLE unit_economics ADD COLUMN IF NOT EXISTS stock INTEGER DEFAULT 0")
                cur.execute("ALTER TABLE unit_economics ADD COLUMN IF NOT EXISTS spp_pct NUMERIC(5,2) DEFAULT 0")
                cur.execute("ALTER TABLE unit_economics ADD COLUMN IF NOT EXISTS tax_system TEXT DEFAULT 'УСН 7%'")
                cur.execute("ALTER TABLE unit_economics ADD COLUMN IF NOT EXISTS tax_pct NUMERIC(5,2) DEFAULT 7")
                cur.execute("ALTER TABLE unit_economics ADD COLUMN IF NOT EXISTS il NUMERIC(6,4) DEFAULT 1")
                cur.execute("ALTER TABLE unit_economics ADD COLUMN IF NOT EXISTS status TEXT DEFAULT ''")
                # ФБС (свой склад) — параллельный расчёт
                cur.execute("ALTER TABLE unit_economics ADD COLUMN IF NOT EXISTS fbs_point TEXT DEFAULT 'ПВЗ'")
                cur.execute("ALTER TABLE unit_economics ADD COLUMN IF NOT EXISTS fbs_sc_coef NUMERIC(8,2) DEFAULT 100")
                cur.execute("ALTER TABLE unit_economics ADD COLUMN IF NOT EXISTS fbs_commission_pct NUMERIC(5,2) DEFAULT 0")
                cur.execute("ALTER TABLE unit_economics ADD COLUMN IF NOT EXISTS fbs_assembly_coef NUMERIC(6,2) DEFAULT 0")
                cur.execute("ALTER TABLE unit_economics ADD COLUMN IF NOT EXISTS fbs_own_cost NUMERIC(12,2) DEFAULT 0")
                cur.execute("ALTER TABLE unit_economics ADD COLUMN IF NOT EXISTS fbs_return_days NUMERIC(6,1) DEFAULT 0")
                cur.execute("ALTER TABLE unit_economics ADD COLUMN IF NOT EXISTS predmet TEXT DEFAULT ''")
                # ФБО: хранение по дням оборачиваемости + приёмка коробов
                cur.execute("ALTER TABLE unit_economics ADD COLUMN IF NOT EXISTS fbo_storage_days NUMERIC(6,1) DEFAULT 30")
                cur.execute("ALTER TABLE unit_economics ADD COLUMN IF NOT EXISTS fbo_accept_coef NUMERIC(6,2) DEFAULT 0")
                # ── Юнит Ozon (параллельная модель) ──────────────────────────
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS unit_ozon (
                        id             SERIAL PRIMARY KEY,
                        project        TEXT NOT NULL DEFAULT '',
                        seller_article TEXT DEFAULT '',
                        sku_ozon       TEXT DEFAULT '',
                        name           TEXT DEFAULT '',
                        category       TEXT DEFAULT '',
                        scheme         TEXT DEFAULT 'FBO',
                        cluster        TEXT DEFAULT '',
                        purchase       NUMERIC(12,2) DEFAULT 0,
                        delivery       NUMERIC(12,2) DEFAULT 0,
                        packaging      NUMERIC(12,2) DEFAULT 0,
                        defect_pct     NUMERIC(6,2) DEFAULT 0,
                        length_cm      NUMERIC(8,2),
                        width_cm       NUMERIC(8,2),
                        height_cm      NUMERIC(8,2),
                        weight_kg      NUMERIC(8,3),
                        price          NUMERIC(12,2) DEFAULT 0,
                        points_pct     NUMERIC(6,2) DEFAULT 0,
                        storage_type   TEXT DEFAULT 'Обычный',
                        free_days      NUMERIC(7,1) DEFAULT 365,
                        turnover_days  NUMERIC(7,1) DEFAULT 60,
                        drr_pct        NUMERIC(6,2) DEFAULT 0,
                        batch          INTEGER DEFAULT 0,
                        fact_payout    NUMERIC(12,2) DEFAULT 0,
                        created_at     TIMESTAMPTZ DEFAULT NOW(),
                        updated_at     TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
                cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS unit_access BOOLEAN DEFAULT FALSE")
                cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS unit_projects TEXT[] DEFAULT '{}'::TEXT[]")
                # Ежедневные бэкапы
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS unit_economics_backup (
                        id          SERIAL PRIMARY KEY,
                        backup_date DATE NOT NULL,
                        project     TEXT NOT NULL,
                        data        JSONB NOT NULL,
                        created_at  TIMESTAMPTZ DEFAULT NOW(),
                        UNIQUE (backup_date, project)
                    )
                """)
                # Логи активности пользователей
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS user_activity (
                        id          SERIAL PRIMARY KEY,
                        email       TEXT NOT NULL,
                        name        TEXT NOT NULL DEFAULT '',
                        date        DATE NOT NULL DEFAULT CURRENT_DATE,
                        first_seen  TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                        last_seen   TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                        page_views  INT NOT NULL DEFAULT 1,
                        UNIQUE (email, date)
                    )
                """)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS article_ff_stocks (
                        id           SERIAL PRIMARY KEY,
                        project      TEXT NOT NULL,
                        vendor_code  TEXT NOT NULL,
                        ff_stock     INTEGER NOT NULL DEFAULT 0,
                        expected_stock INTEGER NOT NULL DEFAULT 0,
                        updated_at   TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                        UNIQUE (project, vendor_code)
                    )
                """)
                # ── План-факт: таблицы, ранее создававшиеся вручную (риск потери
                #    данных, если БД пересоздать). Закрепляем в коде. ──────────────
                # Планы продаж (вводит пользователь; НЕ из WB API)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS sales_plans (
                        id               SERIAL PRIMARY KEY,
                        project          TEXT NOT NULL,
                        month            DATE NOT NULL,
                        vendor_code      TEXT NOT NULL,
                        orders_rub_plan  NUMERIC(14,2) NOT NULL DEFAULT 0,
                        orders_qty_plan  INTEGER NOT NULL DEFAULT 0,
                        sales_rub_plan   NUMERIC(14,2) NOT NULL DEFAULT 0,
                        sales_qty_plan   INTEGER NOT NULL DEFAULT 0,
                        updated_at       TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                        UNIQUE (project, month, vendor_code)
                    )
                """)
                # Постоянные ярлыки сезонности по артикулу
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS article_seasons (
                        id           SERIAL PRIMARY KEY,
                        project      TEXT NOT NULL,
                        vendor_code  TEXT NOT NULL,
                        season_name  TEXT NOT NULL DEFAULT '',
                        updated_at   TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                        UNIQUE (project, vendor_code)
                    )
                """)
                # Коэффициенты сезонности по месяцам
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS season_coefficients (
                        id           SERIAL PRIMARY KEY,
                        project      TEXT NOT NULL,
                        month        DATE NOT NULL,
                        season_name  TEXT NOT NULL,
                        coefficient  NUMERIC(7,2) NOT NULL DEFAULT 0,
                        updated_at   TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                        UNIQUE (project, month, season_name)
                    )
                """)
                # Кэш факта план-факта по месяцам (JSONB-снимок результата
                # collect_planfact_data). Позволяет отдавать страницу из БД мгновенно
                # и не дёргать WB API на каждый просмотр (у sales-funnel свой rate-limit).
                # План/ФФ/сезоны накладываются поверх при отдаче — они дёшевы и меняются
                # независимо от API.
                cur.execute("DROP TABLE IF EXISTS planfact_snapshots")
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS planfact_cache (
                        project      TEXT NOT NULL,
                        month        DATE NOT NULL,
                        data         JSONB NOT NULL,
                        fetched_at   TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                        is_complete  BOOLEAN NOT NULL DEFAULT FALSE,
                        PRIMARY KEY (project, month)
                    )
                """)
                # Бэкапы планов+фактов (раз в 7 дней, по кабинету)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS planfact_backups (
                        id           SERIAL PRIMARY KEY,
                        backup_date  DATE NOT NULL,
                        project      TEXT NOT NULL,
                        plans        JSONB,
                        facts        JSONB,
                        created_at   TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                        UNIQUE (backup_date, project)
                    )
                """)
                # Цели кабинета на месяц (план-факт), напр. план ДРР по заказам.
                # Одна строка на кабинет+месяц (общий, не по артикулам).
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS planfact_targets (
                        project          TEXT NOT NULL,
                        month            DATE NOT NULL,
                        drr_orders_plan  NUMERIC(6,2),
                        updated_at       TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                        PRIMARY KEY (project, month)
                    )
                """)
                # Фоновые задачи генерации WB-отчёта (отчёт долгий из-за rate-limit
                # WB — HTTP-запрос не держим открытым, чтобы не ловить edge-timeout 502)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS wb_report_jobs (
                        id          TEXT PRIMARY KEY,
                        cabinet     TEXT NOT NULL,
                        date_from   DATE,
                        date_to     DATE,
                        status      TEXT NOT NULL DEFAULT 'running',  -- running|done|error
                        html        TEXT,
                        error       TEXT,
                        created_at  TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                        updated_at  TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                """)
                # Подзадачи
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS subtasks (
                        id               SERIAL PRIMARY KEY,
                        task_id          INTEGER NOT NULL REFERENCES tasks(id) ON DELETE CASCADE,
                        title            TEXT NOT NULL,
                        assignee         TEXT,
                        assignee_email   TEXT DEFAULT '',
                        due              DATE,
                        done             BOOLEAN DEFAULT FALSE,
                        done_at          TIMESTAMPTZ,
                        position         INTEGER DEFAULT 0,
                        created_by_email TEXT DEFAULT '',
                        created_at       TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
                cur.execute("CREATE INDEX IF NOT EXISTS idx_subtasks_task ON subtasks(task_id)")
                # Проекты (кабинеты) — динамическое управление
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS projects (
                        id         SERIAL PRIMARY KEY,
                        name       TEXT NOT NULL UNIQUE,
                        emoji      TEXT NOT NULL DEFAULT '⚪',
                        color      TEXT NOT NULL DEFAULT '#888888',
                        sort_order INTEGER NOT NULL DEFAULT 0,
                        active     BOOLEAN NOT NULL DEFAULT TRUE,
                        created_at TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
                # Сеем данные из config.py если таблица пуста
                cur.execute("SELECT COUNT(*) AS cnt FROM projects")
                if cur.fetchone()["cnt"] == 0:
                    _seed = [
                        ("KSF",       "🔵", "#4a9eff", 0),
                        ("KLIFE",     "🟣", "#a78bfa", 1),
                        ("PA",        "🟢", "#4ade80", 2),
                        ("PIELE",     "🟡", "#fbbf24", 3),
                        ("Визенкова", "🟠", "#fb923c", 4),
                    ]
                    for _name, _emoji, _color, _order in _seed:
                        cur.execute(
                            "INSERT INTO projects (name, emoji, color, sort_order) "
                            "VALUES (%s,%s,%s,%s) ON CONFLICT (name) DO NOTHING",
                            (_name, _emoji, _color, _order)
                        )
            conn.commit()
        logging.info("Design tables ready.")
        # Миграция created_by_email — в отдельной транзакции, ошибки не критичны
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute("""
                        UPDATE tasks t
                        SET created_by_email = u.email
                        FROM users u,
                             (SELECT DISTINCT ON (task_id) task_id, changed_by
                              FROM task_history
                              ORDER BY task_id, id ASC) first_hist
                        WHERE t.id = first_hist.task_id
                          AND u.name = first_hist.changed_by
                          AND (t.created_by_email IS NULL OR t.created_by_email = '')
                    """)
                conn.commit()
        except Exception as me:
            logging.warning(f"_ensure_design_tables migration: {me}")
    except Exception as e:
        logging.error(f"_ensure_design_tables error: {e}")

# Вызовы миграций перенесены в _run_startup_migrations() (под advisory-lock),
# чтобы воркеры gunicorn не выполняли ALTER/UPDATE конкурентно и не копили
# подвисшие на локах соединения (это роняло всю БД по числу коннектов).

# ─── Создаём таблицы контентных воронок ──────────────────────────────────────
def _ensure_funnel_tables():
    """Идемпотентная миграция для раздела «Контент-воронки»."""
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                # Флаги доступа на users (как unit_access/unit_projects)
                cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS funnel_access BOOLEAN DEFAULT FALSE")
                cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS funnel_projects TEXT[] DEFAULT '{}'::TEXT[]")

                # Карточка воронки: одна на (project, wb_article, "версия")
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS content_funnels (
                        id              SERIAL PRIMARY KEY,
                        project         TEXT NOT NULL,
                        wb_article      BIGINT NOT NULL,
                        seller_article  TEXT DEFAULT '',
                        title           TEXT DEFAULT '',
                        status          TEXT NOT NULL DEFAULT 'draft',
                            -- draft | generated | edited | sent_to_design | in_design | done
                        own_data        JSONB,            -- снимок "своих данных" на момент генерации
                        ab_context      TEXT DEFAULT '',  -- текстовая выжимка прошлых А/Б, ушедшая в промпт
                        ai_model        TEXT DEFAULT '',
                        ai_raw_response JSONB,            -- сырой ответ Claude (для отладки)
                        design_task_id  INTEGER REFERENCES design_tasks(id) ON DELETE SET NULL,
                        created_by      TEXT DEFAULT '',
                        created_by_email TEXT DEFAULT '',
                        created_at      TIMESTAMPTZ DEFAULT NOW(),
                        updated_at      TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
                cur.execute("CREATE INDEX IF NOT EXISTS idx_content_funnels_project_article ON content_funnels(project, wb_article)")

                # Слайды — гибкая схема: {position, slide_type, pain_point, message, creative_notes}
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS funnel_slides (
                        id              SERIAL PRIMARY KEY,
                        funnel_id       INTEGER NOT NULL REFERENCES content_funnels(id) ON DELETE CASCADE,
                        position        INTEGER NOT NULL DEFAULT 0,
                        slide_type      TEXT NOT NULL DEFAULT '',
                            -- черновая таксономия: cover | pain | solution | feature |
                            -- social_proof | comparison | usage | cta (свободный текст пока)
                        pain_point      TEXT DEFAULT '',
                        message         TEXT DEFAULT '',
                        creative_notes  TEXT DEFAULT '',
                        created_at      TIMESTAMPTZ DEFAULT NOW(),
                        updated_at      TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
                cur.execute("CREATE INDEX IF NOT EXISTS idx_funnel_slides_funnel ON funnel_slides(funnel_id, position)")

                # Загруженные отчёты по конкурентам (HTML/PDF/изображение/текст)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS funnel_competitor_files (
                        id              SERIAL PRIMARY KEY,
                        funnel_id       INTEGER NOT NULL REFERENCES content_funnels(id) ON DELETE CASCADE,
                        name            TEXT NOT NULL,
                        mime_type       TEXT DEFAULT '',
                        file_data       BYTEA,
                        extracted_text  TEXT DEFAULT '',   -- текст для промпта Claude
                        created_by      TEXT DEFAULT '',
                        created_at      TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
                cur.execute("CREATE INDEX IF NOT EXISTS idx_funnel_competitor_files_funnel ON funnel_competitor_files(funnel_id)")

                # Результаты А/Б тестов — копятся со временем, привязаны к воронке (опционально)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS funnel_ab_results (
                        id              SERIAL PRIMARY KEY,
                        funnel_id       INTEGER REFERENCES content_funnels(id) ON DELETE SET NULL,
                        project         TEXT NOT NULL,
                        wb_article      BIGINT NOT NULL,
                        variant_name    TEXT NOT NULL DEFAULT '',
                        period_from     DATE,
                        period_to       DATE,
                        metrics_before  JSONB,   -- {ctr, cart_pct, orders, buyout_pct, ...}
                        metrics_after   JSONB,
                        winner          TEXT DEFAULT '',   -- 'before' | 'after' | 'tie' | свободный текст
                        summary         TEXT DEFAULT '',   -- вывод, идёт в промпт следующих генераций
                        created_by      TEXT DEFAULT '',
                        created_by_email TEXT DEFAULT '',
                        created_at      TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
                cur.execute("CREATE INDEX IF NOT EXISTS idx_funnel_ab_results_project_article ON funnel_ab_results(project, wb_article)")
            conn.commit()
        logging.info("Funnel tables ready.")
    except Exception as e:
        logging.error(f"_ensure_funnel_tables error: {e}")


def _ensure_season_tables():
    """Идемпотентная миграция для раздела «Темп сезона»."""
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                # Флаг доступа на users (как unit_access/funnel_access)
                cur.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS season_access BOOLEAN DEFAULT FALSE")

                # Пресет = группа товаров со своим окном сезона
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS season_presets (
                        id               SERIAL PRIMARY KEY,
                        project          TEXT NOT NULL,
                        name             TEXT NOT NULL DEFAULT '',
                        start_date       DATE,
                        end_date         DATE,
                        buyout_pct       NUMERIC(5,2) NOT NULL DEFAULT 30,  -- константа выкупа (конверсии добавим позже)
                        position         INTEGER NOT NULL DEFAULT 0,
                        created_by_email TEXT DEFAULT '',
                        created_at       TIMESTAMPTZ DEFAULT NOW(),
                        updated_at       TIMESTAMPTZ DEFAULT NOW()
                    )
                """)
                cur.execute("CREATE INDEX IF NOT EXISTS idx_season_presets_project ON season_presets(project, position)")

                # Артикулы, входящие в группу
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS season_preset_items (
                        id              SERIAL PRIMARY KEY,
                        preset_id       INTEGER NOT NULL REFERENCES season_presets(id) ON DELETE CASCADE,
                        wb_article      BIGINT,
                        seller_article  TEXT DEFAULT '',
                        name            TEXT DEFAULT '',
                        position        INTEGER NOT NULL DEFAULT 0
                    )
                """)
                cur.execute("CREATE INDEX IF NOT EXISTS idx_season_preset_items_preset ON season_preset_items(preset_id)")

                # Снапшот метрик из План-факта по кабинету (обновляется утренним кроном/кнопкой).
                # articles = список {nm_id, vendor_code, stock, buyout_pct, avg_price, pace, name}.
                # При обычном открытии данные берём отсюда (без обращения к WB) и мгновенно
                # пересчитываем декомпозицию поверх текущих групп.
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS season_cache (
                        project      TEXT PRIMARY KEY,
                        articles     JSONB NOT NULL DEFAULT '[]'::jsonb,
                        days_elapsed INTEGER NOT NULL DEFAULT 1,
                        pace_month   TEXT DEFAULT '',
                        fetched_at   TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                """)

                # Ручные переопределения %выкупа по артикулу (перекрывают значение из
                # План-факта). Живут отдельно от групп, чтобы не сбрасываться при смене
                # состава группы. NULL/удаление строки = вернуться к План-факту.
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS season_buyout_overrides (
                        project     TEXT NOT NULL,
                        wb_article  BIGINT NOT NULL,
                        buyout_pct  NUMERIC(5,2) NOT NULL,
                        updated_at  TIMESTAMPTZ DEFAULT NOW(),
                        PRIMARY KEY (project, wb_article)
                    )
                """)
            conn.commit()
        logging.info("Season tables ready.")
    except Exception as e:
        logging.error(f"_ensure_season_tables error: {e}")


def _ensure_cockpit_tables():
    """
    Таблицы овнерского кокпита (/cockpit, только admin).
      agency_kpi    — план-факт показателей агентства (воронка привлечение/продажи/
                      предоставление): план/факт в штуках + цена ₽ за единицу.
      cashflow      — денежные потоки (доход/расход) по датам/категориям/кабинетам.
      team_members  — команда: оклад + должность.
      team_cabinets — привязка сотрудника к кабинету с планом и % мотивации от факта.
    """
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS agency_kpi (
                        id          SERIAL PRIMARY KEY,
                        month       DATE NOT NULL,
                        grp         TEXT NOT NULL DEFAULT 'Привлечение',
                        metric      TEXT NOT NULL,
                        plan_qty    INTEGER NOT NULL DEFAULT 0,
                        fact_qty    INTEGER NOT NULL DEFAULT 0,
                        unit_price  NUMERIC(14,2) NOT NULL DEFAULT 0,
                        position    INTEGER NOT NULL DEFAULT 0,
                        updated_at  TIMESTAMPTZ NOT NULL DEFAULT NOW(),
                        UNIQUE (month, grp, metric)
                    )
                """)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS cashflow (
                        id          SERIAL PRIMARY KEY,
                        dt          DATE NOT NULL,
                        direction   TEXT NOT NULL DEFAULT 'income',
                        category    TEXT NOT NULL DEFAULT '',
                        project     TEXT NOT NULL DEFAULT '',
                        amount      NUMERIC(14,2) NOT NULL DEFAULT 0,
                        note        TEXT NOT NULL DEFAULT '',
                        created_at  TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                """)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS team_members (
                        id          SERIAL PRIMARY KEY,
                        name        TEXT NOT NULL,
                        email       TEXT NOT NULL DEFAULT '',
                        role_title  TEXT NOT NULL DEFAULT '',
                        salary_base NUMERIC(14,2) NOT NULL DEFAULT 0,
                        active      BOOLEAN NOT NULL DEFAULT TRUE,
                        position    INTEGER NOT NULL DEFAULT 0,
                        created_at  TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                """)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS team_cabinets (
                        id          SERIAL PRIMARY KEY,
                        member_id   INTEGER NOT NULL REFERENCES team_members(id) ON DELETE CASCADE,
                        cabinet     TEXT NOT NULL,
                        plan_rub    NUMERIC(14,2) NOT NULL DEFAULT 0,
                        bonus_pct   NUMERIC(6,2) NOT NULL DEFAULT 0,
                        UNIQUE (member_id, cabinet)
                    )
                """)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS clients (
                        id          SERIAL PRIMARY KEY,
                        name        TEXT NOT NULL,
                        cabinet     TEXT NOT NULL DEFAULT '',
                        service     TEXT NOT NULL DEFAULT 'Ведение',
                        mrr         NUMERIC(14,2) NOT NULL DEFAULT 0,
                        start_date  DATE,
                        active      BOOLEAN NOT NULL DEFAULT TRUE,
                        note        TEXT NOT NULL DEFAULT '',
                        position    INTEGER NOT NULL DEFAULT 0,
                        created_at  TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                """)
                # Еженедельный снимок финданных кокпита (защита от потери данных)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS cockpit_backups (
                        backup_date  DATE PRIMARY KEY,
                        payload      JSONB NOT NULL,
                        created_at   TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                """)
                # Ручные финансовые цели месяца (оборот/прибыль ₽); факт тянется из
                # операций. Цели по лидам/продажам считаются из воронки, тут не хранятся.
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS cockpit_goals (
                        month         DATE PRIMARY KEY,
                        revenue_goal  NUMERIC(14,2) NOT NULL DEFAULT 0,
                        profit_goal   NUMERIC(14,2) NOT NULL DEFAULT 0,
                        updated_at    TIMESTAMPTZ NOT NULL DEFAULT NOW()
                    )
                """)
                # РНП: ежедневные отметки выполнения показателя воронки (сколько
                # сделано в конкретный день) + недельный план по показателю.
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS agency_kpi_daily (
                        id       SERIAL PRIMARY KEY,
                        kpi_id   INTEGER NOT NULL REFERENCES agency_kpi(id) ON DELETE CASCADE,
                        day      DATE NOT NULL,
                        qty      NUMERIC(12,2) NOT NULL DEFAULT 0,
                        UNIQUE (kpi_id, day)
                    )
                """)
                cur.execute("""
                    CREATE TABLE IF NOT EXISTS agency_kpi_week (
                        id          SERIAL PRIMARY KEY,
                        kpi_id      INTEGER NOT NULL REFERENCES agency_kpi(id) ON DELETE CASCADE,
                        week_start  DATE NOT NULL,
                        plan_qty    NUMERIC(12,2) NOT NULL DEFAULT 0,
                        UNIQUE (kpi_id, week_start)
                    )
                """)
            conn.commit()
        logging.info("Cockpit tables ready.")
    except Exception as e:
        logging.error(f"_ensure_cockpit_tables error: {e}")


def _run_startup_migrations():
    """
    Выполняет все стартовые миграции под Postgres advisory-lock, чтобы разные воркеры
    gunicorn не запускали ALTER/UPDATE одновременно. Раньше два воркера конфликтовали
    (ALTER TABLE ждал AccessExclusiveLock, заблокированный тяжёлым UPDATE соседа), что
    копило подвисшие соединения и упирало БД в лимит коннектов — весь дашборд падал.
    Ключ фиксированный; второй воркер ждёт освобождения и затем прогоняет миграции как
    no-op (все CREATE/ALTER — IF NOT EXISTS).
    """
    LOCK_KEY = 918273645
    try:
        with get_conn() as lock_conn:
            with lock_conn.cursor() as cur:
                cur.execute("SELECT pg_advisory_lock(%s)", (LOCK_KEY,))
            lock_conn.commit()
            try:
                _ensure_design_tables()
                _ensure_funnel_tables()
                _ensure_season_tables()
                _ensure_cockpit_tables()
            finally:
                with lock_conn.cursor() as cur:
                    cur.execute("SELECT pg_advisory_unlock(%s)", (LOCK_KEY,))
                lock_conn.commit()
    except Exception as e:
        logging.error(f"_run_startup_migrations error: {e}")

_run_startup_migrations()

# ─── Google OAuth ─────────────────────────────────────────────────────────────
oauth = OAuth(app)
oauth.register(
    name="google",
    client_id=os.environ.get("GOOGLE_CLIENT_ID"),
    client_secret=os.environ.get("GOOGLE_CLIENT_SECRET"),
    server_metadata_url="https://accounts.google.com/.well-known/openid-configuration",
    client_kwargs={"scope": "openid email profile"},
)

# ─── WB кабинеты ──────────────────────────────────────────────────────────────
def _get_wb_cabinets() -> list[str]:
    env_list = os.environ.get("WB_CABINETS", "")
    if env_list.strip():
        return [c.strip() for c in env_list.split(",") if c.strip()]
    from wb_api import CABINET_ENV_SLUGS
    slug_to_name = {v: k for k, v in CABINET_ENV_SLUGS.items()}
    prefix = "WB_API_KEY_"
    cabinets = []
    for key in sorted(os.environ):
        if key.startswith(prefix) and os.environ[key]:
            slug = key[len(prefix):]
            cabinets.append(slug_to_name.get(slug, slug))
    return cabinets

WB_CABINETS = _get_wb_cabinets()

# ─── DB helpers ───────────────────────────────────────────────────────────────
def _make_session_dict(user: dict) -> dict:
    return {
        "id":                 user["id"],
        "email":              user["email"],
        "name":               user.get("name") or user["email"],
        "picture":            user.get("picture") or "",
        "role":               user["role"],
        "projects":           list(user.get("projects") or []),
        "planfact_projects":  list(user.get("planfact_projects") or []),
        "planfact_edit":      bool(user.get("planfact_edit")),
        "design_access":      bool(user.get("design_access")),
        "unit_access":        bool(user.get("unit_access")),
        "unit_projects":      list(user.get("unit_projects") or []),
        "funnel_access":      bool(user.get("funnel_access")),
        "funnel_projects":    list(user.get("funnel_projects") or []),
        "season_access":      bool(user.get("season_access")),
    }

def _all_projects() -> list:
    """Активные проекты из БД, сортированные по sort_order."""
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM projects WHERE active=TRUE ORDER BY sort_order, id")
            return [dict(r) for r in cur.fetchall()]

def _project_emoji_dict() -> dict:
    return {p["name"]: p["emoji"] for p in _all_projects()}

def _project_names() -> list:
    return [p["name"] for p in _all_projects()]

def _design_projects(u) -> list | None:
    """None = все проекты (admin/employee), [] = нет доступа, [list] = проекты селлера."""
    if u["role"] in ("admin", "employee"):
        return None
    if u["role"] == "seller" and u.get("design_access"):
        return list(u.get("projects") or [])
    return []

def _planfact_allowed(u) -> list | None:
    """None = доступ ко всем кабинетам, [] = нет доступа, [list] = разрешённые кабинеты."""
    if u["role"] in ("admin", "employee"):
        return None   # все кабинеты
    # seller — только явно назначенные кабинеты
    pf = list(u.get("planfact_projects") or [])
    return pf

def _unit_allowed(u) -> bool:
    """Admin и employee всегда имеют доступ к юниту; seller — только если unit_access=True."""
    if u["role"] in ("admin", "employee"):
        return True
    return bool(u.get("unit_access"))

def _unit_projects(u) -> list | None:
    """None = доступ ко всем кабинетам (admin/employee), list = только эти кабинеты."""
    if u["role"] in ("admin", "employee"):
        return None
    return list(u.get("unit_projects") or [])

def _check_unit_project(u, project: str) -> bool:
    """Проверяет, разрешён ли доступ пользователя к конкретному кабинету юнита."""
    allowed = _unit_projects(u)
    if allowed is None:
        return True  # admin/employee — все кабинеты
    return project in allowed

def _funnel_allowed(u) -> bool:
    """Admin и employee всегда имеют доступ к воронкам; seller — только если funnel_access=True."""
    if u["role"] in ("admin", "employee"):
        return True
    return bool(u.get("funnel_access"))

def _funnel_projects(u) -> list | None:
    """None = доступ ко всем кабинетам (admin/employee), list = только эти кабинеты."""
    if u["role"] in ("admin", "employee"):
        return None
    return list(u.get("funnel_projects") or [])

def _check_funnel_project(u, project: str) -> bool:
    """Проверяет, разрешён ли доступ пользователя к конкретному кабинету воронок."""
    allowed = _funnel_projects(u)
    if allowed is None:
        return True  # admin/employee — все кабинеты
    return project in allowed

def _season_allowed(u) -> bool:
    """Admin и employee всегда имеют доступ к темпу сезона; seller — только если season_access=True."""
    if u["role"] in ("admin", "employee"):
        return True
    return bool(u.get("season_access"))

def _planfact_can_edit(u) -> bool:
    """Admin всегда может редактировать; остальные — только если planfact_edit=True."""
    if u["role"] == "admin":
        return True
    return bool(u.get("planfact_edit"))

def get_user_by_id(user_id: int) -> dict | None:
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM users WHERE id=%s", (user_id,))
            row = cur.fetchone()
            return dict(row) if row else None

def get_or_create_user(email: str, name: str, picture: str) -> dict:
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM users WHERE email=%s", (email,))
            user = cur.fetchone()
            if user:
                cur.execute(
                    "UPDATE users SET name=%s, picture=%s, last_login=NOW() "
                    "WHERE email=%s RETURNING *",
                    (name, picture, email),
                )
                user = cur.fetchone()
            else:
                role = "admin" if (ADMIN_EMAIL and email == ADMIN_EMAIL) else "pending"
                cur.execute(
                    "INSERT INTO users (email, name, picture, role) "
                    "VALUES (%s,%s,%s,%s) RETURNING *",
                    (email, name, picture, role),
                )
                user = cur.fetchone()
        conn.commit()
    return dict(user)

def get_tasks(project: str | None = None, projects: list | None = None):
    """
    project  — фильтр по одному проекту
    projects — фильтр по списку (для sellers)
    оба None — все задачи
    """
    base = """
        SELECT t.*, COALESCE(c.cnt,0) AS comment_count,
               COALESCE(s.total,0) AS subtask_count,
               COALESCE(s.done_cnt,0) AS subtask_done_count
        FROM tasks t
        LEFT JOIN (SELECT task_id, COUNT(*) AS cnt FROM comments GROUP BY task_id) c
               ON t.id = c.task_id
        LEFT JOIN (SELECT task_id, COUNT(*) AS total, COUNT(*) FILTER (WHERE done) AS done_cnt
                   FROM subtasks GROUP BY task_id) s
               ON t.id = s.task_id
    """
    with get_conn() as conn:
        with conn.cursor() as cur:
            if project:
                cur.execute(
                    base + " WHERE t.project=%s ORDER BY t.created_at DESC",
                    (project,),
                )
            elif projects is not None:
                cur.execute(
                    base + " WHERE t.project = ANY(%s)"
                           " ORDER BY t.project NULLS LAST, t.created_at DESC",
                    (projects,),
                )
            else:
                cur.execute(
                    base + " ORDER BY t.project NULLS LAST, t.created_at DESC"
                )
            return cur.fetchall()

def serialize(task):
    d = dict(task)
    for key in ("created_at", "done_at"):
        if d.get(key):
            d[key] = d[key].isoformat()
    return d

def serialize_comment(c):
    d = dict(c)
    if d.get("created_at"):
        d["created_at"] = d["created_at"].isoformat()
    return d

def _anon_comments(rows, user):
    """Для селлера полностью обезличиваем автора комментариев.
    Селлер не должен видеть/знать, кто из команды писал (имя остаётся в БД)."""
    seller = bool(user) and user.get("role") == "seller"
    out = []
    for c in rows:
        d = serialize_comment(c)
        if seller:
            d["author"] = ""
        out.append(d)
    return out

# ─── Auth helpers ─────────────────────────────────────────────────────────────
def _session_user() -> dict | None:
    return session.get("user")

def require_auth(f):
    """Декоратор API: нужна сессия, роль не pending."""
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        u = _session_user()
        if not u:
            return jsonify({"error": "unauthorized"}), 401
        if u["role"] == "pending":
            return jsonify({"error": "pending"}), 403
        return f(*args, **kwargs)
    return wrapper

def require_admin(f):
    """Декоратор страниц (redirect): только admin."""
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        u = _session_user()
        if not u:
            return redirect("/auth/login")
        if u["role"] != "admin":
            return Response("Доступ запрещён", status=403)
        return f(*args, **kwargs)
    return wrapper

def require_admin_api(f):
    """Декоратор API: только admin."""
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        u = _session_user()
        if not u or u["role"] != "admin":
            return jsonify({"error": "forbidden"}), 403
        return f(*args, **kwargs)
    return wrapper

def require_unit_page(f):
    """Декоратор страницы /unit: redirect если нет доступа."""
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        u = _session_user()
        if not u:
            return redirect("/auth/login")
        if u["role"] == "pending":
            return redirect("/")
        if not _unit_allowed(u):
            return redirect("/")
        return f(*args, **kwargs)
    return wrapper

def require_unit_api(f):
    """Декоратор API /api/unit-*: 403 если нет доступа."""
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        u = _session_user()
        if not u:
            return jsonify({"error": "unauthorized"}), 401
        if not _unit_allowed(u):
            return jsonify({"error": "forbidden"}), 403
        return f(*args, **kwargs)
    return wrapper

def require_funnel_page(f):
    """Декоратор страницы /funnels: redirect если нет доступа."""
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        u = _session_user()
        if not u:
            return redirect("/auth/login")
        if u["role"] == "pending":
            return redirect("/")
        if not _funnel_allowed(u):
            return redirect("/")
        return f(*args, **kwargs)
    return wrapper

def require_funnel_api(f):
    """Декоратор API /api/funnels*: 403 если нет доступа."""
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        u = _session_user()
        if not u:
            return jsonify({"error": "unauthorized"}), 401
        if not _funnel_allowed(u):
            return jsonify({"error": "forbidden"}), 403
        return f(*args, **kwargs)
    return wrapper

def require_season_page(f):
    """Декоратор страницы /season: redirect если нет доступа."""
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        u = _session_user()
        if not u:
            return redirect("/auth/login")
        if u["role"] == "pending":
            return redirect("/")
        if not _season_allowed(u):
            return redirect("/")
        return f(*args, **kwargs)
    return wrapper

def require_season_api(f):
    """Декоратор API /api/season/*: 403 если нет доступа."""
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        u = _session_user()
        if not u:
            return jsonify({"error": "unauthorized"}), 401
        if not _season_allowed(u):
            return jsonify({"error": "forbidden"}), 403
        return f(*args, **kwargs)
    return wrapper

def api_error(f):
    """Декоратор: ловит Exception → 500 JSON. Убирает try/except boilerplate."""
    @functools.wraps(f)
    def wrapper(*args, **kwargs):
        try:
            return f(*args, **kwargs)
        except RuntimeError as e:
            return jsonify({"error": str(e)}), 400
        except Exception as e:
            logging.exception("%s failure", f.__name__)
            return jsonify({"error": str(e)}), 500
    return wrapper

# ─── Auth routes ──────────────────────────────────────────────────────────────
@app.route("/auth/login")
def auth_login():
    if _session_user():
        return redirect("/")
    redirect_uri = f"{DASHBOARD_URL}/auth/callback"
    return oauth.google.authorize_redirect(redirect_uri)

@app.route("/auth/callback")
def auth_callback():
    try:
        token = oauth.google.authorize_access_token()
    except Exception as e:
        # Битый/устаревший OAuth-state: перезаход, истёкшая сессия, боты, крауллеры,
        # префетч колбэка. Не отдаём 500 — начинаем логин заново (свежий state).
        logging.info("auth callback restart (%s): %s", type(e).__name__, e)
        return redirect("/auth/login")
    info = token.get("userinfo") or {}
    email = info.get("email", "")
    if not email:
        return Response("Не удалось получить email от Google", status=400)
    user = get_or_create_user(
        email=email,
        name=info.get("name", email),
        picture=info.get("picture", ""),
    )
    session.permanent = True
    session["user"] = _make_session_dict(user)
    return redirect("/")

@app.route("/auth/logout")
def auth_logout():
    session.clear()
    return redirect("/auth/login")

# ─── Main page ────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    u = _session_user()
    if not u:
        return redirect("/auth/login")
    # Обновляем из БД при каждой загрузке страницы — чтобы изменения роли вступали в силу
    fresh = get_user_by_id(u["id"])
    if not fresh:
        session.clear()
        return redirect("/auth/login")
    session["user"] = _make_session_dict(fresh)
    u = session["user"]

    if u["role"] == "pending":
        return render_template("pending.html", user=u)

    role = u["role"]
    all_db_projects = _all_projects()
    if role in ("admin", "employee"):
        visible = [{"name": p["name"], "emoji": p["emoji"], "color": p["color"]} for p in all_db_projects]
    else:  # seller
        visible = [{"name": p["name"], "emoji": p["emoji"], "color": p["color"]}
                   for p in all_db_projects if p["name"] in (u["projects"] or [])]

    return render_template(
        "dashboard.html",
        current_user=u,
        visible_projects=visible,
        all_projects_json=json.dumps([
            {"id": p["id"], "name": p["name"], "emoji": p["emoji"],
             "color": p["color"], "sort_order": p["sort_order"], "active": p["active"]}
            for p in all_db_projects
        ]),
        show_wb=(role != "seller"),
        projects_json=json.dumps([p["name"] for p in visible]),
        show_design=_design_auth(u),
        design_projects_json=json.dumps(list(u.get("projects") or []) if role == "seller" else [p["name"] for p in visible]),
    )

# ─── Admin panel ──────────────────────────────────────────────────────────────
@app.route("/admin")
@require_admin
def admin_panel():
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM users ORDER BY (role='pending') DESC, created_at ASC"
            )
            users = [dict(r) for r in cur.fetchall()]
    return render_template("admin.html", users=users, all_projects=_project_names(), wb_cabinets=WB_CABINETS)

@app.route("/api/admin/users/<int:user_id>", methods=["POST"])
@require_admin_api
def api_admin_update_user(user_id):
    d = request.json or {}
    role = d.get("role", "pending")
    projects = d.get("projects", [])
    planfact_projects = d.get("planfact_projects", [])
    planfact_edit = bool(d.get("planfact_edit", False))
    design_access = bool(d.get("design_access", False))
    unit_access    = bool(d.get("unit_access", False))
    unit_projects  = d.get("unit_projects", [])
    funnel_access  = bool(d.get("funnel_access", False))
    funnel_projects = d.get("funnel_projects", [])
    season_access  = bool(d.get("season_access", False))
    if role not in ("pending", "employee", "seller", "admin"):
        return jsonify({"error": "invalid role"}), 400
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE users SET role=%s, projects=%s, planfact_projects=%s, planfact_edit=%s, design_access=%s, unit_access=%s, unit_projects=%s, funnel_access=%s, funnel_projects=%s, season_access=%s WHERE id=%s RETURNING id",
                (role, projects, planfact_projects, planfact_edit, design_access, unit_access, unit_projects, funnel_access, funnel_projects, season_access, user_id),
            )
            if not cur.fetchone():
                return jsonify({"error": "not found"}), 404
        conn.commit()
    return jsonify({"ok": True})

@app.route("/api/admin/users/<int:user_id>", methods=["DELETE"])
@require_admin_api
def api_admin_delete_user(user_id):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM users WHERE id=%s", (user_id,))
        conn.commit()
    return jsonify({"ok": True})

# ─── «Беру сегодня» API ───────────────────────────────────────────────────────
@app.route("/api/today")
@require_auth
def api_today():
    """Возвращает все заявки на сегодня со сведениями о задачах."""
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT tc.id, tc.task_id, tc.user_name,
                           t.title, t.project, t.priority, t.done
                    FROM today_claims tc
                    JOIN tasks t ON t.id = tc.task_id
                    WHERE tc.claimed_date = CURRENT_DATE
                    ORDER BY tc.user_name, tc.id
                """)
                rows = cur.fetchall()
        claims = [dict(r) for r in rows]
        return jsonify(claims)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/today/claim", methods=["POST"])
@require_auth
def api_today_claim():
    """Сотрудник берёт задачу на сегодня."""
    u = _session_user()
    if u["role"] not in ("admin", "employee"):
        return jsonify({"error": "forbidden"}), 403
    d = request.json or {}
    task_id = d.get("task_id")
    if not task_id:
        return jsonify({"error": "task_id required"}), 400
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                # Проверяем, что задача существует
                cur.execute("SELECT id FROM tasks WHERE id=%s AND done=FALSE", (task_id,))
                if not cur.fetchone():
                    return jsonify({"error": "task not found or already done"}), 404
                # INSERT OR IGNORE — задача могла быть взята другим
                cur.execute("""
                    INSERT INTO today_claims (task_id, user_name, claimed_date)
                    VALUES (%s, %s, CURRENT_DATE)
                    ON CONFLICT (task_id, claimed_date) DO NOTHING
                    RETURNING id
                """, (task_id, u["name"]))
                row = cur.fetchone()
            conn.commit()
        if row:
            return jsonify({"ok": True})
        else:
            return jsonify({"ok": False, "error": "already_claimed"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/today/unclaim", methods=["POST"])
@require_auth
def api_today_unclaim():
    """Сотрудник снимает свою заявку."""
    u = _session_user()
    d = request.json or {}
    task_id = d.get("task_id")
    if not task_id:
        return jsonify({"error": "task_id required"}), 400
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    DELETE FROM today_claims
                    WHERE task_id=%s AND claimed_date=CURRENT_DATE AND user_name=%s
                """, (task_id, u["name"]))
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/today/remove", methods=["POST"])
@require_auth
def api_today_remove():
    """Убирает задачу из списка сегодня (не удаляет и не закрывает задачу)."""
    u = _session_user()
    d = request.json or {}
    claim_id = d.get("claim_id")
    if not claim_id:
        return jsonify({"error": "claim_id required"}), 400
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                if u["role"] == "admin":
                    cur.execute("DELETE FROM today_claims WHERE id=%s", (claim_id,))
                else:
                    cur.execute(
                        "DELETE FROM today_claims WHERE id=%s AND user_name=%s AND claimed_date=CURRENT_DATE",
                        (claim_id, u["name"])
                    )
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/today/<int:task_id>/done", methods=["POST"])
@require_auth
def api_today_done(task_id):
    """Закрывает задачу и в today_claims, и в основном списке."""
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE tasks SET done=TRUE, done_at=NOW() WHERE id=%s", (task_id,)
                )
                cur.execute(
                    "DELETE FROM today_claims WHERE task_id=%s AND claimed_date=CURRENT_DATE",
                    (task_id,)
                )
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ─── Notifications API ────────────────────────────────────────────────────────

def _first_name(full_name: str) -> str:
    """Возвращает первое слово из полного имени."""
    return (full_name or "").strip().split()[0] if (full_name or "").strip() else (full_name or "")

def _resolve_owner_email_design(cur, task_id: int, stored_email: str, created_by_name: str) -> str:
    """
    Email создателя design-задачи.
    Fallback: ищет по first_name через users.name (created_by хранит первое имя).
    """
    if stored_email:
        return stored_email
    if created_by_name:
        cur.execute(
            "SELECT email FROM users WHERE name LIKE %s LIMIT 1",
            (created_by_name + "%",)
        )
        row = cur.fetchone()
        if row:
            return row["email"]
    return ""

def _resolve_owner_email(cur, task_id: int, stored_email: str) -> str:
    """
    Возвращает email создателя задачи.
    Если stored_email пуст (старые задачи) — ищет первого автора в task_history,
    затем находит его email по имени в таблице users.
    """
    if stored_email:
        return stored_email
    cur.execute(
        "SELECT changed_by FROM task_history WHERE task_id=%s ORDER BY id ASC LIMIT 1",
        (task_id,)
    )
    row = cur.fetchone()
    if not row:
        return ""
    cur.execute("SELECT email FROM users WHERE name=%s LIMIT 1", (row["changed_by"],))
    u = cur.fetchone()
    return u["email"] if u else ""

def _push_notification(cur, user_email: str, notif_type: str, title: str, message: str = "", task_id: int | None = None):
    """Создаёт запись уведомления в БД (вызывать внутри уже открытого cursor)."""
    if not user_email:
        return
    cur.execute(
        "INSERT INTO notifications (user_email, type, title, message, task_id) "
        "VALUES (%s, %s, %s, %s, %s)",
        (user_email, notif_type, title, message, task_id),
    )


@app.route("/api/users")
@require_auth
def api_users():
    """Список активных пользователей: email + first_name для выпадающего списка исполнителей."""
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT email, name FROM users WHERE role != 'pending' ORDER BY name"
                )
                rows = cur.fetchall()
        return jsonify([
            {"email": r["email"], "first_name": _first_name(r["name"])}
            for r in rows
        ])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/notifications")
@require_auth
def api_notifications():
    """Возвращает непрочитанные уведомления текущего пользователя."""
    u = _session_user()
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id, type, title, message, task_id, created_at "
                    "FROM notifications WHERE user_email=%s AND read=FALSE "
                    "ORDER BY created_at ASC",
                    (u["email"],),
                )
                rows = cur.fetchall()
        return jsonify({"notifications": [
            {
                "id":         r["id"],
                "type":       r["type"],
                "title":      r["title"],
                "message":    r["message"],
                "task_id":    r["task_id"],
                "created_at": r["created_at"].isoformat() if r["created_at"] else None,
            }
            for r in rows
        ]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/notifications/all")
@require_admin_api
def api_notifications_all():
    """Debug: все уведомления + состояние задач (только для admin)."""
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id, user_email, type, title, message, task_id, read, created_at "
                    "FROM notifications ORDER BY created_at DESC LIMIT 50"
                )
                notifs = cur.fetchall()
                cur.execute(
                    "SELECT id, title, assignee, assignee_email, created_by_email "
                    "FROM design_tasks ORDER BY created_at DESC LIMIT 20"
                )
                tasks = cur.fetchall()
        return jsonify({
            "notifications": [dict(r) | {"created_at": r["created_at"].isoformat()} for r in notifs],
            "design_tasks_emails": [dict(r) for r in tasks],
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/notifications/read", methods=["POST"])
@require_auth
def api_notifications_read():
    """Помечает все уведомления текущего пользователя как прочитанные."""
    u = _session_user()
    ids = (request.json or {}).get("ids") or []
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                if ids:
                    cur.execute(
                        "UPDATE notifications SET read=TRUE WHERE user_email=%s AND id = ANY(%s)",
                        (u["email"], ids),
                    )
                else:
                    cur.execute(
                        "UPDATE notifications SET read=TRUE WHERE user_email=%s",
                        (u["email"],),
                    )
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/notifications/dismiss", methods=["POST"])
@require_auth
def api_notifications_dismiss():
    """Скрывает все уведомления из панели (dismissed=TRUE), из БД не удаляет."""
    u = _session_user()
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE notifications SET dismissed=TRUE WHERE user_email=%s",
                    (u["email"],)
                )
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/notifications/inbox")
@require_auth
def api_notifications_inbox():
    """Последние 30 уведомлений (прочитанные + непрочитанные) для bell-панели."""
    u = _session_user()
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id, type, title, message, task_id, read, created_at "
                    "FROM notifications WHERE user_email=%s AND dismissed=FALSE "
                    "ORDER BY created_at DESC LIMIT 30",
                    (u["email"],),
                )
                rows = cur.fetchall()
                unread = sum(1 for r in rows if not r["read"])
        return jsonify({"unread": unread, "notifications": [
            {
                "id":         r["id"],
                "type":       r["type"],
                "title":      r["title"],
                "message":    r["message"],
                "task_id":    r["task_id"],
                "read":       r["read"],
                "created_at": r["created_at"].isoformat() if r["created_at"] else None,
            }
            for r in rows
        ]})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ─── Activity API ─────────────────────────────────────────────────────────────

@app.route("/api/activity/ping", methods=["POST"])
@require_auth
def api_activity_ping():
    """Пинг активности — вызывается JS при загрузке страницы и каждую минуту."""
    u = _session_user()
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO user_activity (email, name, date, first_seen, last_seen, page_views)
                    VALUES (%s, %s, CURRENT_DATE, NOW(), NOW(), 1)
                    ON CONFLICT (email, date) DO UPDATE
                      SET last_seen  = NOW(),
                          page_views = user_activity.page_views + 1,
                          name       = EXCLUDED.name
                """, (u["email"], u.get("name", "")))
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/activity")
@require_admin_api
def api_admin_activity():
    """Логи активности пользователей — последние 30 дней."""
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT email, name, date,
                           first_seen, last_seen,
                           EXTRACT(EPOCH FROM (last_seen - first_seen))::INT AS duration_sec,
                           page_views
                    FROM user_activity
                    WHERE date >= CURRENT_DATE - INTERVAL '30 days'
                    ORDER BY date DESC, last_seen DESC
                """)
                rows = cur.fetchall()
        return jsonify([{
            "email":        r["email"],
            "name":         r["name"],
            "date":         r["date"].isoformat(),
            "first_seen":   r["first_seen"].isoformat() if r["first_seen"] else None,
            "last_seen":    r["last_seen"].isoformat() if r["last_seen"] else None,
            "duration_sec": r["duration_sec"] or 0,
            "page_views":   r["page_views"],
        } for r in rows])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/admin/activity")
@require_admin
def admin_activity():
    return render_template("activity.html")


def _parse_day(s, default):
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except Exception:
        return default

@app.route("/api/admin/completed-summary")
@require_admin_api
def api_admin_completed_summary():
    """Сводка выполненных задач: в какой день кто что закрыл (обе доски).
    Виден только админу."""
    try:
        today_msk = (datetime.utcnow() + timedelta(hours=3)).date()
        to_day    = _parse_day(request.args.get("to", ""),   today_msk)
        from_day  = _parse_day(request.args.get("from", ""), to_day - timedelta(days=29))

        items = []  # {day, who, title, project, board}
        with get_conn() as conn:
            with conn.cursor() as cur:
                # ── Основные задачи ──
                cur.execute("""
                    SELECT t.id, t.title, t.project, t.done_at,
                           (t.done_at + INTERVAL '3 hours')::date AS day,
                           COALESCE((
                               SELECT h.changed_by FROM task_history h
                               WHERE h.task_id = t.id AND h.field='Статус' AND h.new_value='Готово'
                               ORDER BY h.changed_at DESC LIMIT 1
                           ), t.assignee, '—') AS who
                    FROM tasks t
                    WHERE t.done = TRUE AND t.done_at IS NOT NULL
                      AND (t.done_at + INTERVAL '3 hours')::date BETWEEN %s AND %s
                """, (from_day, to_day))
                for r in cur.fetchall():
                    items.append({"day": r["day"].isoformat(), "who": r["who"] or "—",
                                  "title": r["title"], "project": r["project"] or "",
                                  "board": "main"})
                # ── Задачи дизайнеру ──
                cur.execute("""
                    SELECT d.id, d.title, d.project, d.done_at,
                           (d.done_at + INTERVAL '3 hours')::date AS day,
                           COALESCE((
                               SELECT h.changed_by FROM design_task_history h
                               WHERE h.task_id = d.id AND h.field='Статус' AND h.new_value='Готово'
                               ORDER BY h.changed_at DESC LIMIT 1
                           ), d.assignee, '—') AS who
                    FROM design_tasks d
                    WHERE d.done = TRUE AND d.done_at IS NOT NULL
                      AND (d.done_at + INTERVAL '3 hours')::date BETWEEN %s AND %s
                """, (from_day, to_day))
                for r in cur.fetchall():
                    items.append({"day": r["day"].isoformat(), "who": r["who"] or "—",
                                  "title": r["title"], "project": r["project"] or "",
                                  "board": "design"})

        # Группируем: день → человек. Ключ человека — по имени (первое слово),
        # чтобы объединить полное имя (основные) и имя (дизайнеру).
        def norm(who):
            who = (who or "—").strip()
            return who.split()[0] if who and who != "—" else "—"

        by_day = {}
        totals = {}
        for it in items:
            who = norm(it["who"])
            d = by_day.setdefault(it["day"], {})
            p = d.setdefault(who, [])
            p.append({"title": it["title"], "project": it["project"], "board": it["board"]})
            totals[who] = totals.get(who, 0) + 1

        days = []
        for day in sorted(by_day.keys(), reverse=True):
            people = [{"who": who, "count": len(tasks), "tasks": tasks}
                      for who, tasks in sorted(by_day[day].items(), key=lambda kv: -len(kv[1]))]
            days.append({"date": day, "total": sum(p["count"] for p in people), "people": people})

        totals_by_person = [{"who": w, "count": c}
                            for w, c in sorted(totals.items(), key=lambda kv: -kv[1])]

        return jsonify({
            "from": from_day.isoformat(),
            "to": to_day.isoformat(),
            "days": days,
            "totals_by_person": totals_by_person,
            "grand_total": len(items),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/admin/completed")
@require_admin
def admin_completed():
    return render_template("completed.html")

# ─── Управление проектами (кабинетами) ────────────────────────────────────────

@app.route("/api/admin/projects")
@require_auth
def api_admin_get_projects():
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM projects ORDER BY sort_order, id")
                return jsonify([dict(r) for r in cur.fetchall()])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/projects", methods=["POST"])
@require_admin_api
def api_admin_create_project():
    try:
        d = request.json or {}
        name  = (d.get("name") or "").strip()
        emoji = (d.get("emoji") or "⚪").strip()
        color = (d.get("color") or "#888888").strip()
        if not name:
            return jsonify({"error": "name required"}), 400
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT COALESCE(MAX(sort_order),0)+1 AS pos FROM projects"
                )
                pos = cur.fetchone()["pos"]
                cur.execute(
                    "INSERT INTO projects (name, emoji, color, sort_order) "
                    "VALUES (%s,%s,%s,%s) RETURNING *",
                    (name, emoji, color, pos)
                )
                row = dict(cur.fetchone())
            conn.commit()
        return jsonify(row)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/projects/<int:proj_id>", methods=["PUT"])
@require_admin_api
def api_admin_update_project(proj_id):
    try:
        d = request.json or {}
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM projects WHERE id=%s", (proj_id,))
                row = cur.fetchone()
                if not row:
                    return jsonify({"error": "not found"}), 404
                name       = d.get("name",       row["name"])
                emoji      = d.get("emoji",      row["emoji"])
                color      = d.get("color",      row["color"])
                sort_order = d.get("sort_order", row["sort_order"])
                active     = d.get("active",     row["active"])
                cur.execute(
                    "UPDATE projects SET name=%s, emoji=%s, color=%s, sort_order=%s, active=%s WHERE id=%s RETURNING *",
                    (name, emoji, color, sort_order, active, proj_id)
                )
                updated = dict(cur.fetchone())
            conn.commit()
        return jsonify(updated)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/admin/projects/<int:proj_id>", methods=["DELETE"])
@require_admin_api
def api_admin_delete_project(proj_id):
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT name FROM projects WHERE id=%s", (proj_id,))
                row = cur.fetchone()
                if not row:
                    return jsonify({"error": "not found"}), 404
                proj_name = row["name"]
                cur.execute(
                    "SELECT COUNT(*) AS cnt FROM tasks WHERE project=%s AND done=FALSE",
                    (proj_name,)
                )
                active_cnt = cur.fetchone()["cnt"]
                if active_cnt > 0:
                    # Деактивируем, не удаляем — есть активные задачи
                    cur.execute("UPDATE projects SET active=FALSE WHERE id=%s", (proj_id,))
                    conn.commit()
                    return jsonify({"ok": True, "deactivated": True, "active_tasks": active_cnt})
                cur.execute("DELETE FROM projects WHERE id=%s", (proj_id,))
            conn.commit()
        return jsonify({"ok": True, "deleted": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/cabinet-purge", methods=["POST"])
@require_admin_api
def api_admin_cabinet_purge():
    """ПОЛНОЕ удаление кабинета и ВСЕХ его данных из БД (необратимо).
    WB API-ключ остаётся в env Railway — его нужно удалить отдельно."""
    d = request.json or {}
    name = (d.get("name") or "").strip()
    confirm = (d.get("confirm") or "").strip()
    if not name:
        return jsonify({"error": "name required"}), 400
    # Защита от случайного удаления — клиент должен прислать точное имя кабинета
    if confirm != name:
        return jsonify({"error": "confirm must equal cabinet name"}), 400
    # Все таблицы с данными по кабинету (дочерние строки удалятся каскадом по FK)
    purge = [
        ("tasks",                "project"),
        ("design_tasks",         "project"),
        ("content_funnels",      "project"),
        ("unit_economics",       "project"),
        ("unit_economics_backup","project"),
        ("article_ff_stocks",    "project"),
        ("sales_plans",          "project"),
        ("season_coefficients",  "project"),
        ("article_seasons",      "project"),
        ("planfact_cache",       "project"),
        ("planfact_backups",     "project"),
    ]
    deleted = {}
    try:
        # Атомарно: либо удаляется всё, либо ничего (без частичной потери данных)
        with get_conn() as conn:
            with conn.cursor() as cur:
                for tbl, col in purge:
                    cur.execute(f"DELETE FROM {tbl} WHERE {col}=%s", (name,))
                    deleted[tbl] = cur.rowcount
                cur.execute("DELETE FROM projects WHERE name=%s", (name,))
                deleted["projects"] = cur.rowcount
            conn.commit()
        logging.warning("CABINET PURGED: %s by %s — %s", name, _session_user().get("email"), deleted)
        return jsonify({"ok": True, "deleted": deleted,
                        "note": "WB API-ключ в env Railway остался — удалите вручную при необходимости"})
    except Exception as e:
        logging.exception("cabinet purge failed")
        return jsonify({"error": str(e)}), 500


@app.route("/api/admin/planfact-export")
@require_admin_api
def api_admin_planfact_export():
    """Скачать бэкап планов+фактов кабинета отдельным Excel-файлом."""
    cabinet = request.args.get("cabinet", "").strip()
    if not cabinet:
        return jsonify({"error": "cabinet required"}), 400
    try:
        import openpyxl, io
        from openpyxl.styles import Font, PatternFill
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT month, vendor_code, orders_qty_plan, orders_rub_plan, "
                    "sales_qty_plan, sales_rub_plan FROM sales_plans "
                    "WHERE project=%s ORDER BY month, vendor_code",
                    (cabinet,)
                )
                plan_rows = cur.fetchall()
                cur.execute(
                    "SELECT month, data FROM planfact_cache WHERE project=%s ORDER BY month",
                    (cabinet,)
                )
                fact_blobs = cur.fetchall()
        wb = openpyxl.Workbook()
        hdr_font = Font(bold=True, color="FFFFFF")
        hdr_fill = PatternFill("solid", fgColor="1F3A5F")
        def _hdr(ws, cols):
            for i, h in enumerate(cols, 1):
                c = ws.cell(row=1, column=i, value=h)
                c.font = hdr_font; c.fill = hdr_fill
        # Лист «Планы»
        ws1 = wb.active; ws1.title = "Планы"
        _hdr(ws1, ["Месяц", "Артикул", "Заказы шт", "Заказы ₽", "Прод. шт", "Прод. ₽"])
        for r in plan_rows:
            ws1.append([r["month"].isoformat(), r["vendor_code"],
                        int(r["orders_qty_plan"] or 0), float(r["orders_rub_plan"] or 0),
                        int(r["sales_qty_plan"] or 0), float(r["sales_rub_plan"] or 0)])
        # Лист «Факты»
        ws2 = wb.create_sheet("Факты")
        _hdr(ws2, ["Месяц", "Артикул", "Название", "Заказы шт", "Заказы ₽", "Прод. шт", "Прод. ₽"])
        for fb in fact_blobs:
            month_iso = fb["month"].isoformat()
            for a in (fb["data"] or {}).get("articles", []):
                ws2.append([month_iso, a.get("vendor_code", ""), a.get("title", ""),
                            int(a.get("orders_qty_fact") or 0), float(a.get("orders_rub_fact") or 0),
                            int(a.get("sales_qty_fact") or 0), float(a.get("sales_rub_fact") or 0)])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        fname = f"planfact_{cabinet}_{date.today().isoformat()}.xlsx"
        return Response(buf.read(), headers={
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Disposition": f'attachment; filename="{fname}"',
        })
    except Exception as e:
        logging.exception("planfact export failed")
        return jsonify({"error": str(e)}), 500

# ─── Task API ─────────────────────────────────────────────────────────────────
@app.route("/api/tasks")
@require_auth
def api_tasks():
    u = _session_user()
    try:
        if u["role"] in ("admin", "employee"):
            tasks = get_tasks()
        else:  # seller
            tasks = get_tasks(projects=u["projects"] or [])
        return jsonify([serialize(t) for t in tasks])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/tasks", methods=["POST"])
@require_auth
def api_add():
    u = _session_user()
    try:
        d = request.json or {}
        project = d.get("project")
        # Seller может добавлять только в свои проекты
        if u["role"] == "seller" and project and project not in (u["projects"] or []):
            return jsonify({"error": "forbidden"}), 403
        creator_email  = u.get("email", "")
        assignee_email = d.get("assignee_email", "")
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO tasks (chat_id,title,priority,assignee,assignee_email,due,project,created_by_email) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                    (0, d["title"], d.get("priority", "med"),
                     d.get("assignee"), assignee_email,
                     d.get("due"), project, creator_email),
                )
                # Уведомление назначенному исполнителю
                if assignee_email and assignee_email != creator_email:
                    _push_notification(
                        cur, assignee_email, "task_assigned",
                        f"Новая задача: {d['title']}",
                        f"Назначена вам в проекте {project or '—'}",
                    )
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/tasks/<int:task_id>/done", methods=["POST"])
@require_auth
def api_done(task_id):
    try:
        u = _session_user()
        changed_by = (u.get("name") or u.get("email", "unknown")) if u else "unknown"
        with get_conn() as conn:
            with conn.cursor() as cur:
                # Проверяем наличие комментария — обязательное условие закрытия
                cur.execute(
                    "SELECT COUNT(*) AS cnt FROM comments WHERE task_id=%s", (task_id,)
                )
                row = cur.fetchone()
                if not row or int(row["cnt"]) == 0:
                    return jsonify({"error": "Необходимо добавить комментарий или ссылку перед закрытием задачи"}), 400
                cur.execute(
                    "UPDATE tasks SET done=TRUE,done_at=NOW() WHERE id=%s RETURNING title,created_by_email",
                    (task_id,)
                )
                task_row = cur.fetchone()
                cur.execute(
                    "INSERT INTO task_history (task_id,changed_by,field,old_value,new_value) "
                    "VALUES (%s,%s,%s,%s,%s)",
                    (task_id, changed_by, "Статус", "В работе", "Готово")
                )
                # Уведомление создателю задачи
                if task_row:
                    owner_email = _resolve_owner_email(cur, task_id, task_row["created_by_email"] or "")
                    u_email = (u.get("email") or "") if u else ""
                    if owner_email and owner_email != u_email:
                        _push_notification(
                            cur, owner_email, "status_changed",
                            f"Задача закрыта: {task_row['title']}",
                            f"Отметил(а): {changed_by}",
                            task_id,
                        )
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/tasks/<int:task_id>", methods=["PUT"])
@require_auth
def api_update(task_id):
    try:
        u = _session_user()
        changed_by = (u.get("name") or u.get("email", "unknown")) if u else "unknown"
        d = request.json or {}
        field_labels = {
            "title":    "Название",
            "priority": "Приоритет",
            "assignee": "Исполнитель",
            "due":      "Срок",
            "project":  "Проект",
        }
        u_email            = u.get("email", "") if u else ""
        new_assignee_email = d.get("assignee_email", "")
        with get_conn() as conn:
            with conn.cursor() as cur:
                # Читаем старые значения
                cur.execute(
                    "SELECT title,priority,assignee,assignee_email,due,project,created_by_email FROM tasks WHERE id=%s",
                    (task_id,)
                )
                row = cur.fetchone()
                if row:
                    old = {k: row[k] for k in ["title","priority","assignee","due","project"]}
                    new = {
                        "title":    d.get("title"),
                        "priority": d.get("priority", "med"),
                        "assignee": d.get("assignee"),
                        "due":      d.get("due"),
                        "project":  d.get("project"),
                    }
                    changed_fields = []
                    for field, label in field_labels.items():
                        if str(old.get(field) or "") != str(new.get(field) or ""):
                            cur.execute(
                                "INSERT INTO task_history "
                                "(task_id,changed_by,field,old_value,new_value) "
                                "VALUES (%s,%s,%s,%s,%s)",
                                (task_id, changed_by, label,
                                 old.get(field), new.get(field))
                            )
                            changed_fields.append(label)

                    notified = set()

                    # Уведомление новому исполнителю при смене
                    old_assignee_email = row.get("assignee_email") or ""
                    if new_assignee_email and new_assignee_email != old_assignee_email and new_assignee_email != u_email:
                        _push_notification(
                            cur, new_assignee_email, "task_assigned",
                            f"Задача назначена вам: {new.get('title') or old.get('title')}",
                            f"Назначил(а): {changed_by}",
                            task_id,
                        )
                        notified.add(new_assignee_email)

                    # Уведомление создателю при любом изменении (кроме смены исполнителя)
                    owner_email = _resolve_owner_email(cur, task_id, row.get("created_by_email") or "")
                    non_assignee_changes = [f for f in changed_fields if f != "Исполнитель"]
                    if non_assignee_changes and owner_email and owner_email != u_email and owner_email not in notified:
                        _push_notification(
                            cur, owner_email, "task_updated",
                            f"Задача изменена: {new.get('title') or old.get('title')}",
                            f"{changed_by} изменил(а): {', '.join(non_assignee_changes)}",
                            task_id,
                        )

                cur.execute(
                    "UPDATE tasks SET title=%s,priority=%s,assignee=%s,assignee_email=%s,due=%s,project=%s "
                    "WHERE id=%s",
                    (d.get("title"), d.get("priority", "med"),
                     d.get("assignee"), new_assignee_email,
                     d.get("due"), d.get("project"), task_id),
                )
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/tasks/<int:task_id>/history")
@require_auth
def api_task_history(task_id):
    u = _session_user()
    if not u or u["role"] != "admin":
        return jsonify({"error": "forbidden"}), 403
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id,changed_by,field,old_value,new_value,changed_at "
                    "FROM task_history WHERE task_id=%s ORDER BY changed_at DESC",
                    (task_id,)
                )
                rows = cur.fetchall()
                return jsonify([{
                    "id":         r["id"],
                    "changed_by": r["changed_by"],
                    "field":      r["field"],
                    "old_value":  r["old_value"],
                    "new_value":  r["new_value"],
                    "changed_at": r["changed_at"].isoformat() if r["changed_at"] else None,
                } for r in rows])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/tasks/<int:task_id>", methods=["DELETE"])
@require_auth
def api_delete(task_id):
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM tasks WHERE id=%s", (task_id,))
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/tasks/bulk", methods=["PUT"])
@require_auth
def api_bulk_update():
    """Массовое редактирование задач: изменить проект/приоритет/исполнителя."""
    try:
        u = _session_user()
        d = request.json or {}
        ids = [int(i) for i in (d.get("ids") or [])]
        if not ids:
            return jsonify({"error": "no ids"}), 400
        upd = d.get("update", {})
        allowed_fields = {"project", "priority", "assignee", "assignee_email"}
        fields = {k: v for k, v in upd.items() if k in allowed_fields}
        if not fields:
            return jsonify({"error": "no valid fields"}), 400
        set_parts = []
        vals = []
        for k, v in fields.items():
            set_parts.append(f"{k}=%s")
            vals.append(v or None)
        vals.append(ids)
        changed_by = (u.get("name") or u.get("email", "unknown")) if u else "unknown"
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    f"UPDATE tasks SET {', '.join(set_parts)} WHERE id=ANY(%s)",
                    vals
                )
                for field, new_val in fields.items():
                    for task_id in ids:
                        cur.execute(
                            "INSERT INTO task_history (task_id,changed_by,field,old_value,new_value) "
                            "VALUES (%s,%s,%s,%s,%s)",
                            (task_id, changed_by, field, None, str(new_val or ""))
                        )
            conn.commit()
        return jsonify({"ok": True, "updated": len(ids)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/tasks/bulk", methods=["DELETE"])
@require_auth
def api_bulk_delete():
    """Массовое удаление задач."""
    try:
        d = request.json or {}
        ids = [int(i) for i in (d.get("ids") or [])]
        if not ids:
            return jsonify({"error": "no ids"}), 400
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM tasks WHERE id=ANY(%s)", (ids,))
            conn.commit()
        return jsonify({"ok": True, "deleted": len(ids)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/tasks/bulk-done", methods=["POST"])
@require_auth
def api_bulk_done():
    """Массовое закрытие задач (автоматически добавляет системный комментарий)."""
    try:
        u = _session_user()
        changed_by = (u.get("name") or u.get("email", "unknown")) if u else "unknown"
        d = request.json or {}
        ids = [int(i) for i in (d.get("ids") or [])]
        comment = (d.get("comment") or "").strip() or f"Закрыто массовым действием ({changed_by})"
        if not ids:
            return jsonify({"error": "no ids"}), 400
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id FROM tasks WHERE id=ANY(%s) AND done=FALSE",
                    (ids,)
                )
                active_ids = [r["id"] for r in cur.fetchall()]
                for task_id in active_ids:
                    cur.execute(
                        "INSERT INTO comments (task_id, author, text) VALUES (%s,%s,%s)",
                        (task_id, changed_by, comment)
                    )
                    cur.execute(
                        "UPDATE tasks SET done=TRUE, done_at=NOW() WHERE id=%s",
                        (task_id,)
                    )
                    cur.execute(
                        "INSERT INTO task_history (task_id,changed_by,field,old_value,new_value) "
                        "VALUES (%s,%s,%s,%s,%s)",
                        (task_id, changed_by, "Статус", "В работе", "Готово")
                    )
            conn.commit()
        return jsonify({"ok": True, "closed": len(active_ids)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/tasks/<int:task_id>/comments")
@require_auth
def api_get_comments(task_id):
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT * FROM comments WHERE task_id=%s ORDER BY created_at ASC",
                    (task_id,),
                )
                return jsonify(_anon_comments(cur.fetchall(), _session_user()))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/tasks/<int:task_id>/comments", methods=["POST"])
@require_auth
def api_add_comment(task_id):
    try:
        u      = _session_user()
        u_email = u.get("email", "") if u else ""
        author  = _first_name(u.get("name", "")) or "Аноним"
        d = request.json or {}
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO comments (task_id, author, text) VALUES (%s,%s,%s)",
                    (task_id, author, d["text"]),
                )
                # Уведомляем создателя и исполнителя (кроме самого комментатора)
                cur.execute(
                    "SELECT title, created_by_email, assignee_email FROM tasks WHERE id=%s",
                    (task_id,)
                )
                task = cur.fetchone()
                if task:
                    notified = set()
                    for email in [task["created_by_email"], task["assignee_email"]]:
                        if email and email != u_email and email not in notified:
                            _push_notification(
                                cur, email, "comment_added",
                                f"Новый комментарий: {task['title']}",
                                f"{author}: {d['text'][:80]}",
                                task_id,
                            )
                            notified.add(email)
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ── Подзадачи ─────────────────────────────────────────────────────────────────

def serialize_subtask(s):
    d = dict(s)
    for key in ("created_at", "done_at"):
        if d.get(key):
            d[key] = d[key].isoformat()
    if d.get("due"):
        d["due"] = d["due"].isoformat()
    return d

def _recalc_parent_done(cur, task_id, changed_by="system"):
    """Пересчитывает статус родительской задачи: готова, только если ВСЕ подзадачи выполнены."""
    cur.execute(
        "SELECT COUNT(*) AS total, COUNT(*) FILTER (WHERE done) AS done_cnt FROM subtasks WHERE task_id=%s",
        (task_id,)
    )
    row = cur.fetchone()
    total, done_cnt = row["total"], row["done_cnt"]
    if total == 0:
        return
    all_done = (total == done_cnt)
    cur.execute("SELECT done, title, created_by_email FROM tasks WHERE id=%s", (task_id,))
    trow = cur.fetchone()
    if not trow:
        return
    if all_done and not trow["done"]:
        cur.execute("UPDATE tasks SET done=TRUE, done_at=NOW() WHERE id=%s", (task_id,))
        cur.execute(
            "INSERT INTO task_history (task_id,changed_by,field,old_value,new_value) "
            "VALUES (%s,%s,%s,%s,%s)",
            (task_id, changed_by, "Статус", "В работе", "Готово (все подзадачи выполнены)")
        )
        owner_email = _resolve_owner_email(cur, task_id, trow["created_by_email"] or "")
        if owner_email:
            _push_notification(
                cur, owner_email, "status_changed",
                f"Задача закрыта: {trow['title']}",
                "Все подзадачи выполнены — задача автоматически закрыта",
                task_id,
            )
    elif not all_done and trow["done"]:
        cur.execute("UPDATE tasks SET done=FALSE, done_at=NULL WHERE id=%s", (task_id,))
        cur.execute(
            "INSERT INTO task_history (task_id,changed_by,field,old_value,new_value) "
            "VALUES (%s,%s,%s,%s,%s)",
            (task_id, changed_by, "Статус", "Готово", "В работе (есть невыполненные подзадачи)")
        )

@app.route("/api/tasks/<int:task_id>/subtasks")
@require_auth
def api_get_subtasks(task_id):
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT * FROM subtasks WHERE task_id=%s ORDER BY position ASC, id ASC",
                    (task_id,)
                )
                return jsonify([serialize_subtask(s) for s in cur.fetchall()])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/tasks/<int:task_id>/subtasks", methods=["POST"])
@require_auth
def api_add_subtask(task_id):
    try:
        u = _session_user()
        creator_email = u.get("email", "") if u else ""
        d = request.json or {}
        title = (d.get("title") or "").strip()
        if not title:
            return jsonify({"error": "title required"}), 400
        assignee_email = d.get("assignee_email") or ""
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT COALESCE(MAX(position),0)+1 AS pos FROM subtasks WHERE task_id=%s", (task_id,))
                pos = cur.fetchone()["pos"]
                cur.execute(
                    "INSERT INTO subtasks (task_id,title,assignee,assignee_email,due,position,created_by_email) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING *",
                    (task_id, title, d.get("assignee"), assignee_email, d.get("due"), pos, creator_email),
                )
                sub = cur.fetchone()
                if assignee_email and assignee_email != creator_email:
                    cur.execute("SELECT title FROM tasks WHERE id=%s", (task_id,))
                    trow = cur.fetchone()
                    _push_notification(
                        cur, assignee_email, "task_assigned",
                        f"Новая подзадача: {title}",
                        f"В задаче «{trow['title'] if trow else task_id}»",
                        task_id,
                    )
            conn.commit()
        return jsonify(serialize_subtask(sub))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/subtasks/<int:subtask_id>", methods=["PUT"])
@require_auth
def api_update_subtask(subtask_id):
    try:
        d = request.json or {}
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM subtasks WHERE id=%s", (subtask_id,))
                cur_sub = cur.fetchone()
                if not cur_sub:
                    return jsonify({"error": "not found"}), 404
                title          = d.get("title", cur_sub["title"])
                assignee       = d.get("assignee", cur_sub["assignee"])
                assignee_email = d.get("assignee_email", cur_sub["assignee_email"])
                due            = d.get("due", cur_sub["due"])
                cur.execute(
                    "UPDATE subtasks SET title=%s, assignee=%s, assignee_email=%s, due=%s WHERE id=%s RETURNING *",
                    (title, assignee, assignee_email, due, subtask_id),
                )
                sub = cur.fetchone()
            conn.commit()
        return jsonify(serialize_subtask(sub))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/subtasks/<int:subtask_id>/toggle", methods=["POST"])
@require_auth
def api_toggle_subtask(subtask_id):
    try:
        u = _session_user()
        changed_by = (u.get("name") or u.get("email", "unknown")) if u else "unknown"
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT * FROM subtasks WHERE id=%s", (subtask_id,))
                sub = cur.fetchone()
                if not sub:
                    return jsonify({"error": "not found"}), 404
                new_done = not sub["done"]
                if new_done:
                    cur.execute("UPDATE subtasks SET done=TRUE, done_at=NOW() WHERE id=%s RETURNING *", (subtask_id,))
                else:
                    cur.execute("UPDATE subtasks SET done=FALSE, done_at=NULL WHERE id=%s RETURNING *", (subtask_id,))
                sub = cur.fetchone()
                _recalc_parent_done(cur, sub["task_id"], changed_by)
            conn.commit()
        return jsonify(serialize_subtask(sub))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/subtasks/<int:subtask_id>", methods=["DELETE"])
@require_auth
def api_delete_subtask(subtask_id):
    try:
        u = _session_user()
        changed_by = (u.get("name") or u.get("email", "unknown")) if u else "unknown"
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT task_id FROM subtasks WHERE id=%s", (subtask_id,))
                row = cur.fetchone()
                if not row:
                    return jsonify({"error": "not found"}), 404
                task_id = row["task_id"]
                cur.execute("DELETE FROM subtasks WHERE id=%s", (subtask_id,))
                _recalc_parent_done(cur, task_id, changed_by)
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ─── WB routes ────────────────────────────────────────────────────────────────
def _check_wb_access():
    """Проверяет авторизацию и роль для WB-отчётов. Возвращает None если OK."""
    u = _session_user()
    if not u:
        return redirect("/auth/login")
    if u["role"] in ("pending", "seller"):
        return Response("Доступ запрещён", status=403)
    return None

def _parse_period(args) -> tuple[date, date]:
    """Парсит ?from=YYYY-MM-DD&to=YYYY-MM-DD. По умолчанию — прошлая Пн–Вс."""
    today = date.today()
    monday = today - timedelta(days=today.weekday() + 7)
    default_from = monday
    default_to   = monday + timedelta(days=6)
    try:
        f = datetime.strptime(args.get("from", ""), "%Y-%m-%d").date()
    except ValueError:
        f = default_from
    try:
        t = datetime.strptime(args.get("to", ""), "%Y-%m-%d").date()
    except ValueError:
        t = default_to
    return f, t

@app.route("/wb")
def wb_index():
    err = _check_wb_access()
    if err:
        return err
    today = date.today()
    monday = today - timedelta(days=today.weekday() + 7)
    sunday = monday + timedelta(days=6)
    from wb_api import cabinet_env_slug
    from urllib.parse import quote
    cab_options, cab_cards = [], []
    for cab in WB_CABINETS:
        env_name = f"WB_API_KEY_{cabinet_env_slug(cab)}"
        configured = bool(os.environ.get(env_name) or os.environ.get("WB_API_KEY"))
        status = '<div class="status ok">✓ ключ настроен</div>' if configured else \
                 f'<div class="status">⚠ нет {env_name}</div>'
        cab_url = quote(cab)
        href = f"/wb/loading?cabinet={cab_url}&from={monday}&to={sunday}"
        link = f'<a href="{href}">Сформировать</a>' if configured else \
               '<a class="dis">недоступно</a>'
        cab_cards.append(f'<div class="cab"><h3>{cab}</h3>{status}{link}</div>')
        cab_options.append(
            f'<option value="{cab}" {"disabled" if not configured else ""}>{cab}</option>'
        )
    return render_template(
        "wb_index.html",
        cabinet_options="".join(cab_options),
        default_from=str(monday),
        default_to=str(sunday),
        cab_cards="".join(cab_cards),
    )

@app.route("/wb/loading")
def wb_loading():
    err = _check_wb_access()
    if err:
        return err
    cabinet = request.args.get("cabinet", "").strip()
    if not cabinet:
        return Response("?cabinet= обязателен", status=400)
    d_from, d_to = _parse_period(request.args)
    start_url = f"/wb/report/start?cabinet={cabinet}&from={d_from}&to={d_to}"
    months = ["января","февраля","марта","апреля","мая","июня",
              "июля","августа","сентября","октября","ноября","декабря"]
    if d_from.month == d_to.month:
        period = f"{d_from.day:02d} — {d_to.day:02d} {months[d_to.month-1]} {d_to.year}"
    else:
        period = (f"{d_from.day:02d} {months[d_from.month-1]} — "
                  f"{d_to.day:02d} {months[d_to.month-1]} {d_to.year}")
    return render_template(
        "wb_loading.html", cabinet=cabinet, period=period, start_url=start_url
    )

def _build_wb_report_html(cabinet, d_from, d_to):
    """Собирает полный HTML WB-отчёта (WB API + задачи + юнит). Чистая функция —
    можно вызывать как из HTTP-запроса, так и из фонового потока."""
    from wb_api import collect_report_data
    from wb_report import render_html
    data = collect_report_data(cabinet, d_from, d_to)
    # Выполненные задачи проекта за период отчёта
    done_tasks = []
    try:
        import datetime as _dt2
        d_to_dt = _dt2.datetime.combine(d_to, _dt2.time.max)
        d_from_dt = _dt2.datetime.combine(d_from, _dt2.time.min)
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT title, assignee, done_at FROM tasks "
                    "WHERE project = %s AND done = TRUE "
                    "  AND done_at >= %s AND done_at <= %s "
                    "ORDER BY done_at DESC",
                    (cabinet, d_from_dt, d_to_dt)
                )
                done_tasks = [
                    {"title": r["title"], "assignee": r["assignee"],
                     "done_at": r["done_at"].strftime("%d.%m") if r["done_at"] else ""}
                    for r in cur.fetchall()
                ]
    except Exception:
        done_tasks = []
    unit_data = []
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT wb_article, seller_article, brand, cost_price, "
                    "logistics_to_wb, packaging, overhead, defect_pct, liters, "
                    "redemption_pct, warehouse, irp, il, logistics_ktr, "
                    "reception_coef, storage_per_day, commission_pct, "
                    "wb_price, drr_pct, drr_external_rub, stock, spp_pct, "
                    "tax_system, tax_pct, status "
                    "FROM unit_economics WHERE project=%s ORDER BY id ASC",
                    (cabinet,)
                )
                unit_data = [dict(r) for r in cur.fetchall()]
        for row in unit_data:
            for k, v in row.items():
                if hasattr(v, '__float__') and not isinstance(v, (int, bool)):
                    row[k] = float(v) if v is not None else None
    except Exception:
        unit_data = []
    return render_html(data, done_tasks=done_tasks, unit_data=unit_data)


def _run_wb_report_job(job_id, cabinet, d_from, d_to):
    """Фоновая генерация отчёта — пишет результат/ошибку в wb_report_jobs."""
    try:
        html_out = _build_wb_report_html(cabinet, d_from, d_to)
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE wb_report_jobs SET status='done', html=%s, updated_at=NOW() WHERE id=%s",
                    (html_out, job_id)
                )
            conn.commit()
    except Exception as e:
        logging.exception("wb report job %s failed", job_id)
        try:
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "UPDATE wb_report_jobs SET status='error', error=%s, updated_at=NOW() WHERE id=%s",
                        (str(e)[:2000], job_id)
                    )
                conn.commit()
        except Exception:
            logging.exception("wb report job %s: failed to record error", job_id)


@app.route("/wb/report/start")
def wb_report_start():
    """Запускает фоновую генерацию отчёта, сразу возвращает job_id."""
    err = _check_wb_access()
    if err:
        return err
    cabinet = request.args.get("cabinet", "").strip()
    if not cabinet:
        return jsonify({"error": "cabinet required"}), 400
    d_from, d_to = _parse_period(request.args)
    if (d_to - d_from).days < 0:
        return jsonify({"error": "Неверный период: to < from"}), 400
    if (d_to - d_from).days > 31:
        return jsonify({"error": "Период не более 31 дня (ограничение WB API)"}), 400
    import uuid, threading
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM wb_report_jobs WHERE created_at < NOW() - INTERVAL '1 day'")
                # Дедуп: если по этому кабинету+периоду уже есть свежая задача
                # (выполняется или готова < 5 мин назад) — переиспользуем её, а не
                # плодим фоновые потоки (иначе повторные клики/перезагрузки страницы
                # запускают десятки параллельных генераций и жрут WB rate-limit + CPU).
                cur.execute("""
                    SELECT id FROM wb_report_jobs
                    WHERE cabinet=%s AND date_from=%s AND date_to=%s
                      AND status IN ('running','done')
                      AND created_at > NOW() - INTERVAL '5 minutes'
                    ORDER BY created_at DESC LIMIT 1
                """, (cabinet, d_from, d_to))
                existing = cur.fetchone()
                if existing:
                    return jsonify({"job_id": existing["id"], "reused": True})
                job_id = uuid.uuid4().hex
                cur.execute(
                    "INSERT INTO wb_report_jobs (id, cabinet, date_from, date_to, status) "
                    "VALUES (%s,%s,%s,%s,'running')",
                    (job_id, cabinet, d_from, d_to)
                )
            conn.commit()
    except Exception as e:
        logging.exception("wb_report_start failed")
        return jsonify({"error": str(e)}), 500
    threading.Thread(
        target=_run_wb_report_job, args=(job_id, cabinet, d_from, d_to), daemon=True
    ).start()
    return jsonify({"job_id": job_id})


@app.route("/wb/report/status")
def wb_report_status():
    err = _check_wb_access()
    if err:
        return err
    job_id = request.args.get("job", "").strip()
    if not job_id:
        return jsonify({"error": "job required"}), 400
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT status, error, EXTRACT(EPOCH FROM (NOW()-created_at)) AS age "
                "FROM wb_report_jobs WHERE id=%s", (job_id,)
            )
            row = cur.fetchone()
    if not row:
        return jsonify({"status": "error", "error": "Задача не найдена (устарела?)"}), 404
    # Защита от «зависших» задач (перезапуск воркера, бан WB на часы): считаем
    # задачу провалившейся, если она в статусе running дольше 8 минут.
    if row["status"] == "running" and (row["age"] or 0) > 480:
        return jsonify({"status": "error",
                        "error": "Превышено время на сервере (>8 мин). Вероятно, WB API "
                                 "ограничил доступ по rate-limit — попробуйте через 10–15 минут."})
    return jsonify({"status": row["status"], "error": row["error"]})


@app.route("/wb/report/result")
def wb_report_result():
    err = _check_wb_access()
    if err:
        return err
    job_id = request.args.get("job", "").strip()
    if not job_id:
        return Response("job required", status=400)
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT status, html, error FROM wb_report_jobs WHERE id=%s", (job_id,))
            row = cur.fetchone()
    if not row:
        return Response("Отчёт не найден", status=404)
    if row["status"] == "done" and row["html"]:
        return Response(row["html"], mimetype="text/html")
    if row["status"] == "error":
        return Response(f"Ошибка: {row['error']}", status=500)
    return Response("Отчёт ещё готовится", status=425)


@app.route("/wb/report")
def wb_report():
    """Синхронный рендер (прямая ссылка). Для UI используется фоновый /wb/report/start —
    длинный HTTP-запрос режется edge-timeout'ом Railway (502)."""
    err = _check_wb_access()
    if err:
        return err
    cabinet = request.args.get("cabinet", "").strip()
    if not cabinet:
        return Response("?cabinet= обязателен", status=400)
    d_from, d_to = _parse_period(request.args)
    if (d_to - d_from).days < 0:
        return Response("Неверный период: to < from", status=400)
    if (d_to - d_from).days > 31:
        return Response("Период не более 31 дня (ограничение WB API)", status=400)
    try:
        return Response(_build_wb_report_html(cabinet, d_from, d_to), mimetype="text/html")
    except RuntimeError as e:
        return Response(f"Ошибка: {e}", status=400)
    except Exception as e:
        logging.exception("wb_report failure")
        return Response(f"Внутренняя ошибка: {e}", status=500)

@app.route("/api/wb/report")
@require_auth
def api_wb_report():
    u = _session_user()
    if u["role"] in ("pending", "seller"):
        return jsonify({"error": "forbidden"}), 403
    cabinet = request.args.get("cabinet", "").strip()
    if not cabinet:
        return jsonify({"error": "cabinet required"}), 400
    d_from, d_to = _parse_period(request.args)
    try:
        from wb_api import collect_report_data
        data = collect_report_data(cabinet, d_from, d_to)
        data["date_from"] = data["date_from"].isoformat()
        data["date_to"]   = data["date_to"].isoformat()
        data["prev_from"] = data["prev_from"].isoformat()
        data["prev_to"]   = data["prev_to"].isoformat()
        data["campaign_spend"] = {str(k): v for k, v in data["campaign_spend"].items()}
        data["campaign_stats"] = {str(k): v for k, v in data["campaign_stats"].items()}
        return jsonify(data)
    except Exception as e:
        logging.exception("api_wb_report failure")
        return jsonify({"error": str(e)}), 500

@app.route("/api/wb/debug")
@require_auth
def api_wb_debug():
    u = _session_user()
    if u["role"] in ("pending", "seller"):
        return jsonify({"error": "forbidden"}), 403
    cabinet = request.args.get("cabinet", "").strip()
    if not cabinet:
        return jsonify({"error": "cabinet required"}), 400
    d_from, d_to = _parse_period(request.args)
    try:
        from wb_api import WBClient, previous_week, fmt_date
        prev_from, prev_to = previous_week(d_from, d_to)
        out: dict = {
            "cabinet":     cabinet,
            "date_from":   str(d_from),
            "date_to":     str(d_to),
            "prev_from":   str(prev_from),
            "prev_to":     str(prev_to),
            "api_key_env": f"WB_API_KEY_{cabinet.upper()}",
            "api_key_set": bool(os.environ.get(f"WB_API_KEY_{cabinet.upper()}") or
                                os.environ.get("WB_API_KEY")),
        }
        with WBClient(cabinet) as wb:
            r = wb._request("GET", "https://advert-api.wildberries.ru/adv/v1/upd",
                            params={"from": fmt_date(d_from), "to": fmt_date(d_to)})
            out["upd"] = {
                "status": r.status_code,
                "rows_count": (len(r.json()) if r.is_success and isinstance(r.json(), list) else None),
                "preview": r.text[:500],
            }
            r = wb._request("GET", "https://advert-api.wildberries.ru/adv/v1/promotion/count")
            out["promotion_count"] = {"status": r.status_code, "preview": r.text[:500]}
            try:
                ids = list({int(row.get("advertId"))
                            for row in (wb.get_upd(fmt_date(d_from), fmt_date(d_to)) or [])
                            if row.get("advertId")})[:5]
            except Exception:
                ids = []
            out["fullstats"] = {"advert_ids_used": ids}
            if ids:
                r = wb._request("GET", "https://advert-api.wildberries.ru/adv/v3/fullstats",
                                params={"ids": ",".join(map(str, ids)),
                                        "beginDate": fmt_date(d_from),
                                        "endDate":   fmt_date(d_to)})
                out["fullstats"].update({"status": r.status_code, "preview": r.text[:500]})
            out["sales_funnel"] = wb.raw_sales_funnel(
                fmt_date(d_from), fmt_date(d_to),
                fmt_date(prev_from), fmt_date(prev_to),
            )
            out["sales"] = wb.raw_sales(fmt_date(prev_from))
            funnel_json = out["sales_funnel"].get("response_json") or {}
            products = ((funnel_json.get("data") or {}).get("products")) or []
            out["sales_funnel"]["products_count"] = len(products)
            if products:
                out["sales_funnel"]["first_product_keys"] = sorted(products[0].keys())
                stats = products[0].get("statistics") or {}
                out["sales_funnel"]["statistics_keys"] = sorted(stats.keys())
                first_period_key = next(iter(stats), None)
                if first_period_key:
                    period = stats[first_period_key] or {}
                    out["sales_funnel"]["statistics_period_keys"] = (
                        sorted(period.keys()) if isinstance(period, dict) else type(period).__name__
                    )
                    out["sales_funnel"]["first_product_sample"] = products[0]
        return jsonify(out)
    except Exception as e:
        logging.exception("api_wb_debug failure")
        return jsonify({"error": str(e), "cabinet": cabinet}), 500

@app.route("/health")
def health():
    return jsonify({"status": "ok"})

# ─── План-факт ────────────────────────────────────────────────────────────────

@app.route("/plan-fact")
def plan_fact_page():
    u = _session_user()
    if not u:
        return redirect("/auth/login")
    allowed = _planfact_allowed(u)
    if allowed is not None and not allowed:   # seller без доступа
        return Response("Доступ запрещён", status=403)
    cabinets = WB_CABINETS if allowed is None else [c for c in allowed if c in WB_CABINETS]
    if not cabinets:
        return Response("Нет доступных кабинетов", status=403)
    return render_template(
        "plan_fact.html",
        projects=cabinets,
        project_emoji=_project_emoji_dict(),
        current_user=u,
        can_edit=_planfact_can_edit(u),
    )

def _pf_prev_month(month_date: date) -> date:
    """Первое число предыдущего месяца."""
    if month_date.month == 1:
        return month_date.replace(year=month_date.year - 1, month=12, day=1)
    return month_date.replace(month=month_date.month - 1, day=1)


def _pf_load_cache(cabinet: str, month_date: date):
    """Возвращает (data|None, fetched_at|None, is_complete)."""
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT data, fetched_at, is_complete FROM planfact_cache "
                    "WHERE project=%s AND month=%s",
                    (cabinet, month_date)
                )
                row = cur.fetchone()
        if not row:
            return None, None, False
        return row["data"], row["fetched_at"], bool(row["is_complete"])
    except Exception:
        logging.exception("planfact_cache load failed")
        return None, None, False


def _pf_store_cache(cabinet: str, month_date: date, data: dict, is_complete: bool):
    """Сохраняет сырой факт (articles+totals из collect_planfact_data) как JSONB."""
    try:
        blob = json.dumps({"articles": data["articles"], "totals": data["totals"]})
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO planfact_cache (project, month, data, fetched_at, is_complete)
                    VALUES (%s, %s, %s::jsonb, NOW(), %s)
                    ON CONFLICT (project, month) DO UPDATE SET
                        data=EXCLUDED.data, fetched_at=NOW(), is_complete=EXCLUDED.is_complete
                """, (cabinet, month_date, blob, is_complete))
            conn.commit()
    except Exception:
        logging.exception("planfact_cache store failed")


def _pf_maybe_backup(cabinet: str):
    """Раз в 7 дней сохраняет снимок планов+фактов кабинета (на случай потери данных)."""
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT MAX(backup_date) AS last FROM planfact_backups WHERE project=%s",
                    (cabinet,)
                )
                last = cur.fetchone()["last"]
                if last and (date.today() - last).days < 7:
                    return
                cur.execute(
                    "SELECT month, vendor_code, orders_rub_plan, orders_qty_plan, "
                    "sales_rub_plan, sales_qty_plan FROM sales_plans WHERE project=%s",
                    (cabinet,)
                )
                plans = [
                    {"month": r["month"].isoformat(), "vendor_code": r["vendor_code"],
                     "orders_rub_plan": float(r["orders_rub_plan"] or 0),
                     "orders_qty_plan": int(r["orders_qty_plan"] or 0),
                     "sales_rub_plan": float(r["sales_rub_plan"] or 0),
                     "sales_qty_plan": int(r["sales_qty_plan"] or 0)}
                    for r in cur.fetchall()
                ]
                cur.execute(
                    "SELECT month, data FROM planfact_cache WHERE project=%s",
                    (cabinet,)
                )
                facts = [{"month": r["month"].isoformat(), "data": r["data"]} for r in cur.fetchall()]
                if not plans and not facts:
                    return
                cur.execute("""
                    INSERT INTO planfact_backups (backup_date, project, plans, facts)
                    VALUES (CURRENT_DATE, %s, %s::jsonb, %s::jsonb)
                    ON CONFLICT (backup_date, project) DO NOTHING
                """, (cabinet, json.dumps(plans), json.dumps(facts)))
            conn.commit()
    except Exception:
        logging.exception("planfact backup failed")


@app.route("/api/plan-fact")
@require_auth
def api_plan_fact_get():
    u = _session_user()
    allowed = _planfact_allowed(u)
    if allowed is not None and not allowed:
        return jsonify({"error": "forbidden"}), 403
    cabinet    = request.args.get("cabinet", "").strip()
    month_str  = request.args.get("month", "").strip()   # "2026-05"
    force      = request.args.get("refresh", "") in ("1", "true", "yes")
    if not cabinet or not month_str:
        return jsonify({"error": "cabinet and month required"}), 400
    if allowed is not None and cabinet not in allowed:
        return jsonify({"error": "forbidden"}), 403
    import calendar
    try:
        month_date = datetime.strptime(month_str, "%Y-%m").date()
    except ValueError:
        return jsonify({"error": "invalid month format"}), 400
    last_day = calendar.monthrange(month_date.year, month_date.month)[1]
    month_end = month_date.replace(day=last_day)
    d_from   = month_date
    d_to     = min(date.today() - timedelta(days=1), month_end)
    if d_to < d_from:
        d_to = d_from
    try:
        from wb_api import collect_planfact_data

        today = date.today()
        is_past_month    = month_end < today.replace(day=1)
        is_current_month = (month_date.year == today.year and month_date.month == today.month)

        # ── Решаем: брать из кэша БД или дёргать WB API ──
        cached, fetched_at, cache_complete = _pf_load_cache(cabinet, month_date)
        stale = False
        if cached and fetched_at:
            age_h = (datetime.now(fetched_at.tzinfo) - fetched_at).total_seconds() / 3600
            # Прошлый завершённый месяц с полным кэшем — данные финальные, не обновляем.
            # Текущий месяц — обновляем раз в сутки. Будущий — по требованию.
            if is_past_month and cache_complete:
                stale = False
            elif is_current_month:
                stale = age_h >= 24
            else:
                stale = age_h >= 24

        need_fetch = force or (cached is None) or stale
        from_cache = False
        if need_fetch:
            data = collect_planfact_data(cabinet, d_from, d_to)
            is_complete_now = (d_to == month_end and month_end < today)
            _pf_store_cache(cabinet, month_date, data, is_complete_now)
            fetched_at = datetime.now()
        else:
            data = {"articles": cached.get("articles", []), "totals": cached.get("totals", {})}
            from_cache = True

        # ── Раз в 7 дней — бэкап планов+фактов кабинета ──
        _pf_maybe_backup(cabinet)

        # ── Req #5: факт прошлого месяца + откат ср.цены/выкупа берём из кэша БД ──
        # orders_qty_past = ФИНАЛЬНЫЙ факт заказов прошлого месяца (для авто-расчёта плана).
        # avg_price/buyout: берём прошлый месяц ТОЛЬКО если за выбранный месяц их нет.
        # Важно: WB-воронка за «past»-период отдаёт заниженные/нестабильные числа,
        # поэтому источник истины — сохранённый факт прошлого месяца (selected=тот месяц).
        # Если полного снимка прошлого месяца нет (или он неполный, снят до конца месяца),
        # дозапрашиваем один раз из WB — месяц уже закрыт, данные финальны.
        past_month_date = _pf_prev_month(month_date)
        past_last = calendar.monthrange(past_month_date.year, past_month_date.month)[1]
        past_end  = past_month_date.replace(day=past_last)
        past_cached, _pf_ts, past_complete = _pf_load_cache(cabinet, past_month_date)
        if past_end < today and not past_complete:
            try:
                pdata = collect_planfact_data(cabinet, past_month_date, past_end)
                _pf_store_cache(cabinet, past_month_date, pdata, True)
                past_cached = {"articles": pdata["articles"], "totals": pdata["totals"]}
            except Exception:
                logging.exception("prev-month final fetch failed")
        if past_cached:
            past_by_vc = {}
            for pa in past_cached.get("articles", []):
                past_by_vc[pa["vendor_code"]] = {
                    "orders_qty": int(pa.get("orders_qty_fact") or 0),
                    "avg_price":  float(pa.get("avg_price") or 0),
                    "buyout":     float(pa.get("buyout_pct") or 0),
                }
            for a in data["articles"]:
                pv = past_by_vc.get(a["vendor_code"])
                if not pv:
                    continue
                a["orders_qty_past"] = pv["orders_qty"]
                if (not a.get("avg_price")) and pv["avg_price"] > 0:
                    a["avg_price"] = pv["avg_price"]
                    a["avg_price_is_prev"] = True
                if (not a.get("buyout_pct")) and pv["buyout"] > 0:
                    a["buyout_pct"] = pv["buyout"]
                    a["buyout_pct_is_prev"] = True

        # Загружаем план + ярлыки сезонности из БД
        plans: dict = {}
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT vendor_code, orders_rub_plan, orders_qty_plan, "
                    "       sales_rub_plan, sales_qty_plan "
                    "FROM sales_plans WHERE project=%s AND month=%s",
                    (cabinet, month_date)
                )
                for r in cur.fetchall():
                    plans[r["vendor_code"]] = {
                        "orders_rub_plan": float(r["orders_rub_plan"] or 0),
                        "orders_qty_plan": int(r["orders_qty_plan"] or 0),
                        "sales_rub_plan":  float(r["sales_rub_plan"] or 0),
                        "sales_qty_plan":  int(r["sales_qty_plan"] or 0),
                    }
                # Постоянные ярлыки сезонности (не зависят от месяца)
                article_seasons: dict = {}
                cur.execute(
                    "SELECT vendor_code, season_name FROM article_seasons "
                    "WHERE project=%s",
                    (cabinet,)
                )
                for r in cur.fetchall():
                    article_seasons[r["vendor_code"]] = r["season_name"] or ""
                # Все известные сезоны проекта (для дропдауна)
                cur.execute(
                    "SELECT DISTINCT season_name FROM article_seasons "
                    "WHERE project=%s AND season_name != '' ORDER BY season_name",
                    (cabinet,)
                )
                all_season_names = [r["season_name"] for r in cur.fetchall()]
                # Коэффициенты сезонности (за текущий месяц)
                season_coefficients: dict = {}
                cur.execute(
                    "SELECT season_name, coefficient FROM season_coefficients "
                    "WHERE project=%s AND month=%s",
                    (cabinet, month_date)
                )
                for r in cur.fetchall():
                    season_coefficients[r["season_name"]] = float(r["coefficient"] or 0)
                # ФФ-остатки и ожидаемые
                ff_stocks: dict = {}
                cur.execute(
                    "SELECT vendor_code, ff_stock, expected_stock FROM article_ff_stocks "
                    "WHERE project=%s",
                    (cabinet,)
                )
                for r in cur.fetchall():
                    ff_stocks[r["vendor_code"]] = {
                        "ff_stock":       int(r["ff_stock"] or 0),
                        "expected_stock": int(r["expected_stock"] or 0),
                    }
                # Цель кабинета на месяц — план ДРР по заказам
                cur.execute(
                    "SELECT drr_orders_plan FROM planfact_targets WHERE project=%s AND month=%s",
                    (cabinet, month_date)
                )
                _tr = cur.fetchone()
                drr_orders_plan = float(_tr["drr_orders_plan"]) if _tr and _tr["drr_orders_plan"] is not None else None
        today        = date.today()
        days_elapsed = max(1, (d_to - d_from).days + 1)
        days_remaining = max(0, last_day - today.day)
        for a in data["articles"]:
            vc   = a["vendor_code"]
            plan = plans.get(vc, {})
            a.update({
                "orders_rub_plan": plan.get("orders_rub_plan", 0),
                "orders_qty_plan": plan.get("orders_qty_plan", 0),
                "sales_rub_plan":  plan.get("sales_rub_plan",  0),
                "sales_qty_plan":  plan.get("sales_qty_plan",  0),
                # Ярлык сезонности — из постоянной таблицы, не из плана
                "season_name":     article_seasons.get(vc, ""),
                "ff_stock":        ff_stocks.get(vc, {}).get("ff_stock", 0),
                "expected_stock":  ff_stocks.get(vc, {}).get("expected_stock", 0),
            })
            orq  = a["orders_qty_fact"]
            orqp = a["orders_qty_plan"]
            a["orders_per_day"]  = round(orq / days_elapsed, 1)
            a["needed_per_day"]  = round(max(0, (orqp - orq) / days_remaining), 1) if days_remaining > 0 else 0
            a["pct_orders"]      = round(a["orders_rub_fact"] / a["orders_rub_plan"] * 100) if a["orders_rub_plan"] else None
            a["pct_sales"]       = round(a["sales_rub_fact"]  / a["sales_rub_plan"]  * 100) if a["sales_rub_plan"]  else None
        t = data["totals"]
        t.update({
            "orders_rub_plan": sum(a["orders_rub_plan"] for a in data["articles"]),
            "orders_qty_plan": sum(a["orders_qty_plan"] for a in data["articles"]),
            "sales_rub_plan":  sum(a["sales_rub_plan"]  for a in data["articles"]),
            "sales_qty_plan":  sum(a["sales_qty_plan"]  for a in data["articles"]),
        })
        data["meta"] = {
            "d_from":          d_from.isoformat(),
            "d_to":            d_to.isoformat(),
            "days_elapsed":    days_elapsed,
            "days_remaining":  days_remaining,
            "days_in_month":   last_day,
            "pct_orders":      round(t["orders_rub_fact"] / t["orders_rub_plan"] * 100, 1) if t.get("orders_rub_plan") else None,
            "pct_sales":       round(t["sales_rub_fact"]  / t["sales_rub_plan"]  * 100, 1) if t.get("sales_rub_plan")  else None,
            "adv_spend":       t.get("adv_spend", 0),
            "drr_orders":      t.get("drr_orders"),
            "drr_sales":       t.get("drr_sales"),
            "drr_orders_plan": drr_orders_plan,
            "from_cache":      from_cache,
            "fetched_at":      fetched_at.isoformat() if fetched_at else None,
        }
        data["season_coefficients"] = season_coefficients
        data["season_names"]        = all_season_names
        return jsonify(data)
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        logging.exception("api_plan_fact_get failure")
        return jsonify({"error": str(e)}), 500

@app.route("/api/plan-fact", methods=["POST"])
@require_auth
def api_plan_fact_save():
    u = _session_user()
    if not _planfact_can_edit(u):
        return jsonify({"error": "forbidden"}), 403
    allowed = _planfact_allowed(u)
    if allowed is not None and not allowed:
        return jsonify({"error": "forbidden"}), 403
    d       = request.json or {}
    cabinet = d.get("cabinet", "").strip()
    if allowed is not None and cabinet not in allowed:
        return jsonify({"error": "forbidden"}), 403
    month_str = d.get("month", "").strip()
    plans   = d.get("plans", [])
    try:
        month_date = datetime.strptime(month_str, "%Y-%m").date()
    except ValueError:
        return jsonify({"error": "invalid month"}), 400
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                for p in plans:
                    vc = str(p.get("vendor_code") or "").strip()
                    if not vc:
                        continue
                    cur.execute("""
                        INSERT INTO sales_plans
                            (project, month, vendor_code,
                             orders_rub_plan, orders_qty_plan,
                             sales_rub_plan,  sales_qty_plan)
                        VALUES (%s,%s,%s,%s,%s,%s,%s)
                        ON CONFLICT (project, month, vendor_code) DO UPDATE SET
                            orders_rub_plan = EXCLUDED.orders_rub_plan,
                            orders_qty_plan = EXCLUDED.orders_qty_plan,
                            sales_rub_plan  = EXCLUDED.sales_rub_plan,
                            sales_qty_plan  = EXCLUDED.sales_qty_plan
                    """, (
                        cabinet, month_date, vc,
                        float(p.get("orders_rub_plan") or 0),
                        int(p.get("orders_qty_plan") or 0),
                        float(p.get("sales_rub_plan") or 0),
                        int(p.get("sales_qty_plan") or 0),
                    ))
                # Цель кабинета — план ДРР по заказам (общий на месяц)
                if "drr_orders_plan" in d:
                    drr_val = d.get("drr_orders_plan")
                    if drr_val in (None, ""):
                        cur.execute(
                            "DELETE FROM planfact_targets WHERE project=%s AND month=%s",
                            (cabinet, month_date)
                        )
                    else:
                        cur.execute("""
                            INSERT INTO planfact_targets (project, month, drr_orders_plan, updated_at)
                            VALUES (%s, %s, %s, NOW())
                            ON CONFLICT (project, month) DO UPDATE SET
                                drr_orders_plan = EXCLUDED.drr_orders_plan, updated_at = NOW()
                        """, (cabinet, month_date, float(drr_val)))
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/plan-fact/clear", methods=["POST"])
@require_auth
def api_plan_fact_clear():
    """Удаляет все плановые значения кабинета за месяц (факт не трогает)."""
    u = _session_user()
    if not _planfact_can_edit(u):
        return jsonify({"error": "forbidden"}), 403
    allowed = _planfact_allowed(u)
    if allowed is not None and not allowed:
        return jsonify({"error": "forbidden"}), 403
    cabinet   = request.args.get("cabinet", "").strip()
    month_str = request.args.get("month", "").strip()
    if not cabinet or not month_str:
        return jsonify({"error": "cabinet and month required"}), 400
    if allowed is not None and cabinet not in allowed:
        return jsonify({"error": "forbidden"}), 403
    try:
        month_date = datetime.strptime(month_str, "%Y-%m").date()
    except ValueError:
        return jsonify({"error": "invalid month"}), 400
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "DELETE FROM sales_plans WHERE project=%s AND month=%s",
                    (cabinet, month_date)
                )
                n = cur.rowcount
                cur.execute(
                    "DELETE FROM planfact_targets WHERE project=%s AND month=%s",
                    (cabinet, month_date)
                )
            conn.commit()
        return jsonify({"ok": True, "deleted": n})
    except Exception as e:
        logging.exception("plan clear failed")
        return jsonify({"error": str(e)}), 500


# ─── Коэффициенты сезонности ──────────────────────────────────────────────────

@app.route("/api/season-coefficients", methods=["POST"])
@require_auth
def api_season_coeff_save():
    u = _session_user()
    if not _planfact_can_edit(u):
        return jsonify({"error": "forbidden"}), 403
    d        = request.json or {}
    cabinet  = d.get("cabinet", "").strip()
    month_str = d.get("month", "").strip()
    coefficients = d.get("coefficients", {})
    if not cabinet or not month_str:
        return jsonify({"error": "cabinet and month required"}), 400
    try:
        month_date = datetime.strptime(month_str, "%Y-%m").date()
    except ValueError:
        return jsonify({"error": "invalid month"}), 400
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                for season_name, coeff in coefficients.items():
                    sn = str(season_name).strip()
                    if not sn:
                        continue
                    cur.execute("""
                        INSERT INTO season_coefficients
                            (project, month, season_name, coefficient)
                        VALUES (%s, %s, %s, %s)
                        ON CONFLICT (project, month, season_name) DO UPDATE SET
                            coefficient = EXCLUDED.coefficient
                    """, (cabinet, month_date, sn, float(coeff or 0)))
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Ярлыки сезонности (постоянные, по артикулу) ─────────────────────────────

@app.route("/api/ff-stocks/upload", methods=["POST"])
@require_auth
def api_ff_stocks_upload():
    """Загрузка Excel с колонками: Артикул, Остаток на ФФ, Ожидаем остатки."""
    u = _session_user()
    if u["role"] not in ("admin", "employee"):
        return jsonify({"error": "forbidden"}), 403
    cabinet = request.args.get("cabinet", "").strip()
    if not cabinet or cabinet not in WB_CABINETS:
        return jsonify({"error": "unknown cabinet"}), 400
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "no file"}), 400
    try:
        import openpyxl, io
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True)
        ws = wb.active
        # Определяем индексы нужных колонок по заголовку (первая строка)
        headers = [str(c.value or "").strip().lower() for c in ws[1]]
        def find_col(variants):
            for v in variants:
                for i, h in enumerate(headers):
                    if v in h:
                        return i
            return None
        idx_art  = find_col(["артикул"])
        idx_ff   = find_col(["итого остатков", "итого", "фф", "ff", "на складе", "склад"])
        idx_exp  = find_col(["ожидаем", "заказ", "приход", "ожид"])
        if idx_art is None:
            return jsonify({"error": "Колонка 'Артикул' не найдена"}), 400

        rows_upserted = 0
        with get_conn() as conn:
            with conn.cursor() as cur:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    vc = str(row[idx_art] or "").strip()
                    if not vc:
                        continue
                    def _to_int(val):
                        try:
                            return int(float(val)) if val not in (None, "", "#REF!", "#N/A", "#VALUE!", "#DIV/0!", "#NAME?") else 0
                        except (ValueError, TypeError):
                            return 0
                    ff  = _to_int(row[idx_ff]  if idx_ff  is not None else None)
                    exp = _to_int(row[idx_exp] if idx_exp is not None else None)
                    cur.execute("""
                        INSERT INTO article_ff_stocks (project, vendor_code, ff_stock, expected_stock, updated_at)
                        VALUES (%s, %s, %s, %s, NOW())
                        ON CONFLICT (project, vendor_code) DO UPDATE
                          SET ff_stock=EXCLUDED.ff_stock,
                              expected_stock=EXCLUDED.expected_stock,
                              updated_at=NOW()
                    """, (cabinet, vc, ff, exp))
                    rows_upserted += 1
            conn.commit()
        return jsonify({"ok": True, "rows": rows_upserted})
    except Exception as e:
        logging.exception("ff_stocks_upload error")
        return jsonify({"error": str(e)}), 500

@app.route("/api/article-seasons", methods=["POST"])
@require_auth
def api_article_seasons_save():
    u = _session_user()
    if not _planfact_can_edit(u):
        return jsonify({"error": "forbidden"}), 403
    d           = request.json or {}
    cabinet     = d.get("cabinet", "").strip()
    vendor_code = d.get("vendor_code", "").strip()
    season_name = d.get("season_name", "").strip()
    if not cabinet or not vendor_code:
        return jsonify({"error": "cabinet and vendor_code required"}), 400
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO article_seasons (project, vendor_code, season_name, updated_at)
                    VALUES (%s, %s, %s, NOW())
                    ON CONFLICT (project, vendor_code) DO UPDATE SET
                        season_name = EXCLUDED.season_name,
                        updated_at  = NOW()
                """, (cabinet, vendor_code, season_name))
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Дизайнерские задачи ──────────────────────────────────────────────────────

def _design_auth(u):
    """True если пользователь имеет доступ к разделу дизайна."""
    if not u:
        return False
    if u["role"] in ("admin", "employee"):
        return True
    return u["role"] == "seller" and bool(u.get("design_access"))

def serialize_design_task(t):
    d = dict(t)
    for key in ("created_at", "done_at"):
        if d.get(key):
            d[key] = d[key].isoformat()
    return d

def serialize_design_attachment(a):
    d = dict(a)
    if d.get("created_at"):
        d["created_at"] = d["created_at"].isoformat()
    d.pop("file_data", None)   # не отдавать бинарь в JSON
    return d

@app.route("/api/design-tasks")
@require_auth
def api_design_tasks_list():
    u = _session_user()
    allowed = _design_projects(u)
    if allowed is not None and not allowed:
        return jsonify({"error": "forbidden"}), 403
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                if allowed is None:
                    # admin/employee — все задачи
                    cur.execute("""
                        SELECT dt.*,
                               COUNT(da.id) FILTER (WHERE da.attach_role='brief')  AS brief_count,
                               COUNT(da.id) FILTER (WHERE da.attach_role='result') AS result_count,
                               COUNT(dc.id) AS comment_count
                        FROM design_tasks dt
                        LEFT JOIN design_attachments da ON da.task_id = dt.id
                        LEFT JOIN design_comments dc ON dc.task_id = dt.id
                        GROUP BY dt.id
                        ORDER BY dt.done ASC, dt.created_at DESC
                    """)
                else:
                    # seller — только свои проекты
                    cur.execute("""
                        SELECT dt.*,
                               COUNT(da.id) FILTER (WHERE da.attach_role='brief')  AS brief_count,
                               COUNT(da.id) FILTER (WHERE da.attach_role='result') AS result_count,
                               COUNT(dc.id) AS comment_count
                        FROM design_tasks dt
                        LEFT JOIN design_attachments da ON da.task_id = dt.id
                        LEFT JOIN design_comments dc ON dc.task_id = dt.id
                        WHERE dt.project = ANY(%s)
                        GROUP BY dt.id
                        ORDER BY dt.done ASC, dt.created_at DESC
                    """, (allowed,))
                return jsonify([serialize_design_task(dict(r)) for r in cur.fetchall()])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/design-tasks", methods=["POST"])
@require_auth
def api_design_tasks_create():
    u = _session_user()
    allowed = _design_projects(u)
    if allowed is not None and not allowed:
        return jsonify({"error": "forbidden"}), 403
    d = request.json or {}
    title = (d.get("title") or "").strip()
    if not title:
        return jsonify({"error": "title required"}), 400
    project = d.get("project") or None
    # Для селлера проект должен быть из его списка
    if allowed is not None and project and project not in allowed:
        return jsonify({"error": "forbidden"}), 403
    try:
        creator_email  = u.get("email", "")
        assignee_email = d.get("assignee_email", "")
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO design_tasks "
                    "(title, description, priority, assignee, assignee_email, due, project, created_by, created_by_email) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
                    (title, d.get("description", ""), d.get("priority", "med"),
                     d.get("assignee") or None, assignee_email,
                     d.get("due") or None, project,
                     _first_name(u.get("name", "")), creator_email)
                )
                new_id = cur.fetchone()["id"]
                if assignee_email and assignee_email != creator_email:
                    _push_notification(
                        cur, assignee_email, "task_assigned",
                        f"Новая задача дизайнеру: {title}",
                        f"Назначена вам в проекте {project or '—'}",
                        new_id,
                    )
            conn.commit()
        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/design-tasks/<int:task_id>", methods=["PUT"])
@require_auth
def api_design_tasks_update(task_id):
    u = _session_user()
    if not _design_auth(u):
        return jsonify({"error": "forbidden"}), 403
    d = request.json or {}
    project = d.get("project") or None
    allowed = _design_projects(u)
    if allowed is not None and project and project not in allowed:
        return jsonify({"error": "forbidden"}), 403
    try:
        u_email            = u.get("email", "") if u else ""
        changed_by         = _first_name(u.get("name", "")) if u else "unknown"
        new_assignee_email = d.get("assignee_email", "")
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT title, assignee_email, created_by, created_by_email FROM design_tasks WHERE id=%s",
                    (task_id,)
                )
                row = cur.fetchone()
                new_title    = d.get("title","").strip() or None
                new_priority = d.get("priority","med")
                new_due      = d.get("due") or None
                cur.execute(
                    "UPDATE design_tasks SET title=%s, priority=%s, assignee=%s, assignee_email=%s, due=%s, project=%s "
                    "WHERE id=%s",
                    (new_title, new_priority, d.get("assignee") or None,
                     new_assignee_email, new_due, project, task_id)
                )
                if row:
                    # Пишем историю изменений
                    hist_fields = [
                        ("Название",    row.get("title"),          new_title),
                        ("Исполнитель", row.get("assignee_email"), new_assignee_email),
                    ]
                    for fname, old_v, new_v in hist_fields:
                        if str(old_v or "") != str(new_v or ""):
                            cur.execute(
                                "INSERT INTO design_task_history (task_id, changed_by, field, old_value, new_value) "
                                "VALUES (%s,%s,%s,%s,%s)",
                                (task_id, changed_by, fname, old_v, new_v)
                            )
                if row:
                    notified = set()
                    old_assignee_email = row.get("assignee_email") or ""
                    owner_email        = _resolve_owner_email_design(cur, task_id, row.get("created_by_email") or "", row.get("created_by") or "")
                    task_title         = d.get("title") or row.get("title") or ""
                    # Новый исполнитель
                    if new_assignee_email and new_assignee_email != old_assignee_email and new_assignee_email != u_email:
                        _push_notification(cur, new_assignee_email, "task_assigned",
                            f"Задача дизайнеру назначена вам: {task_title}",
                            f"Назначил(а): {changed_by}", task_id)
                        notified.add(new_assignee_email)
                    # Создателю при любом изменении
                    if owner_email and owner_email != u_email and owner_email not in notified:
                        _push_notification(cur, owner_email, "task_updated",
                            f"Задача дизайнеру изменена: {task_title}",
                            f"Изменил(а): {changed_by}", task_id)
                    logging.info(f"[notif] design update task={task_id} owner={owner_email} u={u_email}")
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/design-tasks/<int:task_id>/done", methods=["POST"])
@require_auth
def api_design_tasks_done(task_id):
    u = _session_user()
    if not _design_auth(u):
        return jsonify({"error": "forbidden"}), 403
    try:
        d          = request.json or {}
        comment    = (d.get("comment") or "").strip()
        url        = (d.get("url") or "").strip()
        u_email    = u.get("email", "") if u else ""
        changed_by = _first_name(u.get("name", "")) if u else "unknown"
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE design_tasks SET done=TRUE, done_at=NOW() WHERE id=%s RETURNING title, created_by, created_by_email",
                    (task_id,)
                )
                row = cur.fetchone()
                if row:
                    # История: кто и когда закрыл (нужно для сводки по выполненным)
                    cur.execute(
                        "INSERT INTO design_task_history (task_id, changed_by, field, old_value, new_value) "
                        "VALUES (%s,%s,%s,%s,%s)",
                        (task_id, changed_by, "Статус", "В работе", "Готово")
                    )
                    # Сохраняем комментарий закрытия
                    text = comment
                    if url:
                        text = (text + "\n\nСсылка: " + url) if text else "Ссылка: " + url
                    if text:
                        cur.execute(
                            "INSERT INTO design_comments (task_id, author, text) VALUES (%s,%s,%s)",
                            (task_id, changed_by, text)
                        )
                    owner_email = _resolve_owner_email_design(cur, task_id, row.get("created_by_email") or "", row.get("created_by") or "")
                    if owner_email and owner_email != u_email:
                        _push_notification(cur, owner_email, "status_changed",
                            f"Задача дизайнеру закрыта: {row['title']}",
                            f"Закрыл(а): {changed_by}", task_id)
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/design-tasks/<int:task_id>/undone", methods=["POST"])
@require_auth
def api_design_tasks_undone(task_id):
    u = _session_user()
    if not _design_auth(u):
        return jsonify({"error": "forbidden"}), 403
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "UPDATE design_tasks SET done=FALSE, done_at=NULL WHERE id=%s", (task_id,)
                )
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/design-tasks/<int:task_id>", methods=["DELETE"])
@require_auth
def api_design_tasks_delete(task_id):
    u = _session_user()
    if not _design_auth(u):
        return jsonify({"error": "forbidden"}), 403
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM design_tasks WHERE id=%s", (task_id,))
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/design-tasks/<int:task_id>/comments")
@require_auth
def api_design_get_comments(task_id):
    if not _design_auth(_session_user()):
        return jsonify({"error": "forbidden"}), 403
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id, task_id, author, text, created_at FROM design_comments "
                    "WHERE task_id=%s ORDER BY created_at ASC", (task_id,)
                )
                return jsonify(_anon_comments(cur.fetchall(), _session_user()))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/design-tasks/<int:task_id>/comments", methods=["POST"])
@require_auth
def api_design_add_comment(task_id):
    u = _session_user()
    if not _design_auth(u):
        return jsonify({"error": "forbidden"}), 403
    try:
        u_email    = u.get("email", "") if u else ""
        author     = _first_name(u.get("name", "")) or "Аноним"
        d          = request.json or {}
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO design_comments (task_id, author, text) VALUES (%s,%s,%s)",
                    (task_id, author, d.get("text",""))
                )
                cur.execute(
                    "SELECT title, created_by_email, assignee_email FROM design_tasks WHERE id=%s",
                    (task_id,)
                )
                task = cur.fetchone()
                if task:
                    notified = set()
                    for email in [task["created_by_email"], task["assignee_email"]]:
                        if email and email != u_email and email not in notified:
                            _push_notification(cur, email, "comment_added",
                                f"Новый комментарий: {task['title']}",
                                f"{author}: {d.get('text','')[:80]}", task_id)
                            notified.add(email)
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/design-tasks/<int:task_id>/history")
@require_auth
def api_design_task_history(task_id):
    u = _session_user()
    if not u or u["role"] != "admin":
        return jsonify({"error": "forbidden"}), 403
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id, changed_by, field, old_value, new_value, changed_at "
                    "FROM design_task_history WHERE task_id=%s ORDER BY changed_at DESC",
                    (task_id,)
                )
                return jsonify([{
                    "id":         r["id"],
                    "changed_by": r["changed_by"],
                    "field":      r["field"],
                    "old_value":  r["old_value"],
                    "new_value":  r["new_value"],
                    "changed_at": r["changed_at"].isoformat() if r["changed_at"] else None,
                } for r in cur.fetchall()])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/design-today")
@require_auth
def api_design_today():
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT dtc.id, dtc.task_id, dtc.user_name,
                           dt.title, dt.project, dt.priority, dt.done
                    FROM design_today_claims dtc
                    JOIN design_tasks dt ON dt.id = dtc.task_id
                    WHERE dtc.claimed_date = CURRENT_DATE
                    ORDER BY dtc.user_name, dtc.id
                """)
                return jsonify([dict(r) for r in cur.fetchall()])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/design-today/claim", methods=["POST"])
@require_auth
def api_design_today_claim():
    u = _session_user()
    if u["role"] not in ("admin", "employee"):
        return jsonify({"error": "forbidden"}), 403
    d = request.json or {}
    task_id = d.get("task_id")
    if not task_id:
        return jsonify({"error": "task_id required"}), 400
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT id FROM design_tasks WHERE id=%s AND done=FALSE", (task_id,))
                if not cur.fetchone():
                    return jsonify({"error": "task not found or already done"}), 404
                cur.execute("""
                    INSERT INTO design_today_claims (task_id, user_name, claimed_date)
                    VALUES (%s, %s, CURRENT_DATE)
                    ON CONFLICT (task_id, claimed_date) DO NOTHING
                    RETURNING id
                """, (task_id, u["name"]))
                row = cur.fetchone()
            conn.commit()
        return jsonify({"ok": True}) if row else jsonify({"ok": False, "error": "already_claimed"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/design-today/unclaim", methods=["POST"])
@require_auth
def api_design_today_unclaim():
    u = _session_user()
    d = request.json or {}
    task_id = d.get("task_id")
    if not task_id:
        return jsonify({"error": "task_id required"}), 400
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "DELETE FROM design_today_claims WHERE task_id=%s AND claimed_date=CURRENT_DATE AND user_name=%s",
                    (task_id, u["name"])
                )
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/design-tasks/<int:task_id>/attachments")
@require_auth
def api_design_attachments_list(task_id):
    u = _session_user()
    if not _design_auth(u):
        return jsonify({"error": "forbidden"}), 403
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id, task_id, attach_role, attach_type, name, url, mime_type, created_by, created_at "
                    "FROM design_attachments WHERE task_id=%s ORDER BY created_at ASC",
                    (task_id,)
                )
                return jsonify([serialize_design_attachment(dict(r)) for r in cur.fetchall()])
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/design-tasks/<int:task_id>/attachments", methods=["POST"])
@require_auth
def api_design_attachments_add(task_id):
    u = _session_user()
    if not _design_auth(u):
        return jsonify({"error": "forbidden"}), 403
    try:
        ct = request.content_type or ""
        if "multipart" in ct:
            # Загрузка файла
            f = request.files.get("file")
            if not f or not f.filename:
                return jsonify({"error": "no file"}), 400
            attach_role = request.form.get("role", "brief")
            name = f.filename
            mime_type = f.content_type or "application/octet-stream"
            file_data = f.read()
            if len(file_data) > 20 * 1024 * 1024:
                return jsonify({"error": "file too large (max 20MB)"}), 400
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "INSERT INTO design_attachments "
                        "(task_id, attach_role, attach_type, name, mime_type, file_data, created_by) "
                        "VALUES (%s,%s,'file',%s,%s,%s,%s)",
                        (task_id, attach_role, name, mime_type, file_data, u.get("name", ""))
                    )
                conn.commit()
            return jsonify({"ok": True})
        else:
            # Ссылка
            d = request.json or {}
            url = (d.get("url") or "").strip()
            name = (d.get("name") or url).strip()
            attach_role = d.get("role", "brief")
            if not url:
                return jsonify({"error": "url required"}), 400
            with get_conn() as conn:
                with conn.cursor() as cur:
                    cur.execute(
                        "INSERT INTO design_attachments "
                        "(task_id, attach_role, attach_type, name, url, created_by) "
                        "VALUES (%s,%s,'link',%s,%s,%s)",
                        (task_id, attach_role, name, url, u.get("name", ""))
                    )
                conn.commit()
            return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/design-attachments/<int:att_id>/download")
@require_auth
def api_design_attachment_download(att_id):
    u = _session_user()
    if not _design_auth(u):
        return Response("Доступ запрещён", status=403)
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT name, mime_type, file_data FROM design_attachments "
                    "WHERE id=%s AND attach_type='file'",
                    (att_id,)
                )
                row = cur.fetchone()
        if not row or not row["file_data"]:
            return Response("Файл не найден", status=404)
        name      = row["name"]
        mime_type = row["mime_type"]
        file_data = row["file_data"]
        raw = bytes(file_data)
        from urllib.parse import quote
        encoded_name = quote(name)
        return Response(
            raw,
            mimetype=mime_type or "application/octet-stream",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}",
                "Content-Length": str(len(raw)),
            }
        )
    except Exception as e:
        return Response(f"Ошибка: {e}", status=500)

@app.route("/api/design-attachments/<int:att_id>", methods=["DELETE"])
@require_auth
def api_design_attachment_delete(att_id):
    u = _session_user()
    if not _design_auth(u):
        return jsonify({"error": "forbidden"}), 403
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM design_attachments WHERE id=%s", (att_id,))
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Юнит-экономика ───────────────────────────────────────────────────────────
@app.route("/unit")
@require_unit_page
def unit_page():
    u = _session_user()
    allowed = _unit_projects(u)
    cabinets = WB_CABINETS if allowed is None else [c for c in WB_CABINETS if c in allowed]
    return render_template(
        "unit.html",
        current_user=u,
        projects=cabinets,
        project_emoji=_project_emoji_dict(),
    )


def _unit_row_params(d):
    return {
        "project":          d.get("project", ""),
        "brand":            d.get("brand", ""),
        "wb_article":       d.get("wb_article") or None,
        "seller_article":   d.get("seller_article", ""),
        "cost_price":       d.get("cost_price") or 0,
        "logistics_to_wb":  d.get("logistics_to_wb") or 0,
        "packaging":        d.get("packaging") or 0,
        "overhead":         d.get("overhead") or 0,
        "defect_pct":       d.get("defect_pct") or 0,
        "width_cm":         d.get("width_cm") or None,
        "length_cm":        d.get("length_cm") or None,
        "height_cm":        d.get("height_cm") or None,
        "liters":           d.get("liters") or None,
        "redemption_pct":   d.get("redemption_pct") or 100,
        "warehouse":        d.get("warehouse", ""),
        "irp":              d.get("irp") or 1.0,
        "il":               d.get("il") or 1.0,
        "logistics_ktr":    d.get("logistics_ktr") or None,
        "reception_coef":   d.get("reception_coef", "x0"),
        "storage_per_day":  0,
        "commission_pct":   d.get("commission_pct") or 36,
        "wb_price":         d.get("wb_price") or 0,
        "drr_pct":          d.get("drr_pct") or 0,
        "drr_external_rub": d.get("drr_external_rub") or 0,
        "stock":            d.get("stock") or 0,
        "spp_pct":          d.get("spp_pct") or 0,
        "tax_system":       d.get("tax_system") or "УСН 7%",
        "tax_pct":          d.get("tax_pct") or 7,
        "status":           d.get("status") or "",
        "fbs_point":            d.get("fbs_point", "ПВЗ"),
        "fbs_sc_coef":          d.get("fbs_sc_coef") if d.get("fbs_sc_coef") not in (None, "") else 100,
        "fbs_commission_pct":   d.get("fbs_commission_pct") or 0,
        "fbs_assembly_coef":    d.get("fbs_assembly_coef") or 0,
        "fbs_own_cost":         d.get("fbs_own_cost") or 0,
        "fbs_return_days":      d.get("fbs_return_days") or 0,
        "predmet":              d.get("predmet", ""),
        "fbo_storage_days":     d.get("fbo_storage_days") if d.get("fbo_storage_days") not in (None, "") else 30,
        "fbo_accept_coef":      d.get("fbo_accept_coef") or 0,
    }


@app.route("/api/unit-rows")
@require_unit_api
def api_unit_rows_get():
    project = request.args.get("project", "")
    if not _check_unit_project(_session_user(), project):
        return jsonify({"error": "forbidden"}), 403
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT id, project, brand, wb_article, seller_article,
                           cost_price, logistics_to_wb, packaging, overhead,
                           defect_pct, width_cm, length_cm, height_cm, liters,
                           redemption_pct, warehouse, irp, il, logistics_ktr,
                           reception_coef, storage_per_day, commission_pct,
                           wb_price, drr_pct, drr_external_rub, stock, spp_pct,
                           tax_system, tax_pct, status,
                           fbs_point, fbs_sc_coef, fbs_commission_pct,
                           fbs_assembly_coef, fbs_own_cost, fbs_return_days, predmet,
                           fbo_storage_days, fbo_accept_coef
                    FROM unit_economics
                    WHERE project = %s
                    ORDER BY id ASC
                """, (project,))
                rows = [dict(r) for r in cur.fetchall()]  # dict_row → copy
        for row in rows:
            for k, v in row.items():
                if hasattr(v, '__float__') and not isinstance(v, (int, bool)):
                    row[k] = float(v) if v is not None else None
        # Ежедневный бэкап (ON CONFLICT DO NOTHING = не дублируем)
        if project and rows:
            try:
                with get_conn() as conn:
                    with conn.cursor() as cur:
                        cur.execute("""
                            INSERT INTO unit_economics_backup (backup_date, project, data)
                            SELECT CURRENT_DATE, %s, jsonb_agg(row_to_json(ue))
                            FROM unit_economics ue WHERE project = %s
                            ON CONFLICT (backup_date, project) DO NOTHING
                        """, (project, project))
                    conn.commit()
            except Exception:
                pass
        return jsonify(rows)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/unit-rows", methods=["POST"])
@require_unit_api
def api_unit_rows_create():
    raw = request.get_json(silent=True) or {}
    try:
        p = _unit_row_params(raw)
        if not _check_unit_project(_session_user(), p.get("project", "")):
            return jsonify({"error": "forbidden"}), 403
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO unit_economics
                        (project, brand, wb_article, seller_article,
                         cost_price, logistics_to_wb, packaging, overhead,
                         defect_pct, width_cm, length_cm, height_cm, liters,
                         redemption_pct, warehouse, irp, il, logistics_ktr,
                         reception_coef, storage_per_day, commission_pct,
                         wb_price, drr_pct, drr_external_rub, stock, spp_pct,
                         tax_system, tax_pct, status,
                         fbs_point, fbs_sc_coef, fbs_commission_pct,
                         fbs_assembly_coef, fbs_own_cost, fbs_return_days, predmet,
                         fbo_storage_days, fbo_accept_coef)
                    VALUES
                        (%(project)s, %(brand)s, %(wb_article)s, %(seller_article)s,
                         %(cost_price)s, %(logistics_to_wb)s, %(packaging)s, %(overhead)s,
                         %(defect_pct)s, %(width_cm)s, %(length_cm)s, %(height_cm)s, %(liters)s,
                         %(redemption_pct)s, %(warehouse)s, %(irp)s, %(il)s, %(logistics_ktr)s,
                         %(reception_coef)s, %(storage_per_day)s, %(commission_pct)s,
                         %(wb_price)s, %(drr_pct)s, %(drr_external_rub)s, %(stock)s, %(spp_pct)s,
                         %(tax_system)s, %(tax_pct)s, %(status)s,
                         %(fbs_point)s, %(fbs_sc_coef)s, %(fbs_commission_pct)s,
                         %(fbs_assembly_coef)s, %(fbs_own_cost)s, %(fbs_return_days)s, %(predmet)s,
                         %(fbo_storage_days)s, %(fbo_accept_coef)s)
                    RETURNING id
                """, p)
                new_id = cur.fetchone()["id"]
            conn.commit()
        return jsonify({"id": new_id})
    except Exception as e:
        logging.error(f"unit-rows-create {raw.get('project','')}: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/unit-rows/<int:row_id>", methods=["PUT"])
@require_unit_api
def api_unit_rows_update(row_id):
    raw = request.get_json(silent=True) or {}
    try:
        p = _unit_row_params(raw)
        p["id"] = row_id
        # Проверяем project как из тела запроса, так и из БД (защита от подмены)
        u = _session_user()
        if not _check_unit_project(u, p.get("project", "")):
            return jsonify({"error": "forbidden"}), 403
    except Exception as e:
        logging.error(f"unit-rows-update {row_id} (params): {e}")
        return jsonify({"error": str(e)}), 500
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT project FROM unit_economics WHERE id=%s", (row_id,))
                existing = cur.fetchone()
        if existing and not _check_unit_project(u, existing["project"]):
            return jsonify({"error": "forbidden"}), 403
    except Exception:
        pass
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE unit_economics SET
                        project=%(project)s, brand=%(brand)s,
                        wb_article=%(wb_article)s, seller_article=%(seller_article)s,
                        cost_price=%(cost_price)s, logistics_to_wb=%(logistics_to_wb)s,
                        packaging=%(packaging)s, overhead=%(overhead)s,
                        defect_pct=%(defect_pct)s, width_cm=%(width_cm)s,
                        length_cm=%(length_cm)s, height_cm=%(height_cm)s, liters=%(liters)s,
                        redemption_pct=%(redemption_pct)s, warehouse=%(warehouse)s,
                        irp=%(irp)s, il=%(il)s, logistics_ktr=%(logistics_ktr)s,
                        reception_coef=%(reception_coef)s, storage_per_day=%(storage_per_day)s,
                        commission_pct=%(commission_pct)s, wb_price=%(wb_price)s,
                        drr_pct=%(drr_pct)s, drr_external_rub=%(drr_external_rub)s,
                        stock=%(stock)s, spp_pct=%(spp_pct)s,
                        tax_system=%(tax_system)s, tax_pct=%(tax_pct)s,
                        status=%(status)s,
                        fbs_point=%(fbs_point)s, fbs_sc_coef=%(fbs_sc_coef)s,
                        fbs_commission_pct=%(fbs_commission_pct)s,
                        fbs_assembly_coef=%(fbs_assembly_coef)s,
                        fbs_own_cost=%(fbs_own_cost)s, fbs_return_days=%(fbs_return_days)s,
                        predmet=%(predmet)s,
                        fbo_storage_days=%(fbo_storage_days)s, fbo_accept_coef=%(fbo_accept_coef)s,
                        updated_at=NOW()
                    WHERE id=%(id)s
                """, p)
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        logging.error(f"unit-rows-update {row_id}: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/unit-rows/<int:row_id>", methods=["DELETE"])
@require_unit_api
def api_unit_rows_delete(row_id):
    u = _session_user()
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT project FROM unit_economics WHERE id=%s", (row_id,))
                existing = cur.fetchone()
        if existing and not _check_unit_project(u, existing["project"]):
            return jsonify({"error": "forbidden"}), 403
    except Exception:
        pass
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM unit_economics WHERE id=%s", (row_id,))
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Юнит Ozon: CRUD ─────────────────────────────────────────────────────────
_OZON_FIELDS = [
    "seller_article", "sku_ozon", "name", "category", "scheme", "cluster",
    "purchase", "delivery", "packaging", "defect_pct",
    "length_cm", "width_cm", "height_cm", "weight_kg",
    "price", "points_pct", "storage_type", "free_days", "turnover_days",
    "drr_pct", "batch", "fact_payout",
]
_OZON_NUM = {"purchase", "delivery", "packaging", "defect_pct", "length_cm",
             "width_cm", "height_cm", "weight_kg", "price", "points_pct",
             "free_days", "turnover_days", "drr_pct", "batch", "fact_payout"}
_OZON_NULLABLE = {"length_cm", "width_cm", "height_cm", "weight_kg"}


def _ozon_row_params(d):
    p = {"project": d.get("project", "")}
    for f in _OZON_FIELDS:
        v = d.get(f)
        if f in _OZON_NUM:
            if v in (None, ""):
                p[f] = None if f in _OZON_NULLABLE else 0
            else:
                p[f] = v
        else:
            p[f] = v if v is not None else ""
    # дефолты текстовых
    if not p.get("scheme"):
        p["scheme"] = "FBO"
    if not p.get("storage_type"):
        p["storage_type"] = "Обычный"
    return p


@app.route("/api/ozon-rows")
@require_unit_api
def api_ozon_rows_get():
    project = request.args.get("project", "")
    if not _check_unit_project(_session_user(), project):
        return jsonify({"error": "forbidden"}), 403
    try:
        cols = "id, project, " + ", ".join(_OZON_FIELDS)
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    f"SELECT {cols} FROM unit_ozon WHERE project=%s ORDER BY id ASC",
                    (project,),
                )
                rows = [dict(r) for r in cur.fetchall()]
        for row in rows:
            for k, v in row.items():
                if hasattr(v, "__float__") and not isinstance(v, (int, bool)):
                    row[k] = float(v) if v is not None else None
        return jsonify(rows)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/ozon-rows", methods=["POST"])
@require_unit_api
def api_ozon_rows_create():
    raw = request.get_json(silent=True) or {}
    try:
        p = _ozon_row_params(raw)
        if not _check_unit_project(_session_user(), p.get("project", "")):
            return jsonify({"error": "forbidden"}), 403
        cols = ["project"] + _OZON_FIELDS
        collist = ", ".join(cols)
        vallist = ", ".join(f"%({c})s" for c in cols)
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    f"INSERT INTO unit_ozon ({collist}) VALUES ({vallist}) RETURNING id", p
                )
                new_id = cur.fetchone()["id"]
            conn.commit()
        return jsonify({"id": new_id})
    except Exception as e:
        logging.error(f"ozon-rows-create {raw.get('project','')}: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/ozon-rows/<int:row_id>", methods=["PUT"])
@require_unit_api
def api_ozon_rows_update(row_id):
    raw = request.get_json(silent=True) or {}
    try:
        p = _ozon_row_params(raw)
        p["id"] = row_id
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT project FROM unit_ozon WHERE id=%s", (row_id,))
                ex = cur.fetchone()
                if ex and not _check_unit_project(_session_user(), ex["project"]):
                    return jsonify({"error": "forbidden"}), 403
                sets = ", ".join(f"{c}=%({c})s" for c in (["project"] + _OZON_FIELDS))
                cur.execute(
                    f"UPDATE unit_ozon SET {sets}, updated_at=NOW() WHERE id=%(id)s", p
                )
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        logging.error(f"ozon-rows-update {row_id}: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/ozon-rows/<int:row_id>", methods=["DELETE"])
@require_unit_api
def api_ozon_rows_delete(row_id):
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT project FROM unit_ozon WHERE id=%s", (row_id,))
                ex = cur.fetchone()
                if ex and not _check_unit_project(_session_user(), ex["project"]):
                    return jsonify({"error": "forbidden"}), 403
                cur.execute("DELETE FROM unit_ozon WHERE id=%s", (row_id,))
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/ozon-import-from-wb", methods=["POST"])
@require_unit_api
def api_ozon_import_from_wb():
    """Переносит базу товара из ВБ-юнитки в Ozon: артикул, бренд, себестоимость,
    упаковку, доставку, брак, габариты. Уже существующие (по артикулу продавца)
    пропускаются, чтобы не затирать введённые в Ozon данные."""
    project = request.args.get("project", "")
    if not project:
        return jsonify({"error": "project required"}), 400
    if not _check_unit_project(_session_user(), project):
        return jsonify({"error": "forbidden"}), 403
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT seller_article FROM unit_ozon WHERE project=%s", (project,)
                )
                existing = {(r["seller_article"] or "").strip().lower()
                            for r in cur.fetchall()}
                cur.execute(
                    "SELECT seller_article, brand, predmet, cost_price, logistics_to_wb, "
                    "packaging, defect_pct, length_cm, width_cm, height_cm "
                    "FROM unit_economics WHERE project=%s ORDER BY id ASC",
                    (project,),
                )
                wb_rows = cur.fetchall()
                added = 0
                for w in wb_rows:
                    sa = (w["seller_article"] or "").strip()
                    if not sa or sa.lower() in existing:
                        continue
                    existing.add(sa.lower())
                    p = _ozon_row_params({
                        "project": project,
                        "seller_article": sa,
                        "name": w["brand"] or "",
                        "category": w["predmet"] or "",
                        "purchase": w["cost_price"] or 0,
                        "delivery": w["logistics_to_wb"] or 0,
                        "packaging": w["packaging"] or 0,
                        "defect_pct": w["defect_pct"] or 0,
                        "length_cm": w["length_cm"],
                        "width_cm": w["width_cm"],
                        "height_cm": w["height_cm"],
                    })
                    cols = ["project"] + _OZON_FIELDS
                    collist = ", ".join(cols)
                    vallist = ", ".join(f"%({c})s" for c in cols)
                    cur.execute(
                        f"INSERT INTO unit_ozon ({collist}) VALUES ({vallist})", p
                    )
                    added += 1
            conn.commit()
        return jsonify({"added": added, "skipped": len(wb_rows) - added})
    except Exception as e:
        logging.error(f"ozon-import-from-wb {project}: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/unit-import-articles")
@require_unit_api
def api_unit_import_articles():
    """
    Загружает список артикулов WB для импорта.
    Воронка продаж (30 дней) не покрывает товары без показов/продаж за период,
    поэтому артикулы дополнительно добираются из остатков на складах и из
    карточек с ценами — иначе товары с остатком, но без активности в воронке,
    не попадают в список (репортилась пропажа части артикулов по проекту).
    """
    project = request.args.get("project", "")
    if not project:
        return jsonify({"error": "project required"}), 400
    if not _check_unit_project(_session_user(), project):
        return jsonify({"error": "forbidden"}), 403
    try:
        from datetime import date, timedelta
        from wb_api import WBClient, fmt_date
        today    = date.today()
        d_from   = fmt_date(today - timedelta(days=30))
        d_to     = fmt_date(today)
        p_from   = fmt_date(today - timedelta(days=60))
        p_to     = fmt_date(today - timedelta(days=31))
        with WBClient(project) as wb:
            raw    = wb.raw_sales_funnel(d_from, d_to, past_from=p_from, past_to=p_to)
            stocks = wb.get_stocks(fmt_date(today - timedelta(days=180)))
            goods  = wb.get_goods_prices()
        status = raw.get("status")
        if status != 200:
            msg = raw.get("response_text_preview") or f"HTTP {status}"
            logging.error(f"unit-import-articles {project}: WB вернул {status}: {msg[:200]}")
            return jsonify({"error": f"WB API вернул {status}: {msg[:150]}"}), 502
        rj = raw.get("response_json") or {}
        products = (rj.get("data") or {}).get("products") or []

        by_id = {}
        for p in products:
            prod  = p.get("product") or {}
            nm_id = prod.get("nmId")
            if not nm_id:
                continue
            by_id[nm_id] = {
                "wb_article":     nm_id,
                "seller_article": prod.get("vendorCode", ""),
                "name":           prod.get("name", ""),
                "brand":          prod.get("brandName", ""),
                "stock":          0,
            }

        # Добавляем/обогащаем остатком артикулы со склада ВБ
        for s in stocks or []:
            nm_id = s.get("nmId")
            if not nm_id:
                continue
            qty = s.get("quantityFull") or s.get("quantity") or 0
            if nm_id in by_id:
                by_id[nm_id]["stock"] += qty
            else:
                by_id[nm_id] = {
                    "wb_article":     nm_id,
                    "seller_article": s.get("supplierArticle", ""),
                    "name":           "",
                    "brand":          "",
                    "stock":          qty,
                }

        # Добираем активные карточки без продаж и без остатка (например, новые)
        for g in goods or []:
            nm_id = g.get("nmID")
            if not nm_id or nm_id in by_id:
                continue
            by_id[nm_id] = {
                "wb_article":     nm_id,
                "seller_article": g.get("vendorCode", ""),
                "name":           "",
                "brand":          "",
                "stock":          0,
            }

        return jsonify(list(by_id.values()))
    except Exception as e:
        logging.error(f"unit-import-articles {project}: {e}")
        return jsonify({"error": str(e)}), 500


def _parse_price_import_excel(file_bytes):
    """
    Парсит Excel с колонками типа "Артикул продавца" / "Артикул ВБ" /
    "Цена прихода" / "ФФ". Названия колонок ищутся по заголовку (первая
    строка), порядок столбцов не важен.
    """
    import io
    import re
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    header = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def find_col(*variants):
        for i, h in enumerate(header):
            if any(v in h for v in variants):
                return i
        return None

    i_seller = find_col("артикул продавца")
    i_wb     = find_col("артикул вб", "артикул wb")
    i_cost   = find_col("цена прихода")
    i_pack   = find_col("фф")
    if i_seller is None or i_wb is None or i_cost is None or i_pack is None:
        raise ValueError(
            "Не найдены нужные столбцы (Артикул продавца / Артикул ВБ / Цена прихода / ФФ)"
        )

    items = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if i_seller >= len(row) or i_wb >= len(row):
            continue
        seller_raw = row[i_seller]
        wb_raw     = row[i_wb]
        if seller_raw is None or wb_raw is None:
            continue
        seller = str(seller_raw).strip()
        m = re.search(r"(\d+)", str(wb_raw))
        if not seller or not m:
            continue
        cost = row[i_cost] if i_cost < len(row) else None
        pack = row[i_pack] if i_pack < len(row) else None
        items.append({
            "seller_article": seller,
            "wb_article":     int(m.group(1)),
            "cost_price":     float(cost) if cost is not None else None,
            "packaging":      float(pack) if pack is not None else None,
        })
    return items


@app.route("/api/unit-import-excel", methods=["POST"])
@require_unit_api
def api_unit_import_excel():
    """
    Импортирует себестоимость и упаковку из Excel-таблицы поставщика,
    сверяя ОБА артикула (продавца и ВБ) — обновляет только полные совпадения.
    """
    project = request.args.get("project", "")
    if not project:
        return jsonify({"error": "project required"}), 400
    if not _check_unit_project(_session_user(), project):
        return jsonify({"error": "forbidden"}), 403
    f = request.files.get("file")
    if not f:
        return jsonify({"error": "файл не получен"}), 400
    try:
        items = _parse_price_import_excel(f.read())
    except Exception as e:
        return jsonify({"error": f"Не удалось прочитать файл: {e}"}), 400

    updated, mismatched, not_found = [], [], []
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id, wb_article, seller_article FROM unit_economics WHERE project=%s",
                    (project,)
                )
                db_rows = cur.fetchall()

            by_key, by_wb, by_seller = {}, {}, {}
            for r in db_rows:
                sa = (r["seller_article"] or "").strip().lower()
                by_key[(r["wb_article"], sa)] = r["id"]
                by_wb.setdefault(r["wb_article"], []).append(r)
                by_seller.setdefault(sa, []).append(r)

            with conn.cursor() as cur:
                for it in items:
                    sa_l = it["seller_article"].lower()
                    row_id = by_key.get((it["wb_article"], sa_l))
                    if row_id:
                        cur.execute(
                            "UPDATE unit_economics SET cost_price=%s, packaging=%s, updated_at=NOW() WHERE id=%s",
                            (it["cost_price"], it["packaging"], row_id)
                        )
                        updated.append(it)
                        continue
                    wb_cand     = by_wb.get(it["wb_article"])
                    seller_cand = by_seller.get(sa_l)
                    if wb_cand or seller_cand:
                        mismatched.append({
                            **it,
                            "found_seller_article": wb_cand[0]["seller_article"] if wb_cand else None,
                            "found_wb_article":     seller_cand[0]["wb_article"] if seller_cand else None,
                        })
                    else:
                        not_found.append(it)
            conn.commit()
    except Exception as e:
        logging.error(f"unit-import-excel {project}: {e}")
        return jsonify({"error": str(e)}), 500

    return jsonify({
        "updated":   len(updated),
        "mismatched": mismatched,
        "not_found": not_found,
    })


def _build_unit_xlsx(project, columns, data_rows, footer=None):
    """Строит xlsx из уже посчитанных на клиенте значений (columns + rows)."""
    import io
    import urllib.parse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    fmt_map = {
        "money": '#,##0" ₽"',
        "pct":   '0.0"%"',
        "int":   '#,##0',
        "num":   '0.0',
        "text":  None,
    }

    wb = Workbook()
    ws = wb.active
    ws.title = (project or "Юнит")[:31]

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F3A5F")
    center   = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for ci, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col.get("label", ""))
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center

    for ri, row in enumerate(data_rows, 2):
        for ci, col in enumerate(columns, 1):
            v = row[ci - 1] if ci - 1 < len(row) else None
            cell = ws.cell(row=ri, column=ci, value=v)
            nf = fmt_map.get(col.get("fmt", "text"))
            if nf and isinstance(v, (int, float)) and not isinstance(v, bool):
                cell.number_format = nf

    if footer:
        fr = len(data_rows) + 2
        foot_font = Font(bold=True)
        foot_fill = PatternFill("solid", fgColor="E8EDF4")
        for ci, col in enumerate(columns, 1):
            v = footer[ci - 1] if ci - 1 < len(footer) else None
            cell = ws.cell(row=fr, column=ci, value=v)
            cell.font = foot_font
            cell.fill = foot_fill
            nf = fmt_map.get(col.get("fmt", "text"))
            if nf and isinstance(v, (int, float)) and not isinstance(v, bool):
                cell.number_format = nf

    for ci, col in enumerate(columns, 1):
        label = str(col.get("label", ""))
        ws.column_dimensions[get_column_letter(ci)].width = max(11, min(20, len(label) + 3))
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = urllib.parse.quote(f"unit_{project}.xlsx")
    return Response(
        buf.read(),
        headers={
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Content-Disposition": f"attachment; filename=\"unit.xlsx\"; filename*=UTF-8''{safe}",
        }
    )


@app.route("/api/unit-export-excel", methods=["POST"])
@require_unit_api
def api_unit_export_excel():
    """
    Выгрузка юнит-экономики в Excel со ВСЕМИ расчётными метриками.
    Клиент присылает уже посчитанные значения (columns + rows + footer),
    чтобы выгрузка 1:1 совпадала с тем, что видно в таблице (включая
    прибыль, маржу, ROI, налог, хранение, капитал и т.д.).
    """
    payload = request.get_json(silent=True) or {}
    project = payload.get("project", "")
    if not project:
        return jsonify({"error": "project required"}), 400
    if not _check_unit_project(_session_user(), project):
        return jsonify({"error": "forbidden"}), 403
    columns = payload.get("columns") or []
    data_rows = payload.get("rows") or []
    footer = payload.get("footer")
    if not columns:
        return jsonify({"error": "no columns"}), 400
    try:
        return _build_unit_xlsx(project, columns, data_rows, footer)
    except Exception as e:
        logging.error(f"unit-export-excel {project}: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/unit-refresh", methods=["POST"])
@require_unit_api
def api_unit_refresh():
    """Обновляет из WB API: Остаток, %выкупа, Литраж, Цена ВБ."""
    project = request.args.get("project", "")
    if not project:
        return jsonify({"error": "project required"}), 400
    if not _check_unit_project(_session_user(), project):
        return jsonify({"error": "forbidden"}), 403
    try:
        from wb_api import WBClient, fmt_date, _funnel_metric, get_product_field
        today  = date.today()
        d_from = fmt_date(today - timedelta(days=30))
        d_to   = fmt_date(today)
        p_from = fmt_date(today - timedelta(days=60))
        p_to   = fmt_date(today - timedelta(days=31))

        # Получаем текущие артикулы из БД
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id, wb_article FROM unit_economics WHERE project=%s AND wb_article IS NOT NULL",
                    (project,)
                )
                db_rows = {r["wb_article"]: r["id"] for r in cur.fetchall()}

        if not db_rows:
            return jsonify({"updated": 0})

        with WBClient(project) as wb:
            # 1. Остатки со склада
            stocks_raw = wb.get_stocks(fmt_date(today - timedelta(days=180)))
            # 2. Воронка продаж — %выкупа
            funnel = wb.get_sales_funnel(d_from, d_to, past_from=p_from, past_to=p_to, limit=1000)
            # 3. Цены и скидки
            prices_raw = wb.get_goods_prices()
            # 4. Габариты карточек — литраж
            liters_by_nm = wb.get_cards_dimensions()

        # ── Агрегация остатков по nmId ──────────────────────────────────────
        # Остаток = на складе (quantity) + в пути от клиента (inWayFromClient)
        # НЕ включаем inWayToClient (в пути до клиента)
        stock_by_nm: dict[int, int] = {}
        for s in stocks_raw:
            nm = s.get("nmId")
            if not nm:
                continue
            qty = (s.get("quantity") or 0) + (s.get("inWayFromClient") or 0)
            stock_by_nm[nm] = stock_by_nm.get(nm, 0) + qty

        # ── %выкупа по nmId из воронки ─────────────────────────────────────
        redemption_by_nm: dict[int, float] = {}
        for p in (funnel.get("data") or {}).get("products") or []:
            nm = get_product_field(p, "nmId")
            if not nm:
                continue
            orders  = _funnel_metric(p, "orderCount",  "selected")
            buyouts = _funnel_metric(p, "buyoutCount", "selected")
            if orders > 0:
                redemption_by_nm[nm] = round(buyouts / orders * 100, 1)

        # ── Цена ВБ по nmId из goods/filter ────────────────────────────────
        # discountedPrice = цена после скидки продавца (Цена ВБ до СПП)
        price_by_nm: dict[int, float] = {}
        for g in prices_raw:
            nm = g.get("nmID")
            if not nm:
                continue
            sizes = g.get("sizes") or []
            if sizes:
                dp = sizes[0].get("discountedPrice") or sizes[0].get("price")
                if dp:
                    price_by_nm[nm] = float(dp)

        # ── Обновляем в БД ──────────────────────────────────────────────────
        updated = 0
        with get_conn() as conn:
            with conn.cursor() as cur:
                for nm_id, row_id in db_rows.items():
                    stock   = stock_by_nm.get(nm_id)
                    red_pct = redemption_by_nm.get(nm_id)
                    liters  = liters_by_nm.get(nm_id)
                    wb_price = price_by_nm.get(nm_id)

                    if all(v is None for v in (stock, red_pct, liters, wb_price)):
                        continue

                    sets, vals = [], []
                    if stock is not None:
                        sets.append("stock=%s"); vals.append(stock)
                    if red_pct is not None:
                        sets.append("redemption_pct=%s"); vals.append(red_pct)
                    if liters is not None:
                        sets.append("liters=%s"); vals.append(liters)
                    if wb_price is not None:
                        sets.append("wb_price=%s"); vals.append(wb_price)
                    sets.append("updated_at=NOW()")
                    vals.append(row_id)
                    cur.execute(f"UPDATE unit_economics SET {', '.join(sets)} WHERE id=%s", vals)
                    updated += 1
            conn.commit()

        return jsonify({
            "updated": updated,
            "stock_found": len(stock_by_nm),
            "redemption_found": len(redemption_by_nm),
            "liters_found": len(liters_by_nm),
            "prices_found": len(price_by_nm),
        })
    except Exception as e:
        logging.error(f"unit-refresh {project}: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/unit-auto-import", methods=["POST"])
@require_unit_api
def api_unit_auto_import():
    """Первичный авто-импорт: добавляет артикулы из WB только если их нет в БД.
    Удаляет «пустые» строки (wb_article IS NULL). Возвращает все строки после вставки."""
    project = request.args.get("project", "")
    if not project:
        return jsonify({"error": "project required"}), 400
    if not _check_unit_project(_session_user(), project):
        return jsonify({"error": "forbidden"}), 403
    try:
        from wb_api import WBClient, fmt_date
        today  = date.today()
        d_from = fmt_date(today - timedelta(days=30))
        d_to   = fmt_date(today)
        p_from = fmt_date(today - timedelta(days=60))
        p_to   = fmt_date(today - timedelta(days=31))

        # Сначала чистим мусорные строки (независимо от результата WB)
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "DELETE FROM unit_economics WHERE project=%s AND (wb_article IS NULL OR wb_article=0)",
                    (project,)
                )
            conn.commit()

        with WBClient(project) as wb:
            raw = wb.raw_sales_funnel(d_from, d_to, past_from=p_from, past_to=p_to)

        status = raw.get("status")
        if status != 200:
            msg = raw.get("response_text_preview") or f"HTTP {status}"
            logging.error(f"unit-auto-import {project}: WB вернул {status}: {msg[:200]}")
            return jsonify({"error": f"WB API вернул {status}: {msg[:150]}"}), 502

        products = ((raw.get("response_json") or {}).get("data") or {}).get("products") or []
        if not products:
            return jsonify({"error": "WB API вернул пустой список артикулов — попробуйте «Загрузить из WB»"}), 502

        # Собираем уникальные артикулы из WB
        wb_articles = {}
        for p in products:
            prod  = p.get("product") or {}
            nm_id = prod.get("nmId")
            if nm_id and nm_id not in wb_articles:
                wb_articles[nm_id] = {
                    "wb_article":     nm_id,
                    "seller_article": prod.get("vendorCode", ""),
                    "brand":          prod.get("brandName", "") or project,
                }

        with get_conn() as conn:
            with conn.cursor() as cur:
                # Узнаём какие wb_article уже есть
                cur.execute(
                    "SELECT wb_article FROM unit_economics WHERE project=%s AND wb_article IS NOT NULL",
                    (project,)
                )
                existing = {r[0] for r in cur.fetchall()}

                # Вставляем только новые
                to_insert = [a for nm, a in wb_articles.items() if nm not in existing]
                for a in to_insert:
                    cur.execute("""
                        INSERT INTO unit_economics
                            (project, brand, wb_article, seller_article,
                             commission_pct, redemption_pct, irp, stock)
                        VALUES (%s,%s,%s,%s, 36, 100, 1.0, 0)
                    """, (project, a["brand"], a["wb_article"], a["seller_article"]))

                # Читаем все строки
                cur.execute("""
                    SELECT id, project, brand, wb_article, seller_article,
                           cost_price, logistics_to_wb, packaging, overhead,
                           defect_pct, width_cm, length_cm, height_cm, liters,
                           redemption_pct, warehouse, irp, logistics_ktr,
                           reception_coef, storage_per_day, commission_pct,
                           wb_price, drr_pct, drr_external_rub, stock
                    FROM unit_economics WHERE project=%s ORDER BY id ASC
                """, (project,))
                rows = [dict(r) for r in cur.fetchall()]  # dict_row → copy
            conn.commit()

        for row in rows:
            for k, v in row.items():
                if hasattr(v, '__float__') and not isinstance(v, (int, bool)):
                    row[k] = float(v) if v is not None else None

        return jsonify({"inserted": len(to_insert), "rows": rows})
    except Exception as e:
        logging.error(f"unit-auto-import {project}: {e}")
        return jsonify({"error": str(e)}), 500


# ─── Контент-воронки ──────────────────────────────────────────────────────────
@app.route("/funnels")
@require_funnel_page
def funnels_page():
    u = _session_user()
    allowed = _funnel_projects(u)
    cabinets = WB_CABINETS if allowed is None else [c for c in WB_CABINETS if c in allowed]
    return render_template("funnels.html", current_user=u, projects=cabinets, project_emoji=_project_emoji_dict())


def serialize_funnel(f):
    d = dict(f)
    for key in ("created_at", "updated_at"):
        if d.get(key):
            d[key] = d[key].isoformat()
    return d

def serialize_slide(s):
    d = dict(s)
    for key in ("created_at", "updated_at"):
        if d.get(key):
            d[key] = d[key].isoformat()
    return d

def serialize_ab_result(r):
    d = dict(r)
    for key in ("created_at", "period_from", "period_to"):
        if d.get(key):
            d[key] = d[key].isoformat()
    return d


def _collect_own_data(project: str, wb_article: int) -> dict:
    """Собирает «свои данные» по артикулу: юнит-экономика, воронка продаж, цена, остатки."""
    from wb_api import WBClient, fmt_date, _funnel_metric, get_product_field

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT cost_price, wb_price, redemption_pct, commission_pct, stock,
                       seller_article, brand
                FROM unit_economics WHERE project=%s AND wb_article=%s LIMIT 1
            """, (project, wb_article))
            ue_row = cur.fetchone()
    unit_economics = None
    if ue_row:
        unit_economics = {
            k: (float(v) if hasattr(v, '__float__') and not isinstance(v, (int, bool)) else v)
            for k, v in dict(ue_row).items()
        }

    today  = date.today()
    d_from = fmt_date(today - timedelta(days=30))
    d_to   = fmt_date(today)
    p_from = fmt_date(today - timedelta(days=60))
    p_to   = fmt_date(today - timedelta(days=31))

    funnel_data: dict = {}
    price_data: dict = {}
    stocks_data: dict = {"total": 0, "by_warehouse": []}
    photo_url: str | None = None
    try:
        with WBClient(project) as wb:
            funnel = wb.get_sales_funnel(d_from, d_to, past_from=p_from, past_to=p_to, nm_ids=[wb_article])
            prices_raw = wb.get_goods_prices(nm_id=wb_article)
            stocks_raw = wb.get_stocks(fmt_date(today - timedelta(days=180)))
            photo_url = wb.get_card_photo(wb_article)

        products = (funnel.get("data") or {}).get("products") or []
        product = next((p for p in products if get_product_field(p, "nmId") == wb_article), None)
        if product:
            views      = _funnel_metric(product, "openCount", "selected")
            cart       = _funnel_metric(product, "cartCount", "selected")
            orders     = _funnel_metric(product, "orderCount", "selected")
            order_sum  = _funnel_metric(product, "orderSum", "selected")
            buyouts    = _funnel_metric(product, "buyoutCount", "selected")
            buyout_sum = _funnel_metric(product, "buyoutSum", "selected")
            funnel_data = {
                "title":              get_product_field(product, "title", ""),
                "brand":              get_product_field(product, "brandName", ""),
                "vendor_code":        get_product_field(product, "vendorCode", ""),
                "views":              views,
                "views_prev":         _funnel_metric(product, "openCount", "past"),
                "cart":               cart,
                "cart_prev":          _funnel_metric(product, "cartCount", "past"),
                "orders":             orders,
                "orders_prev":        _funnel_metric(product, "orderCount", "past"),
                "order_sum":          order_sum,
                "buyouts":            buyouts,
                "buyouts_prev":       _funnel_metric(product, "buyoutCount", "past"),
                "buyout_sum":         buyout_sum,
                "ctr_pct":            round(cart / views * 100, 2) if views else 0,
                "cart_to_order_pct":  round(orders / cart * 100, 2) if cart else 0,
                "buyout_pct":         round(buyouts / orders * 100, 2) if orders else 0,
            }

        for g in prices_raw:
            if g.get("nmID") == wb_article:
                sizes = g.get("sizes") or []
                if sizes:
                    price_data = {
                        "price":            sizes[0].get("price"),
                        "discounted_price": sizes[0].get("discountedPrice"),
                    }
                break

        by_wh: dict[str, int] = {}
        for s in stocks_raw:
            if s.get("nmId") != wb_article:
                continue
            wh = s.get("warehouseName") or "—"
            qty = (s.get("quantity") or 0) + (s.get("inWayFromClient") or 0)
            by_wh[wh] = by_wh.get(wh, 0) + qty
        stocks_data = {
            "total": sum(by_wh.values()),
            "by_warehouse": [{"warehouse": k, "qty": v} for k, v in sorted(by_wh.items(), key=lambda x: -x[1])],
        }
    except Exception as e:
        logging.error(f"_collect_own_data {project}/{wb_article}: {e}")

    return {
        "project":        project,
        "wb_article":     wb_article,
        "unit_economics": unit_economics,
        "funnel":         funnel_data,
        "price":          price_data,
        "stocks":         stocks_data,
        "photo_url":      photo_url,
        "collected_at":   datetime.utcnow().isoformat(),
    }


def _format_ab_context(ab_rows: list[dict]) -> str:
    """Форматирует прошлые А/Б результаты в текст для промпта Claude."""
    if not ab_rows:
        return ""
    lines = []
    for r in ab_rows:
        period = ""
        if r.get("period_from") and r.get("period_to"):
            period = f" ({r['period_from'].strftime('%d.%m')}–{r['period_to'].strftime('%d.%m')})"
        winner  = r.get("winner") or "—"
        summary = r.get("summary") or ""
        lines.append(f"Вариант «{r.get('variant_name','')}»{period}: победитель — {winner}. {summary}".strip())
    return "\n".join(lines)


@app.route("/api/funnels")
@require_funnel_api
def api_funnels_list():
    project = request.args.get("project", "")
    if not project:
        return jsonify({"error": "project required"}), 400
    if not _check_funnel_project(_session_user(), project):
        return jsonify({"error": "forbidden"}), 403
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT cf.id, cf.project, cf.wb_article, cf.seller_article, cf.title,
                           cf.status, cf.design_task_id, cf.created_by, cf.created_by_email,
                           cf.created_at, cf.updated_at, cf.own_data->>'photo_url' AS photo_url,
                           COUNT(fs.id) AS slide_count
                    FROM content_funnels cf
                    LEFT JOIN funnel_slides fs ON fs.funnel_id = cf.id
                    WHERE cf.project=%s
                    GROUP BY cf.id
                    ORDER BY cf.updated_at DESC
                """, (project,))
                return jsonify([serialize_funnel(dict(r)) for r in cur.fetchall()])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/funnels", methods=["POST"])
@require_funnel_api
def api_funnels_create():
    u = _session_user()
    d = request.json or {}
    project = (d.get("project") or "").strip()
    wb_article = d.get("wb_article")
    if not project or not wb_article:
        return jsonify({"error": "project and wb_article required"}), 400
    if not _check_funnel_project(u, project):
        return jsonify({"error": "forbidden"}), 403
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO content_funnels (project, wb_article, seller_article, title, created_by, created_by_email)
                    VALUES (%s,%s,%s,%s,%s,%s) RETURNING id
                """, (project, wb_article, d.get("seller_article", ""), d.get("title", ""),
                      _first_name(u.get("name", "")), u.get("email", "")))
                new_id = cur.fetchone()["id"]
            conn.commit()
        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def _load_funnel(funnel_id: int) -> dict | None:
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM content_funnels WHERE id=%s", (funnel_id,))
            row = cur.fetchone()
    return dict(row) if row else None


@app.route("/api/funnels/own-data")
@require_funnel_api
def api_funnels_own_data():
    project = request.args.get("project", "")
    wb_article = request.args.get("wb_article", "")
    if not project or not wb_article:
        return jsonify({"error": "project and wb_article required"}), 400
    if not _check_funnel_project(_session_user(), project):
        return jsonify({"error": "forbidden"}), 403
    try:
        wb_article_int = int(wb_article)
    except ValueError:
        return jsonify({"error": "invalid wb_article"}), 400
    try:
        return jsonify(_collect_own_data(project, wb_article_int))
    except Exception as e:
        logging.error(f"funnels own-data {project}/{wb_article}: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/funnels/<int:funnel_id>")
@require_funnel_api
def api_funnels_detail(funnel_id):
    u = _session_user()
    funnel = _load_funnel(funnel_id)
    if not funnel:
        return jsonify({"error": "not found"}), 404
    if not _check_funnel_project(u, funnel["project"]):
        return jsonify({"error": "forbidden"}), 403
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT * FROM funnel_slides WHERE funnel_id=%s ORDER BY position ASC, id ASC",
                    (funnel_id,)
                )
                slides = [serialize_slide(dict(r)) for r in cur.fetchall()]

                cur.execute(
                    "SELECT id, name, mime_type, extracted_text, created_by, created_at "
                    "FROM funnel_competitor_files WHERE funnel_id=%s ORDER BY created_at ASC",
                    (funnel_id,)
                )
                files = []
                for r in cur.fetchall():
                    rd = dict(r)
                    if rd.get("created_at"):
                        rd["created_at"] = rd["created_at"].isoformat()
                    rd["has_text"] = bool(rd.pop("extracted_text", ""))
                    files.append(rd)

                cur.execute("""
                    SELECT * FROM funnel_ab_results
                    WHERE project=%s AND wb_article=%s
                    ORDER BY created_at DESC
                """, (funnel["project"], funnel["wb_article"]))
                ab_results = [serialize_ab_result(dict(r)) for r in cur.fetchall()]

        result = serialize_funnel(funnel)
        result["slides"] = slides
        result["competitor_files"] = files
        result["ab_results"] = ab_results
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/funnels/<int:funnel_id>", methods=["PUT"])
@require_funnel_api
def api_funnels_update(funnel_id):
    u = _session_user()
    funnel = _load_funnel(funnel_id)
    if not funnel:
        return jsonify({"error": "not found"}), 404
    if not _check_funnel_project(u, funnel["project"]):
        return jsonify({"error": "forbidden"}), 403
    d = request.json or {}
    title  = d.get("title", funnel.get("title", ""))
    status = d.get("status", funnel.get("status", "draft"))
    slides = d.get("slides")
    own_data = d.get("own_data")
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                if own_data is not None:
                    cur.execute(
                        "UPDATE content_funnels SET title=%s, status=%s, own_data=%s, updated_at=NOW() WHERE id=%s",
                        (title, status, json.dumps(own_data, ensure_ascii=False), funnel_id)
                    )
                else:
                    cur.execute(
                        "UPDATE content_funnels SET title=%s, status=%s, updated_at=NOW() WHERE id=%s",
                        (title, status, funnel_id)
                    )
                if slides is not None:
                    cur.execute("DELETE FROM funnel_slides WHERE funnel_id=%s", (funnel_id,))
                    for i, s in enumerate(slides):
                        cur.execute("""
                            INSERT INTO funnel_slides (funnel_id, position, slide_type, pain_point, message, creative_notes)
                            VALUES (%s,%s,%s,%s,%s,%s)
                        """, (funnel_id, s.get("position", i + 1), s.get("slide_type", ""),
                              s.get("pain_point", ""), s.get("message", ""), s.get("creative_notes", "")))
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/funnels/<int:funnel_id>", methods=["DELETE"])
@require_funnel_api
def api_funnels_delete(funnel_id):
    u = _session_user()
    funnel = _load_funnel(funnel_id)
    if not funnel:
        return jsonify({"error": "not found"}), 404
    if not _check_funnel_project(u, funnel["project"]):
        return jsonify({"error": "forbidden"}), 403
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM content_funnels WHERE id=%s", (funnel_id,))
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Файлы конкурентов ──────────────────────────────────────────────────────
@app.route("/api/funnels/<int:funnel_id>/competitor-files")
@require_funnel_api
def api_funnel_competitor_files_list(funnel_id):
    u = _session_user()
    funnel = _load_funnel(funnel_id)
    if not funnel:
        return jsonify({"error": "not found"}), 404
    if not _check_funnel_project(u, funnel["project"]):
        return jsonify({"error": "forbidden"}), 403
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT id, name, mime_type, extracted_text, created_by, created_at "
                    "FROM funnel_competitor_files WHERE funnel_id=%s ORDER BY created_at ASC",
                    (funnel_id,)
                )
                rows = []
                for r in cur.fetchall():
                    rd = dict(r)
                    if rd.get("created_at"):
                        rd["created_at"] = rd["created_at"].isoformat()
                    rd["has_text"] = bool(rd.pop("extracted_text", ""))
                    rows.append(rd)
        return jsonify(rows)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/funnels/<int:funnel_id>/competitor-files", methods=["POST"])
@require_funnel_api
def api_funnel_competitor_files_add(funnel_id):
    u = _session_user()
    funnel = _load_funnel(funnel_id)
    if not funnel:
        return jsonify({"error": "not found"}), 404
    if not _check_funnel_project(u, funnel["project"]):
        return jsonify({"error": "forbidden"}), 403
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"error": "no file"}), 400
    name = f.filename
    mime_type = f.content_type or "application/octet-stream"
    file_data = f.read()
    if len(file_data) > 20 * 1024 * 1024:
        return jsonify({"error": "file too large (max 20MB)"}), 400
    try:
        from funnel_ai import extract_text_from_upload
        extracted_text = extract_text_from_upload(name, mime_type, file_data)
    except Exception as e:
        logging.error(f"extract_text_from_upload {name}: {e}")
        extracted_text = ""
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO funnel_competitor_files (funnel_id, name, mime_type, file_data, extracted_text, created_by)
                    VALUES (%s,%s,%s,%s,%s,%s) RETURNING id
                """, (funnel_id, name, mime_type, file_data, extracted_text, u.get("name", "")))
                new_id = cur.fetchone()["id"]
            conn.commit()
        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/funnels/competitor-files/<int:file_id>/download")
@require_funnel_api
def api_funnel_competitor_file_download(file_id):
    u = _session_user()
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT cf.name, cf.mime_type, cf.file_data, fn.project
                    FROM funnel_competitor_files cf
                    JOIN content_funnels fn ON fn.id = cf.funnel_id
                    WHERE cf.id=%s
                """, (file_id,))
                row = cur.fetchone()
        if not row or not row["file_data"]:
            return Response("Файл не найден", status=404)
        if not _check_funnel_project(u, row["project"]):
            return Response("Доступ запрещён", status=403)
        raw = bytes(row["file_data"])
        from urllib.parse import quote
        encoded_name = quote(row["name"])
        return Response(
            raw,
            mimetype=row["mime_type"] or "application/octet-stream",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_name}",
                "Content-Length": str(len(raw)),
            }
        )
    except Exception as e:
        return Response(f"Ошибка: {e}", status=500)


@app.route("/api/funnels/competitor-files/<int:file_id>", methods=["DELETE"])
@require_funnel_api
def api_funnel_competitor_file_delete(file_id):
    u = _session_user()
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT fn.project FROM funnel_competitor_files cf
                    JOIN content_funnels fn ON fn.id = cf.funnel_id
                    WHERE cf.id=%s
                """, (file_id,))
                row = cur.fetchone()
                if not row:
                    return jsonify({"error": "not found"}), 404
                if not _check_funnel_project(u, row["project"]):
                    return jsonify({"error": "forbidden"}), 403
                cur.execute("DELETE FROM funnel_competitor_files WHERE id=%s", (file_id,))
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Генерация и отправка дизайнеру ─────────────────────────────────────────
@app.route("/api/funnels/<int:funnel_id>/generate", methods=["POST"])
@require_funnel_api
def api_funnel_generate(funnel_id):
    u = _session_user()
    funnel = _load_funnel(funnel_id)
    if not funnel:
        return jsonify({"error": "not found"}), 404
    if not _check_funnel_project(u, funnel["project"]):
        return jsonify({"error": "forbidden"}), 403
    d = request.json or {}
    notes = (d.get("notes") or "").strip()
    try:
        own_data = _collect_own_data(funnel["project"], funnel["wb_article"])

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT extracted_text FROM funnel_competitor_files "
                    "WHERE funnel_id=%s AND extracted_text != ''",
                    (funnel_id,)
                )
                competitor_text = "\n\n---\n\n".join(r["extracted_text"] for r in cur.fetchall())

                cur.execute("""
                    SELECT variant_name, period_from, period_to, winner, summary
                    FROM funnel_ab_results
                    WHERE project=%s AND wb_article=%s
                    ORDER BY created_at DESC LIMIT 10
                """, (funnel["project"], funnel["wb_article"]))
                ab_rows = [dict(r) for r in cur.fetchall()]

        ab_context = _format_ab_context(ab_rows)

        from funnel_ai import generate_funnel
        slides, raw = generate_funnel(own_data, competitor_text, ab_context, notes)

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    UPDATE content_funnels
                    SET own_data=%s, ab_context=%s, ai_model=%s, ai_raw_response=%s,
                        status='generated', updated_at=NOW()
                    WHERE id=%s
                """, (json.dumps(own_data, ensure_ascii=False), ab_context, raw.get("model", ""),
                      json.dumps(raw, ensure_ascii=False), funnel_id))
                cur.execute("DELETE FROM funnel_slides WHERE funnel_id=%s", (funnel_id,))
                for s in slides:
                    cur.execute("""
                        INSERT INTO funnel_slides (funnel_id, position, slide_type, pain_point, message, creative_notes)
                        VALUES (%s,%s,%s,%s,%s,%s)
                    """, (funnel_id, s["position"], s["slide_type"], s["pain_point"],
                          s["message"], s["creative_notes"]))
            conn.commit()
    except RuntimeError as e:
        return jsonify({"error": str(e)}), 502
    except Exception as e:
        logging.error(f"funnel generate {funnel_id}: {e}")
        return jsonify({"error": str(e)}), 502
    return api_funnels_detail(funnel_id)


def _funnel_brief_inputs(funnel_id):
    funnel = _load_funnel(funnel_id)
    if not funnel:
        return None, None, None
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM funnel_slides WHERE funnel_id=%s ORDER BY position ASC, id ASC",
                (funnel_id,)
            )
            slides = [dict(r) for r in cur.fetchall()]
    return funnel, slides, (funnel.get("own_data") or {})


@app.route("/api/funnels/<int:funnel_id>/brief.pdf")
@require_funnel_api
def api_funnel_brief_pdf(funnel_id):
    u = _session_user()
    funnel, slides, own_data = _funnel_brief_inputs(funnel_id)
    if not funnel:
        return jsonify({"error": "not found"}), 404
    if not _check_funnel_project(u, funnel["project"]):
        return jsonify({"error": "forbidden"}), 403
    try:
        from funnel_ai import render_brief_pdf
        pdf_bytes = render_brief_pdf(funnel, slides, own_data)
    except Exception as e:
        logging.error(f"funnel brief.pdf {funnel_id}: {e}")
        return jsonify({"error": str(e)}), 500
    from urllib.parse import quote
    name = quote(f"funnel_{funnel['wb_article']}_brief.pdf")
    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{name}",
            "Content-Length": str(len(pdf_bytes)),
        }
    )


@app.route("/api/funnels/<int:funnel_id>/brief.docx")
@require_funnel_api
def api_funnel_brief_docx(funnel_id):
    u = _session_user()
    funnel, slides, own_data = _funnel_brief_inputs(funnel_id)
    if not funnel:
        return jsonify({"error": "not found"}), 404
    if not _check_funnel_project(u, funnel["project"]):
        return jsonify({"error": "forbidden"}), 403
    try:
        from funnel_ai import render_brief_docx
        docx_bytes = render_brief_docx(funnel, slides, own_data)
    except Exception as e:
        logging.error(f"funnel brief.docx {funnel_id}: {e}")
        return jsonify({"error": str(e)}), 500
    from urllib.parse import quote
    name = quote(f"funnel_{funnel['wb_article']}_brief.docx")
    return Response(
        docx_bytes,
        mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{name}",
            "Content-Length": str(len(docx_bytes)),
        }
    )


@app.route("/api/funnels/<int:funnel_id>/send-to-design", methods=["POST"])
@require_funnel_api
def api_funnel_send_to_design(funnel_id):
    u = _session_user()
    funnel = _load_funnel(funnel_id)
    if not funnel:
        return jsonify({"error": "not found"}), 404
    if not _check_funnel_project(u, funnel["project"]):
        return jsonify({"error": "forbidden"}), 403
    d = request.json or {}
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT * FROM funnel_slides WHERE funnel_id=%s ORDER BY position ASC, id ASC",
                    (funnel_id,)
                )
                slides = [dict(r) for r in cur.fetchall()]
        if not slides:
            return jsonify({"error": "у воронки нет слайдов — сначала сгенерируйте или заполните вручную"}), 400

        own_data = funnel.get("own_data") or {}
        from funnel_ai import render_brief_html
        brief_html = render_brief_html(funnel, slides, own_data)

        title = f"Контент-воронка: {funnel.get('seller_article') or funnel['wb_article']} ({funnel['project']})"
        assignee_email = (d.get("assignee_email") or "").strip()
        creator_email  = u.get("email", "")

        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO design_tasks "
                    "(title, description, priority, assignee_email, due, project, created_by, created_by_email) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id",
                    (title,
                     d.get("description") or f"Контентная воронка для артикула {funnel['wb_article']} — см. вложение.",
                     d.get("priority", "med"), assignee_email, d.get("due") or None,
                     funnel["project"], _first_name(u.get("name", "")), creator_email)
                )
                task_id = cur.fetchone()["id"]
                cur.execute(
                    "INSERT INTO design_attachments "
                    "(task_id, attach_role, attach_type, name, mime_type, file_data, created_by) "
                    "VALUES (%s,'brief','file',%s,%s,%s,%s)",
                    (task_id, f"funnel_{funnel['wb_article']}_brief.html", "text/html",
                     brief_html.encode("utf-8"), u.get("name", ""))
                )
                if assignee_email and assignee_email != creator_email:
                    _push_notification(
                        cur, assignee_email, "task_assigned",
                        f"Новая задача дизайнеру: {title}",
                        f"Назначена вам в проекте {funnel['project']}",
                        task_id,
                    )
                cur.execute(
                    "UPDATE content_funnels SET design_task_id=%s, status='sent_to_design', updated_at=NOW() WHERE id=%s",
                    (task_id, funnel_id)
                )
            conn.commit()
        return jsonify({"ok": True, "design_task_id": task_id})
    except Exception as e:
        logging.error(f"funnel send-to-design {funnel_id}: {e}")
        return jsonify({"error": str(e)}), 500


# ── А/Б тесты ───────────────────────────────────────────────────────────────
@app.route("/api/funnels/ab-results")
@require_funnel_api
def api_funnel_ab_results_list():
    project = request.args.get("project", "")
    wb_article = request.args.get("wb_article", "")
    if not project or not wb_article:
        return jsonify({"error": "project and wb_article required"}), 400
    if not _check_funnel_project(_session_user(), project):
        return jsonify({"error": "forbidden"}), 403
    try:
        wb_article_int = int(wb_article)
    except ValueError:
        return jsonify({"error": "invalid wb_article"}), 400
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    SELECT * FROM funnel_ab_results
                    WHERE project=%s AND wb_article=%s
                    ORDER BY created_at DESC
                """, (project, wb_article_int))
                return jsonify([serialize_ab_result(dict(r)) for r in cur.fetchall()])
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/funnels/ab-results", methods=["POST"])
@require_funnel_api
def api_funnel_ab_results_create():
    u = _session_user()
    d = request.json or {}
    project = (d.get("project") or "").strip()
    wb_article = d.get("wb_article")
    if not project or not wb_article:
        return jsonify({"error": "project and wb_article required"}), 400
    if not _check_funnel_project(u, project):
        return jsonify({"error": "forbidden"}), 403
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO funnel_ab_results
                        (funnel_id, project, wb_article, variant_name, period_from, period_to,
                         metrics_before, metrics_after, winner, summary, created_by, created_by_email)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id
                """, (d.get("funnel_id") or None, project, wb_article, d.get("variant_name", ""),
                      d.get("period_from") or None, d.get("period_to") or None,
                      json.dumps(d.get("metrics_before") or {}, ensure_ascii=False),
                      json.dumps(d.get("metrics_after") or {}, ensure_ascii=False),
                      d.get("winner", ""), d.get("summary", ""),
                      _first_name(u.get("name", "")), u.get("email", "")))
                new_id = cur.fetchone()["id"]
            conn.commit()
        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/funnels/ab-results/<int:result_id>", methods=["PUT"])
@require_funnel_api
def api_funnel_ab_results_update(result_id):
    u = _session_user()
    d = request.json or {}
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT project FROM funnel_ab_results WHERE id=%s", (result_id,))
                row = cur.fetchone()
                if not row:
                    return jsonify({"error": "not found"}), 404
                if not _check_funnel_project(u, row["project"]):
                    return jsonify({"error": "forbidden"}), 403
                cur.execute("""
                    UPDATE funnel_ab_results SET
                        variant_name=%s, period_from=%s, period_to=%s,
                        metrics_before=%s, metrics_after=%s, winner=%s, summary=%s
                    WHERE id=%s
                """, (d.get("variant_name", ""), d.get("period_from") or None, d.get("period_to") or None,
                      json.dumps(d.get("metrics_before") or {}, ensure_ascii=False),
                      json.dumps(d.get("metrics_after") or {}, ensure_ascii=False),
                      d.get("winner", ""), d.get("summary", ""), result_id))
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/funnels/ab-results/<int:result_id>", methods=["DELETE"])
@require_funnel_api
def api_funnel_ab_results_delete(result_id):
    u = _session_user()
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT project FROM funnel_ab_results WHERE id=%s", (result_id,))
                row = cur.fetchone()
                if not row:
                    return jsonify({"error": "not found"}), 404
                if not _check_funnel_project(u, row["project"]):
                    return jsonify({"error": "forbidden"}), 403
                cur.execute("DELETE FROM funnel_ab_results WHERE id=%s", (result_id,))
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Темп сезона ──────────────────────────────────────────────────────────────
#
# Декомпозиция распродажи сезонного стока: для каждой группы товаров (пресета) со
# своим окном сезона считаем, сколько заказов в день нужно делать, чтобы
# распродать остаток к концу сезона, и сравниваем с текущим темпом.
#
# Все показатели берём из План-факта за ТЕКУЩИЙ месяц (тот же источник, что и вкладка
# «План-Факт»): остаток (= quantityFull = склад + в пути к клиенту + в пути от клиента),
# %выкупа по артикулу, средняя цена, факт заказов за месяц.
#
#   нужно продать  = остаток
#   нужно заказов  = ceil(остаток / %выкупа_из_планфакта)
#   заказов/день   = ceil(нужно заказов / дней до конца сезона)
#   темп сейчас    = заказы за текущий месяц / дней прошло  (из План-факта)
#   отставание     = заказов/день нужно − темп сейчас
#   рублёвые       = соответствующий показатель × средняя цена артикула
#
# Конверсии (CTR/CR) и расход на рекламу будут добавлены позже.


def _season_cabinets(u) -> list:
    """Доступные кабинеты для темпа сезона (пока — все кабинеты WB при наличии доступа)."""
    return list(WB_CABINETS)


def _season_serialize_preset(p: dict, items: list) -> dict:
    return {
        "id":         p["id"],
        "project":    p["project"],
        "name":       p["name"],
        "start_date": p["start_date"].isoformat() if p.get("start_date") else None,
        "end_date":   p["end_date"].isoformat()   if p.get("end_date")   else None,
        "buyout_pct": float(p.get("buyout_pct") or 0),
        "position":   p.get("position") or 0,
        "items": [
            {
                "wb_article":     it["wb_article"],
                "seller_article": it.get("seller_article") or "",
                "name":           it.get("name") or "",
            }
            for it in items
        ],
    }


def _season_load_presets(project: str) -> list:
    """Все пресеты кабинета с их артикулами (без обращения к WB)."""
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM season_presets WHERE project=%s ORDER BY position, id",
                (project,),
            )
            presets = [dict(r) for r in cur.fetchall()]
            if not presets:
                return []
            ids = [p["id"] for p in presets]
            cur.execute(
                "SELECT * FROM season_preset_items WHERE preset_id = ANY(%s) ORDER BY position, id",
                (ids,),
            )
            items_by_preset: dict = {}
            for r in cur.fetchall():
                items_by_preset.setdefault(r["preset_id"], []).append(dict(r))
    return [_season_serialize_preset(p, items_by_preset.get(p["id"], [])) for p in presets]


def _season_build_snapshot(cabinet: str, today: date, force: bool = False) -> dict:
    """
    Собирает снапшот метрик кабинета из План-факта за ТЕКУЩИЙ месяц (источник правды):
    %выкупа, темп заказов/день, средняя цена, остаток (= quantityFull = склад + в пути
    к клиенту + в пути от клиента). Обращается к WB и кладёт результат в `season_cache`.

    force=False — можно переиспользовать прогретый кэш План-факта (для первого
    открытия). force=True (кнопка/крон) — всегда тянет свежие данные из WB.

    Возвращает снапшот {articles, days_elapsed, pace_month, fetched_at}.
    """
    import calendar
    month_start = today.replace(day=1)

    pf_articles = None
    if not force:
        cached, _fetched_at, _complete = _pf_load_cache(cabinet, month_start)
        if cached:
            pf_articles = cached.get("articles", []) or []
    if pf_articles is None:
        from wb_api import collect_planfact_data
        last_day  = calendar.monthrange(month_start.year, month_start.month)[1]
        month_end = month_start.replace(day=last_day)
        d_to = min(today - timedelta(days=1), month_end)
        if d_to < month_start:
            d_to = month_start
        data = collect_planfact_data(cabinet, month_start, d_to)
        _pf_store_cache(cabinet, month_start, data, False)
        pf_articles = data.get("articles", []) or []

    # Дней с фактическими данными в текущем месяце (факт идёт по вчерашний день включ.)
    days_elapsed = max(1, today.day - 1)

    articles = []
    for a in pf_articles:
        articles.append({
            "nm_id":          a.get("nm_id"),
            "vendor_code":    a.get("vendor_code") or "",
            "name":           a.get("title") or "",
            "stock":          int(a.get("stocks") or 0),
            "buyout_pct":     round(float(a.get("buyout_pct") or 0), 1),
            "avg_price":      round(float(a.get("avg_price") or 0)),
            "orders_qty":     int(a.get("orders_qty_fact") or 0),
        })

    snapshot = {
        "articles":     articles,
        "days_elapsed": days_elapsed,
        "pace_month":   today.strftime("%Y-%m"),
        "fetched_at":   datetime.now().isoformat(),
    }
    _season_store_snapshot(cabinet, snapshot)
    return snapshot


def _season_store_snapshot(cabinet: str, snapshot: dict):
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("""
                    INSERT INTO season_cache (project, articles, days_elapsed, pace_month, fetched_at)
                    VALUES (%s, %s::jsonb, %s, %s, NOW())
                    ON CONFLICT (project) DO UPDATE SET
                        articles=EXCLUDED.articles, days_elapsed=EXCLUDED.days_elapsed,
                        pace_month=EXCLUDED.pace_month, fetched_at=NOW()
                """, (cabinet, json.dumps(snapshot["articles"]),
                      snapshot["days_elapsed"], snapshot["pace_month"]))
            conn.commit()
    except Exception:
        logging.exception("season_cache store failed")


def _season_load_snapshot(cabinet: str) -> dict | None:
    """Читает снапшот метрик кабинета из season_cache (без обращения к WB)."""
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT articles, days_elapsed, pace_month, fetched_at "
                    "FROM season_cache WHERE project=%s", (cabinet,)
                )
                row = cur.fetchone()
        if not row:
            return None
        return {
            "articles":     row["articles"] or [],
            "days_elapsed": row["days_elapsed"] or 1,
            "pace_month":   row["pace_month"] or "",
            "fetched_at":   row["fetched_at"].isoformat() if row["fetched_at"] else None,
        }
    except Exception:
        logging.exception("season_cache load failed")
        return None


def _season_index_from_snapshot(snapshot: dict):
    """Строит индексы (by_nm, by_vc) из снапшота для быстрого расчёта декомпозиции."""
    days_elapsed = max(1, int(snapshot.get("days_elapsed") or 1))
    by_nm: dict = {}
    by_vc: dict = {}
    for a in snapshot.get("articles", []):
        rec = {
            "stock":          int(a.get("stock") or 0),
            "buyout_pct":     float(a.get("buyout_pct") or 0),
            "avg_price":      float(a.get("avg_price") or 0),
            "pace":           (int(a.get("orders_qty") or 0) / days_elapsed),
            "seller_article": a.get("vendor_code") or "",
            "name":           a.get("name") or "",
        }
        if a.get("nm_id"):
            by_nm[a["nm_id"]] = rec
        if a.get("vendor_code"):
            by_vc[str(a["vendor_code"])] = rec
    return by_nm, by_vc


def _season_compute(preset: dict, by_nm: dict, by_vc: dict, today: date, overrides: dict | None = None) -> dict:
    """Считает декомпозицию по одному пресету на данных План-факта. Возвращает строки и итоги.
    overrides: {wb_article: buyout_pct} — ручные переопределения %выкупа (перекрывают План-факт)."""
    overrides = overrides or {}
    end = preset.get("end_date")
    start = preset.get("start_date")
    end_date   = date.fromisoformat(end)   if end   else None
    start_date = date.fromisoformat(start) if start else None
    # Считаем от позднейшей из дат «сегодня»/«старт сезона»: если сезон ещё не начался —
    # берём полную длину, если идёт — оставшиеся дни до конца.
    from_day = max(today, start_date) if start_date else today
    days_left = (end_date - from_day).days if end_date else None
    season_over = days_left is not None and days_left <= 0

    rows = []
    tot = {"stock": 0, "need_orders": 0, "need_per_day": 0, "pace": 0.0,
           "need_sales_rub": 0.0, "need_orders_rub": 0.0, "pace_rub": 0.0,
           "need_day_rub": 0.0, "buyout_sum": 0.0, "buyout_cnt": 0}
    for it in preset["items"]:
        nm = it["wb_article"]
        sa = it.get("seller_article") or ""
        rec = by_nm.get(nm) or (by_vc.get(sa) if sa else None) or {}
        stock  = int(rec.get("stock", 0))
        price  = float(rec.get("avg_price", 0) or 0)
        pace   = float(rec.get("pace", 0.0) or 0.0)
        # %выкупа: ручное переопределение перекрывает значение из План-факта
        ov = overrides.get(nm)
        is_override = ov is not None
        buyout = float(ov) if is_override else float(rec.get("buyout_pct", 0) or 0)

        buyout_frac = buyout / 100 if buyout > 0 else 0
        need_sales  = stock
        need_orders = math.ceil(stock / buyout_frac) if buyout_frac > 0 else stock
        if days_left is None:
            need_day = None
        elif season_over:
            need_day = need_orders          # сезон закончился — всё «надо было вчера»
        else:
            need_day = math.ceil(need_orders / days_left) if days_left > 0 else need_orders
        lag = (round(need_day - pace, 1) if need_day is not None else None)

        rows.append({
            "wb_article":      nm,
            "seller_article":  sa or rec.get("seller_article") or str(nm),
            "name":            it.get("name") or rec.get("name") or "",
            "buyout_pct":      round(buyout, 1),
            "buyout_override": is_override,
            "avg_price":       round(price),
            "stock":           stock,
            "need_sales":      need_sales,
            "need_orders":     need_orders,
            "need_per_day":    need_day,
            "pace":            round(pace, 1),
            "lag":             lag,
            # рублёвые
            "need_sales_rub":  round(need_sales * price),
            "need_orders_rub": round(need_orders * price),
            "pace_rub":        round(pace * price),
            "lag_rub":         (round(lag * price) if lag is not None else None),
        })
        tot["stock"]           += stock
        tot["need_orders"]     += need_orders
        tot["need_per_day"]    += (need_day or 0)
        tot["pace"]            += pace
        tot["need_sales_rub"]  += need_sales * price
        tot["need_orders_rub"] += need_orders * price
        tot["pace_rub"]        += pace * price
        tot["need_day_rub"]    += (need_day or 0) * price
        if buyout > 0:
            tot["buyout_sum"]  += buyout
            tot["buyout_cnt"]  += 1

    lag_total = tot["need_per_day"] - tot["pace"]
    avg_buyout = (tot["buyout_sum"] / tot["buyout_cnt"]) if tot["buyout_cnt"] else 0
    return {
        "id":          preset["id"],
        "name":        preset["name"],
        "start_date":  preset.get("start_date"),
        "end_date":    end,
        "days_left":   days_left,
        "season_over": season_over,
        "rows":        rows,
        "totals": {
            "stock":           tot["stock"],
            "buyout_pct":      round(avg_buyout, 1),
            "need_orders":     tot["need_orders"],
            "need_per_day":    tot["need_per_day"],
            "pace":            round(tot["pace"], 1),
            "lag":             round(lag_total, 1),
            "need_sales_rub":  round(tot["need_sales_rub"]),
            "need_orders_rub": round(tot["need_orders_rub"]),
            "pace_rub":        round(tot["pace_rub"]),
            "lag_rub":         round(tot["need_day_rub"] - tot["pace_rub"]),
        },
    }


@app.route("/season")
@require_season_page
def season_page():
    u = _session_user()
    return render_template(
        "season.html",
        current_user=u,
        projects=_season_cabinets(u),
        project_emoji=_project_emoji_dict(),
    )


@app.route("/api/season/articles")
@require_season_api
def api_season_articles():
    """
    Список артикулов кабинета с остатком для пикера пресета. Берём из снапшота метрик
    (season_cache) — быстро и совпадает с таблицей. Если снапшота нет — соберём его.
    """
    project = request.args.get("project", "").strip()
    if not project or project not in WB_CABINETS:
        return jsonify({"error": "project required"}), 400
    try:
        # Только из снапшота, без обращения к WB. Если снапшота ещё нет — вернём пусто
        # (пользователь сначала жмёт «Обновить из WB»).
        snapshot = _season_load_snapshot(project)
        if snapshot is None:
            return jsonify([])
        rows = [
            {"wb_article": a["nm_id"], "seller_article": a.get("vendor_code") or "",
             "stock": int(a.get("stock") or 0)}
            for a in snapshot.get("articles", []) if a.get("nm_id")
        ]
        rows.sort(key=lambda r: r["seller_article"])
        return jsonify(rows)
    except Exception as e:
        logging.error(f"season/articles {project}: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/season/presets")
@require_season_api
def api_season_presets_get():
    """Пресеты кабинета с артикулами (без живых остатков — для редактирования)."""
    project = request.args.get("project", "").strip()
    if not project or project not in WB_CABINETS:
        return jsonify({"error": "project required"}), 400
    return jsonify(_season_load_presets(project))


@app.route("/api/season/presets", methods=["POST"])
@require_season_api
def api_season_presets_create():
    u = _session_user()
    d = request.json or {}
    project = (d.get("project") or "").strip()
    name    = (d.get("name") or "").strip()
    if not project or project not in WB_CABINETS:
        return jsonify({"error": "project required"}), 400
    if not name:
        return jsonify({"error": "name required"}), 400
    start_date = (d.get("start_date") or None) or None
    end_date   = (d.get("end_date") or None) or None
    buyout     = d.get("buyout_pct")
    buyout     = 30 if buyout in (None, "") else buyout
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT COALESCE(MAX(position), -1) + 1 AS pos FROM season_presets WHERE project=%s",
                    (project,),
                )
                pos = cur.fetchone()["pos"]
                cur.execute(
                    "INSERT INTO season_presets (project, name, start_date, end_date, buyout_pct, position, created_by_email) "
                    "VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id",
                    (project, name, start_date, end_date, buyout, pos, u.get("email", "")),
                )
                new_id = cur.fetchone()["id"]
            conn.commit()
        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/season/presets/<int:preset_id>", methods=["PUT"])
@require_season_api
def api_season_presets_update(preset_id):
    d = request.json or {}
    fields, vals = [], []
    if "name" in d:
        fields.append("name=%s");       vals.append((d.get("name") or "").strip())
    if "start_date" in d:
        fields.append("start_date=%s"); vals.append(d.get("start_date") or None)
    if "end_date" in d:
        fields.append("end_date=%s");   vals.append(d.get("end_date") or None)
    if "buyout_pct" in d:
        b = d.get("buyout_pct")
        fields.append("buyout_pct=%s"); vals.append(30 if b in (None, "") else b)
    if not fields:
        return jsonify({"error": "nothing to update"}), 400
    fields.append("updated_at=NOW()")
    vals.append(preset_id)
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    f"UPDATE season_presets SET {', '.join(fields)} WHERE id=%s RETURNING id",
                    tuple(vals),
                )
                if not cur.fetchone():
                    return jsonify({"error": "not found"}), 404
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/season/presets/<int:preset_id>", methods=["DELETE"])
@require_season_api
def api_season_presets_delete(preset_id):
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM season_presets WHERE id=%s RETURNING id", (preset_id,))
                if not cur.fetchone():
                    return jsonify({"error": "not found"}), 404
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/season/presets/<int:preset_id>/items", methods=["PUT"])
@require_season_api
def api_season_presets_set_items(preset_id):
    """Полная замена набора артикулов в пресете."""
    d = request.json or {}
    items = d.get("items") or []
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT id FROM season_presets WHERE id=%s", (preset_id,))
                if not cur.fetchone():
                    return jsonify({"error": "not found"}), 404
                cur.execute("DELETE FROM season_preset_items WHERE preset_id=%s", (preset_id,))
                for pos, it in enumerate(items):
                    nm = it.get("wb_article")
                    if not nm:
                        continue
                    cur.execute(
                        "INSERT INTO season_preset_items (preset_id, wb_article, seller_article, name, position) "
                        "VALUES (%s,%s,%s,%s,%s)",
                        (preset_id, nm, it.get("seller_article", ""), it.get("name", ""), pos),
                    )
                cur.execute("UPDATE season_presets SET updated_at=NOW() WHERE id=%s", (preset_id,))
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/season/buyout", methods=["POST"])
@require_season_api
def api_season_set_buyout():
    """Ручное переопределение %выкупа по артикулу. Пустое/None значение — сбросить (вернуть из План-факта)."""
    d = request.json or {}
    project = (d.get("project") or "").strip()
    nm = d.get("wb_article")
    if not project or project not in WB_CABINETS or not nm:
        return jsonify({"error": "project and wb_article required"}), 400
    raw = d.get("buyout_pct")
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                if raw in (None, "", False):
                    cur.execute(
                        "DELETE FROM season_buyout_overrides WHERE project=%s AND wb_article=%s",
                        (project, nm),
                    )
                    cleared = True
                else:
                    val = max(0, min(100, float(raw)))
                    cur.execute("""
                        INSERT INTO season_buyout_overrides (project, wb_article, buyout_pct, updated_at)
                        VALUES (%s, %s, %s, NOW())
                        ON CONFLICT (project, wb_article) DO UPDATE SET
                            buyout_pct=EXCLUDED.buyout_pct, updated_at=NOW()
                    """, (project, nm, val))
                    cleared = False
            conn.commit()
        return jsonify({"ok": True, "cleared": cleared})
    except (ValueError, TypeError):
        return jsonify({"error": "invalid buyout_pct"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


def _season_load_overrides(project: str) -> dict:
    """Ручные переопределения %выкупа кабинета: {wb_article: buyout_pct}."""
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT wb_article, buyout_pct FROM season_buyout_overrides WHERE project=%s",
                    (project,),
                )
                return {r["wb_article"]: float(r["buyout_pct"]) for r in cur.fetchall()}
    except Exception:
        logging.exception("season overrides load failed")
        return {}


def _season_decomp_payload(project: str, snapshot: dict, today: date) -> dict:
    """Считает декомпозицию по всем группам кабинета поверх снапшота метрик."""
    presets = _season_load_presets(project)
    by_nm, by_vc = _season_index_from_snapshot(snapshot)
    overrides = _season_load_overrides(project)
    computed = [_season_compute(p, by_nm, by_vc, today, overrides) for p in presets]
    return {
        "presets":    computed,
        "pace_days":  snapshot.get("days_elapsed", 1),
        "pace_month": snapshot.get("pace_month", ""),
        "fetched_at": snapshot.get("fetched_at"),
    }


@app.route("/api/season/decomposition")
@require_season_api
def api_season_decomposition():
    """
    Декомпозиция по всем группам кабинета. При обычном открытии — мгновенно из снапшота
    метрик в БД (`season_cache`), без обращения к WB. С ?refresh=1 (кнопка «Обновить из
    WB») — тянет свежие данные из План-факта и обновляет снапшот. Если снапшота ещё нет
    — собирает его один раз.
    """
    project = request.args.get("project", "").strip()
    if not project or project not in WB_CABINETS:
        return jsonify({"error": "project required"}), 400
    force = request.args.get("refresh", "") in ("1", "true", "yes")
    try:
        today = date.today()
        if force:
            # Явное обновление (кнопка) — единственный путь, который ходит в WB.
            snapshot = _season_build_snapshot(project, today, force=True)
        else:
            # Обычное открытие — ТОЛЬКО из БД, без обращения к WB (иначе медленный
            # запрос блокирует воркеры gunicorn и роняет весь дашборд).
            snapshot = _season_load_snapshot(project)
        if snapshot is None:
            # Снапшота ещё нет — отдаём структуру групп с нулевыми метриками и флагом,
            # чтобы фронт показал «нажмите Обновить из WB». WB не трогаем.
            snapshot = {"articles": [], "days_elapsed": 1, "pace_month": "", "fetched_at": None}
            payload = _season_decomp_payload(project, snapshot, today)
            payload["no_snapshot"] = True
            payload["from_cache"]  = True
            return jsonify(payload)
        payload = _season_decomp_payload(project, snapshot, today)
        payload["from_cache"] = not force
        return jsonify(payload)
    except Exception as e:
        logging.error(f"season/decomposition {project}: {e}")
        return jsonify({"error": str(e)}), 500


@app.route("/api/season/cron-refresh", methods=["GET", "POST"])
def api_season_cron_refresh():
    """
    Утреннее обновление снапшотов по ВСЕМ кабинетам (для Railway cron). Защита — токен
    в env `SEASON_CRON_TOKEN` (передаётся как ?token= или заголовок X-Cron-Token).
    Если токен не задан в окружении — endpoint выключен (503).
    """
    expected = os.environ.get("SEASON_CRON_TOKEN", "")
    if not expected:
        return jsonify({"error": "cron disabled (no SEASON_CRON_TOKEN)"}), 503
    got = request.args.get("token", "") or request.headers.get("X-Cron-Token", "")
    if got != expected:
        return jsonify({"error": "forbidden"}), 403
    # Опционально можно обновлять по одному кабинету (?cabinet=), чтобы не упереться в
    # timeout gunicorn при большом числе кабинетов — крон дёргает их по очереди.
    one = request.args.get("cabinet", "").strip()
    cabinets = [one] if one else list(WB_CABINETS)
    if one and one not in WB_CABINETS:
        return jsonify({"error": "unknown cabinet"}), 400
    today = date.today()
    refreshed, errors = [], {}
    for cabinet in cabinets:
        try:
            snap = _season_build_snapshot(cabinet, today, force=True)
            refreshed.append({"cabinet": cabinet, "articles": len(snap["articles"])})
        except Exception as e:
            logging.error(f"season cron-refresh {cabinet}: {e}")
            errors[cabinet] = str(e)
    return jsonify({"ok": True, "refreshed": refreshed, "errors": errors,
                    "at": datetime.now().isoformat()})


# ═══════════════════════════════════════════════════════════════════════════════
#  КОКПИТ АГЕНТСТВА (/cockpit) — только владелец (admin)
#  Аналитика бизнеса: воронка агентства (KPI), денежные потоки, цели, задачи,
#  сетка планов по кабинетам, команда/ЗП/мотивация.
# ═══════════════════════════════════════════════════════════════════════════════

# Дефолтные показатели воронки агентства (как в Google-таблице): группа → метрики
COCKPIT_KPI_DEFAULTS = [
    ("Привлечение", "Рассылки в личку",       300, 0),
    ("Привлечение", "Рекламное объявление",     4, 20000),
    ("Продажи",     "Обработано лидов",         30, 0),
    ("Продажи",     "Разовые услуги",            2, 10000),
    ("Продажи",     "Продажа (договор)",         3, 150000),
    ("Предоставление", "Аудиты",                 6, 0),
    ("Предоставление", "Стратегия на месяц",     3, 0),
    ("Предоставление", "Ведение кабинета",       3, 0),
]


def _cockpit_month(arg: str | None) -> date:
    """'YYYY-MM' → первое число месяца. Пусто/битое → текущий месяц."""
    today = date.today()
    if arg:
        try:
            y, m = arg.split("-")[:2]
            return date(int(y), int(m), 1)
        except Exception:
            pass
    return today.replace(day=1)


def _cockpit_seed_kpi(month_date: date):
    """Если за месяц ещё нет показателей — засеять дефолтной воронкой."""
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT COUNT(*) AS n FROM agency_kpi WHERE month=%s", (month_date,))
            if (cur.fetchone()["n"] or 0) > 0:
                return
            for i, (grp, metric, plan, price) in enumerate(COCKPIT_KPI_DEFAULTS):
                cur.execute(
                    "INSERT INTO agency_kpi (month,grp,metric,plan_qty,fact_qty,unit_price,position) "
                    "VALUES (%s,%s,%s,%s,0,%s,%s) ON CONFLICT (month,grp,metric) DO NOTHING",
                    (month_date, grp, metric, plan, price, i),
                )
        conn.commit()


def _cockpit_days(month_date: date) -> tuple[int, int, int]:
    """(дней в месяце, прошло, осталось) — прошло не больше, чем весь месяц."""
    import calendar as _cal
    dim = _cal.monthrange(month_date.year, month_date.month)[1]
    today = date.today()
    if today < month_date:
        elapsed = 0
    elif today.year == month_date.year and today.month == month_date.month:
        elapsed = today.day
    else:
        elapsed = dim
    return dim, elapsed, max(0, dim - elapsed)


def _week_start(d: date) -> date:
    """Понедельник недели, в которую попадает дата."""
    return d - timedelta(days=d.weekday())


def _month_weeks(month_date: date) -> list:
    """Список понедельников недель, пересекающих месяц (для навигатора РНП)."""
    import calendar
    dim = calendar.monthrange(month_date.year, month_date.month)[1]
    last = month_date.replace(day=dim)
    weeks, ws = [], _week_start(month_date)
    while ws <= last:
        weeks.append(ws)
        ws = ws + timedelta(days=7)
    return weeks


def _cockpit_daily_by_kpi(kpi_ids: list) -> dict:
    """{kpi_id: сумма всех дневных отметок} — факт месяца из РНП."""
    if not kpi_ids:
        return {}
    out = {}
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT kpi_id, COALESCE(SUM(qty),0) AS s FROM agency_kpi_daily "
                    "WHERE kpi_id = ANY(%s) GROUP BY kpi_id",
                    (kpi_ids,),
                )
                for r in cur.fetchall():
                    out[r["kpi_id"]] = float(r["s"] or 0)
    except Exception as e:
        logging.error(f"cockpit daily_by_kpi: {e}")
    return out


def _cockpit_cabinet_facts(month_date: date) -> dict:
    """
    По каждому WB-кабинету: факт продаж (₽) — из planfact_cache (снимок факта),
    план продаж (₽) — из таблицы sales_plans (план в кэше не хранится, он
    накладывается поверх при отдаче План-Факта). Без обращения к WB.
    """
    # План продаж по кабинетам за месяц — тот же источник, что и вкладка План-Факт
    plans = {}
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT project, COALESCE(SUM(sales_rub_plan),0) AS splan, "
                    "COALESCE(SUM(orders_rub_plan),0) AS oplan "
                    "FROM sales_plans WHERE month=%s GROUP BY project",
                    (month_date,),
                )
                for r in cur.fetchall():
                    plans[r["project"]] = {
                        "sales_rub_plan": float(r["splan"] or 0),
                        "orders_rub_plan": float(r["oplan"] or 0),
                    }
    except Exception as e:
        logging.error(f"cockpit cabinet plans: {e}")

    out = {}
    for cab in WB_CABINETS:
        data, _fetched, _complete = _pf_load_cache(cab, month_date)
        t = (data or {}).get("totals", {}) if data else {}
        orders_fact = float(t.get("orders_rub_fact") or 0)
        orders_plan = plans.get(cab, {}).get("orders_rub_plan", 0.0)
        pct = round(orders_fact / orders_plan * 100, 1) if orders_plan else None
        out[cab] = {
            "orders_rub_fact": orders_fact,
            "orders_rub_plan": orders_plan,
            "pct": pct,
        }
    return out


def _cockpit_task_stats(month_date: date) -> dict:
    """Операционная загрузка за месяц из общей таблицы tasks (created_at в месяце)."""
    if month_date.month == 12:
        month_end = date(month_date.year + 1, 1, 1)
    else:
        month_end = date(month_date.year, month_date.month + 1, 1)
    stats = {"total": 0, "done": 0, "in_work": 0, "by_project": {}, "by_priority": {}}
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "SELECT project, priority, done FROM tasks "
                    "WHERE created_at >= %s AND created_at < %s",
                    (month_date, month_end),
                )
                for r in cur.fetchall():
                    stats["total"] += 1
                    if r["done"]:
                        stats["done"] += 1
                    else:
                        stats["in_work"] += 1
                    proj = r["project"] or "Без проекта"
                    prio = r["priority"] or "med"
                    stats["by_project"][proj] = stats["by_project"].get(proj, 0) + 1
                    stats["by_priority"][prio] = stats["by_priority"].get(prio, 0) + 1
    except Exception as e:
        logging.error(f"cockpit task_stats: {e}")
    return stats


def _cockpit_maybe_backup():
    """Раз в 7 дней — снимок всех финтаблиц кокпита в cockpit_backups (JSONB)."""
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute("SELECT MAX(backup_date) AS last FROM cockpit_backups")
                last = cur.fetchone()["last"]
                if last and (date.today() - last).days < 7:
                    return
                snap = {}
                for tbl in ("agency_kpi", "agency_kpi_daily", "agency_kpi_week",
                            "cashflow", "clients", "team_members", "team_cabinets",
                            "cockpit_goals"):
                    cur.execute(f"SELECT * FROM {tbl}")
                    rows = []
                    for r in cur.fetchall():
                        d = {}
                        for k, v in dict(r).items():
                            if isinstance(v, (date, datetime)):
                                d[k] = v.isoformat()
                            elif isinstance(v, Decimal):
                                d[k] = float(v)
                            else:
                                d[k] = v
                        rows.append(d)
                    snap[tbl] = rows
                cur.execute(
                    "INSERT INTO cockpit_backups (backup_date, payload) "
                    "VALUES (CURRENT_DATE, %s::jsonb) ON CONFLICT (backup_date) DO NOTHING",
                    (json.dumps(snap),),
                )
            conn.commit()
        logging.info("Cockpit weekly backup saved.")
    except Exception:
        logging.exception("cockpit backup failed")


@app.route("/cockpit")
@require_admin
def cockpit_page():
    u = _session_user()
    return render_template("cockpit.html", current_user=u, cabinets=WB_CABINETS)


@app.route("/api/cockpit/overview")
@require_admin_api
def api_cockpit_overview():
    month_date = _cockpit_month(request.args.get("month"))
    _cockpit_seed_kpi(month_date)
    _cockpit_maybe_backup()
    dim, elapsed, remaining = _cockpit_days(month_date)

    # ── Воронка агентства (KPI) ─────────────────────────────────────────────
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM agency_kpi WHERE month=%s ORDER BY position, id", (month_date,)
            )
            kpi_rows = cur.fetchall()
    daily_fact = _cockpit_daily_by_kpi([r["id"] for r in kpi_rows])
    kpi = []
    goal_money_plan = goal_money_fact = 0.0
    leads_plan = leads_fact = sales_plan_q = sales_fact_q = 0
    for r in kpi_rows:
        plan_q = int(r["plan_qty"] or 0)
        # Факт месяца — из ежедневных отметок РНП (agency_kpi_daily), не из fact_qty
        fact_q = daily_fact.get(r["id"], float(r["fact_qty"] or 0))
        fact_q = int(round(fact_q))
        price = float(r["unit_price"] or 0)
        should = round(plan_q * elapsed / dim) if dim else 0
        run_rate = round(fact_q / elapsed * dim) if elapsed else 0
        pct = round(fact_q / plan_q * 100) if plan_q else None
        kpi.append({
            "id": r["id"], "grp": r["grp"], "metric": r["metric"],
            "plan_qty": plan_q, "fact_qty": fact_q, "unit_price": price,
            "money_plan": plan_q * price, "money_fact": fact_q * price,
            "should_be": should, "run_rate": run_rate, "pct": pct,
        })
        goal_money_plan += plan_q * price
        goal_money_fact += fact_q * price
        ml = r["metric"].lower()
        if "лид" in ml:
            leads_plan += plan_q; leads_fact += fact_q
        if "договор" in ml or "разов" in ml:
            sales_plan_q += plan_q; sales_fact_q += fact_q

    # ── Денежные потоки за месяц ─────────────────────────────────────────────
    if month_date.month == 12:
        month_end = date(month_date.year + 1, 1, 1)
    else:
        month_end = date(month_date.year, month_date.month + 1, 1)
    income = expense = 0.0
    cf_by_day = {}
    cf_rows = []
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT * FROM cashflow WHERE dt >= %s AND dt < %s ORDER BY dt DESC, id DESC",
                (month_date, month_end),
            )
            cf_rows = cur.fetchall()
    n_income_ops = 0
    for r in cf_rows:
        amt = float(r["amount"] or 0)
        key = r["dt"].isoformat()
        d = cf_by_day.setdefault(key, {"income": 0.0, "expense": 0.0})
        if r["direction"] == "income":
            income += amt; d["income"] += amt; n_income_ops += 1
        else:
            expense += amt; d["expense"] += amt
    avg_check = round(income / n_income_ops) if n_income_ops else 0

    # ── Сетка планов по кабинетам + команда/ЗП ──────────────────────────────
    cab_facts = _cockpit_cabinet_facts(month_date)
    cabinet_grid = [
        {"cabinet": cab, **vals} for cab, vals in cab_facts.items()
    ]
    cabinet_grid.sort(key=lambda x: -(x["orders_rub_fact"] or 0))

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM team_members ORDER BY position, id")
            members = cur.fetchall()
            cur.execute("SELECT * FROM team_cabinets ORDER BY id")
            tc_rows = cur.fetchall()
    tc_by_member = {}
    for r in tc_rows:
        tc_by_member.setdefault(r["member_id"], []).append(r)
    team = []
    payroll = 0.0
    for m in members:
        base = float(m["salary_base"] or 0)
        cabs = []
        bonus = 0.0
        for tc in tc_by_member.get(m["id"], []):
            fact = float(cab_facts.get(tc["cabinet"], {}).get("orders_rub_fact") or 0)
            plan = float(tc["plan_rub"] or 0)
            bp = float(tc["bonus_pct"] or 0)
            b = round(fact * bp / 100)
            bonus += b
            cabs.append({
                "id": tc["id"], "cabinet": tc["cabinet"], "plan_rub": plan,
                "fact_rub": fact, "bonus_pct": bp, "bonus_rub": b,
                "pct": round(fact / plan * 100, 1) if plan else None,
            })
        total = round(base + bonus)
        payroll += total
        team.append({
            "id": m["id"], "name": m["name"], "email": m["email"],
            "role_title": m["role_title"], "salary_base": base, "active": bool(m["active"]),
            "cabinets": cabs, "bonus_rub": round(bonus), "salary_total": total,
        })

    # ── Клиенты + Финмодель (P&L / РНП) ─────────────────────────────────────
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM clients ORDER BY active DESC, position, id")
            client_rows = cur.fetchall()
    clients = []
    recurring_rev = 0.0
    n_active = 0
    for c in client_rows:
        mrr = float(c["mrr"] or 0)
        act = bool(c["active"])
        if act:
            recurring_rev += mrr
            n_active += 1
        clients.append({
            "id": c["id"], "name": c["name"], "cabinet": c["cabinet"],
            "service": c["service"], "mrr": mrr, "active": act,
            "start_date": c["start_date"].isoformat() if c["start_date"] else "",
            "note": c["note"],
        })

    # Выручка = ведение (клиенты) + прочие поступления (доходы cashflow, сгруппированы)
    income_by_cat = {}
    expense_by_cat = {}
    for r in cf_rows:
        amt = float(r["amount"] or 0)
        cat = (r["category"] or "").strip() or "Прочее"
        if r["direction"] == "income":
            income_by_cat[cat] = income_by_cat.get(cat, 0.0) + amt
        else:
            expense_by_cat[cat] = expense_by_cat.get(cat, 0.0) + amt

    # Итоги считаем ТОЛЬКО по факту операций (нарастающий итог за месяц).
    # Клиентский доход (recurring) и ФОТ (payroll) НЕ подставляем автоматически —
    # они справочные (план ведения / плановый ФОТ). Всё фактическое движение денег
    # владелец пишет вручную в операциях.
    revenue = income
    costs_total = expense
    profit = revenue - costs_total
    avg_check = round(recurring_rev / n_active) if n_active else 0

    revenue_lines = [
        {"category": cat, "amount": round(amt),
         "pct": round(amt / revenue * 100, 1) if revenue else None, "count": None}
        for cat, amt in sorted(income_by_cat.items(), key=lambda x: -x[1])
    ]
    cost_lines = [
        {"category": cat, "amount": round(amt),
         "pct": round(amt / revenue * 100, 1) if revenue else None, "auto": False}
        for cat, amt in sorted(expense_by_cat.items(), key=lambda x: -x[1])
    ]

    finance = {
        "revenue": round(revenue), "costs_total": round(costs_total),
        "profit": round(profit),
        "margin": round(profit / revenue * 100, 1) if revenue else None,
        "revenue_lines": revenue_lines, "cost_lines": cost_lines,
        # справочные (в итоги не входят)
        "recurring_plan": round(recurring_rev), "n_clients": n_active,
        "avg_check": avg_check, "payroll_plan": round(payroll),
    }

    # ── Ручные финансовые цели (оборот/прибыль); факт — из финмодели ─────────
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT revenue_goal, profit_goal FROM cockpit_goals WHERE month=%s", (month_date,))
            grow = cur.fetchone()
    rev_goal = float(grow["revenue_goal"]) if grow else 0.0
    prof_goal = float(grow["profit_goal"]) if grow else 0.0
    fin_goals = {
        "revenue_goal": round(rev_goal), "revenue_fact": round(revenue),
        "revenue_pct": round(revenue / rev_goal * 100, 1) if rev_goal else None,
        "profit_goal": round(prof_goal), "profit_fact": round(profit),
        "profit_pct": round(profit / prof_goal * 100, 1) if prof_goal else None,
    }

    return jsonify({
        "month": month_date.isoformat(),
        "days": {"in_month": dim, "elapsed": elapsed, "remaining": remaining},
        "clients": clients,
        "finance": finance,
        "fin_goals": fin_goals,
        "kpi": kpi,
        "goals": {
            "money_plan": round(goal_money_plan), "money_fact": round(goal_money_fact),
            "money_pct": round(goal_money_fact / goal_money_plan * 100, 1) if goal_money_plan else None,
            "leads_plan": leads_plan, "leads_fact": leads_fact,
            "sales_plan": sales_plan_q, "sales_fact": sales_fact_q,
        },
        "cashflow": {
            "income": round(income), "expense": round(expense),
            "net": round(income - expense), "avg_check": avg_check,
            "by_day": cf_by_day,
            "rows": [{
                "id": r["id"], "dt": r["dt"].isoformat(),
                "direction": r["direction"], "category": r["category"],
                "project": r["project"], "amount": float(r["amount"] or 0),
                "note": r["note"],
            } for r in cf_rows],
        },
        "cabinet_grid": cabinet_grid,
        "tasks": _cockpit_task_stats(month_date),
        "team": team,
        "payroll": round(payroll),
    })


@app.route("/api/cockpit/kpi", methods=["POST"])
@require_admin_api
def api_cockpit_kpi_save():
    d = request.json or {}
    month_date = _cockpit_month(d.get("month"))
    kid = d.get("id")
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                if kid:
                    cur.execute(
                        "UPDATE agency_kpi SET grp=%s,metric=%s,plan_qty=%s,fact_qty=%s,"
                        "unit_price=%s,updated_at=NOW() WHERE id=%s",
                        (d.get("grp", "Привлечение"), d.get("metric", ""),
                         int(d.get("plan_qty") or 0), int(d.get("fact_qty") or 0),
                         float(d.get("unit_price") or 0), kid),
                    )
                else:
                    cur.execute("SELECT COALESCE(MAX(position),0)+1 AS p FROM agency_kpi WHERE month=%s", (month_date,))
                    pos = cur.fetchone()["p"]
                    cur.execute(
                        "INSERT INTO agency_kpi (month,grp,metric,plan_qty,fact_qty,unit_price,position) "
                        "VALUES (%s,%s,%s,%s,%s,%s,%s) "
                        "ON CONFLICT (month,grp,metric) DO UPDATE SET "
                        "plan_qty=EXCLUDED.plan_qty,fact_qty=EXCLUDED.fact_qty,"
                        "unit_price=EXCLUDED.unit_price,updated_at=NOW()",
                        (month_date, d.get("grp", "Привлечение"), d.get("metric", ""),
                         int(d.get("plan_qty") or 0), int(d.get("fact_qty") or 0),
                         float(d.get("unit_price") or 0), pos),
                    )
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/cockpit/kpi/<int:kid>", methods=["DELETE"])
@require_admin_api
def api_cockpit_kpi_del(kid):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM agency_kpi WHERE id=%s", (kid,))
        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/cockpit/rnp")
@require_admin_api
def api_cockpit_rnp():
    """РНП: месячная сводка показателей + выбранная неделя (план недели, отметки
    по дням Пн–Вс, факт недели, %) + деньги по дням для графика."""
    month_date = _cockpit_month(request.args.get("month"))
    _cockpit_seed_kpi(month_date)
    dim, elapsed, remaining = _cockpit_days(month_date)
    weeks = _month_weeks(month_date)
    sel = None
    if request.args.get("week"):
        try:
            sel = datetime.strptime(request.args["week"], "%Y-%m-%d").date()
        except Exception:
            sel = None
    if sel is None:
        today = date.today()
        sel = _week_start(today) if _week_start(today) in weeks else weeks[0]
    days = [sel + timedelta(days=i) for i in range(7)]
    day_iso = [d.isoformat() for d in days]

    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM agency_kpi WHERE month=%s ORDER BY position, id", (month_date,))
            kpi_rows = cur.fetchall()
            ids = [r["id"] for r in kpi_rows]
            daily, fact_before = {}, {}
            if ids:
                cur.execute(
                    "SELECT kpi_id, day, qty FROM agency_kpi_daily "
                    "WHERE kpi_id = ANY(%s) AND day >= %s AND day < %s",
                    (ids, sel, sel + timedelta(days=7)),
                )
                for r in cur.fetchall():
                    daily[(r["kpi_id"], r["day"].isoformat())] = float(r["qty"] or 0)
                # Факт, накопленный ДО текущей недели (для переноса долга в недельный план)
                cur.execute(
                    "SELECT kpi_id, COALESCE(SUM(qty),0) AS s FROM agency_kpi_daily "
                    "WHERE kpi_id = ANY(%s) AND day >= %s AND day < %s GROUP BY kpi_id",
                    (ids, month_date, sel),
                )
                for r in cur.fetchall():
                    fact_before[r["kpi_id"]] = float(r["s"] or 0)
    daily_month = _cockpit_daily_by_kpi(ids)

    # Недельный план считается автоматом из месячного: цель к концу текущей недели
    # (пропорционально дням месяца) минус уже сделанное до недели. Так недовыполнение
    # накапливается в план следующей недели (+), перевыполнение — уменьшает (−, до 0).
    month_end = month_date.replace(day=dim)
    end_ref = min(days[6], month_end)
    cum_days = (end_ref - month_date).days + 1 if end_ref >= month_date else 0
    cum_days = max(0, min(cum_days, dim))

    metrics = []
    daily_money = [0.0] * 7
    for r in kpi_rows:
        price = float(r["unit_price"] or 0)
        cells = [daily.get((r["id"], di), 0.0) for di in day_iso]
        for i, c in enumerate(cells):
            daily_money[i] += c * price
        wfact = sum(cells)
        mplan = int(r["plan_qty"] or 0)
        target_through = mplan * cum_days / dim if dim else 0
        wp = max(0, round(target_through - fact_before.get(r["id"], 0.0)))
        mfact = daily_month.get(r["id"], float(r["fact_qty"] or 0))
        metrics.append({
            "id": r["id"], "grp": r["grp"], "metric": r["metric"], "unit_price": price,
            "month_plan": mplan, "month_fact": round(mfact),
            "month_pct": round(mfact / mplan * 100) if mplan else None,
            "should": round(mplan * elapsed / dim) if dim else 0,
            "run_rate": round(mfact / elapsed * dim) if elapsed else 0,
            "week_plan": wp, "daily": cells, "week_fact": wfact,
            "week_pct": round(wfact / wp * 100) if wp else None,
        })

    return jsonify({
        "month": month_date.isoformat(),
        "week_start": sel.isoformat(), "week_end": days[6].isoformat(),
        "days": day_iso,
        "day_labels": ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"],
        "day_nums": [d.day for d in days],
        "in_month": [d.month == month_date.month for d in days],
        "weeks": [w.isoformat() for w in weeks],
        "prev_week": (sel - timedelta(days=7)).isoformat(),
        "next_week": (sel + timedelta(days=7)).isoformat(),
        "metrics": metrics,
        "daily_money": [round(x) for x in daily_money],
    })


@app.route("/api/cockpit/rnp/daily", methods=["POST"])
@require_admin_api
def api_cockpit_rnp_daily():
    d = request.json or {}
    if not d.get("kpi_id") or not d.get("day"):
        return jsonify({"error": "kpi_id and day required"}), 400
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO agency_kpi_daily (kpi_id, day, qty) VALUES (%s,%s,%s) "
                    "ON CONFLICT (kpi_id, day) DO UPDATE SET qty=EXCLUDED.qty",
                    (int(d["kpi_id"]), d["day"], float(d.get("qty") or 0)),
                )
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/cockpit/cashflow", methods=["POST"])
@require_admin_api
def api_cockpit_cashflow_add():
    d = request.json or {}
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO cashflow (dt,direction,category,project,amount,note) "
                    "VALUES (%s,%s,%s,%s,%s,%s)",
                    (d.get("dt") or date.today().isoformat(),
                     d.get("direction", "income"), d.get("category", ""),
                     d.get("project", ""), float(d.get("amount") or 0), d.get("note", "")),
                )
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/cockpit/cashflow/<int:cid>", methods=["DELETE"])
@require_admin_api
def api_cockpit_cashflow_del(cid):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM cashflow WHERE id=%s", (cid,))
        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/cockpit/team", methods=["POST"])
@require_admin_api
def api_cockpit_team_save():
    d = request.json or {}
    mid = d.get("id")
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                if mid:
                    cur.execute(
                        "UPDATE team_members SET name=%s,email=%s,role_title=%s,"
                        "salary_base=%s,active=%s WHERE id=%s",
                        (d.get("name", ""), d.get("email", ""), d.get("role_title", ""),
                         float(d.get("salary_base") or 0), bool(d.get("active", True)), mid),
                    )
                else:
                    cur.execute(
                        "INSERT INTO team_members (name,email,role_title,salary_base) "
                        "VALUES (%s,%s,%s,%s) RETURNING id",
                        (d.get("name", ""), d.get("email", ""), d.get("role_title", ""),
                         float(d.get("salary_base") or 0)),
                    )
                    mid = cur.fetchone()["id"]
                # Полная замена привязок к кабинетам, если переданы
                if "cabinets" in d:
                    cur.execute("DELETE FROM team_cabinets WHERE member_id=%s", (mid,))
                    for c in d.get("cabinets") or []:
                        if not c.get("cabinet"):
                            continue
                        cur.execute(
                            "INSERT INTO team_cabinets (member_id,cabinet,plan_rub,bonus_pct) "
                            "VALUES (%s,%s,%s,%s) ON CONFLICT (member_id,cabinet) DO UPDATE SET "
                            "plan_rub=EXCLUDED.plan_rub,bonus_pct=EXCLUDED.bonus_pct",
                            (mid, c["cabinet"], float(c.get("plan_rub") or 0),
                             float(c.get("bonus_pct") or 0)),
                        )
            conn.commit()
        return jsonify({"ok": True, "id": mid})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/cockpit/team/<int:mid>", methods=["DELETE"])
@require_admin_api
def api_cockpit_team_del(mid):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM team_members WHERE id=%s", (mid,))
        conn.commit()
    return jsonify({"ok": True})


@app.route("/api/cockpit/clients", methods=["POST"])
@require_admin_api
def api_cockpit_client_save():
    d = request.json or {}
    cid = d.get("id")
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                if cid:
                    cur.execute(
                        "UPDATE clients SET name=%s,cabinet=%s,service=%s,mrr=%s,"
                        "start_date=%s,active=%s,note=%s WHERE id=%s",
                        (d.get("name", ""), d.get("cabinet", ""), d.get("service", "Ведение"),
                         float(d.get("mrr") or 0), d.get("start_date") or None,
                         bool(d.get("active", True)), d.get("note", ""), cid),
                    )
                else:
                    cur.execute(
                        "INSERT INTO clients (name,cabinet,service,mrr,start_date,active,note) "
                        "VALUES (%s,%s,%s,%s,%s,%s,%s)",
                        (d.get("name", ""), d.get("cabinet", ""), d.get("service", "Ведение"),
                         float(d.get("mrr") or 0), d.get("start_date") or None,
                         bool(d.get("active", True)), d.get("note", "")),
                    )
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/cockpit/goals", methods=["POST"])
@require_admin_api
def api_cockpit_goals_save():
    d = request.json or {}
    month_date = _cockpit_month(d.get("month"))
    try:
        with get_conn() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    "INSERT INTO cockpit_goals (month, revenue_goal, profit_goal, updated_at) "
                    "VALUES (%s,%s,%s,NOW()) ON CONFLICT (month) DO UPDATE SET "
                    "revenue_goal=EXCLUDED.revenue_goal, profit_goal=EXCLUDED.profit_goal, updated_at=NOW()",
                    (month_date, float(d.get("revenue_goal") or 0), float(d.get("profit_goal") or 0)),
                )
            conn.commit()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/cockpit/clients/<int:cid>", methods=["DELETE"])
@require_admin_api
def api_cockpit_client_del(cid):
    with get_conn() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM clients WHERE id=%s", (cid,))
        conn.commit()
    return jsonify({"ok": True})


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
